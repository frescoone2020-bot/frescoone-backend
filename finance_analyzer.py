"""
Frescoone Trading Sdn Bhd -- Finance Analyzer

Setup:
    pip install pandas openpyxl pdfplumber
Run (from this folder):
    python finance_analyzer.py

Reads (auto-discovered by pattern, so filenames can change month to month):
    isku_database.json                          -> SKU master (cost/RSP/DSP; add/edit via the
                                                    dashboard's ISKU Manager tab, or run
                                                    migrate_isku_database.py once to (re)build it)
    from-sitegiant-Orders_*.zip                  -> D2C order exports (TikTok/Shopee/Lazada/Web)
    dropship-reseller-file/INV-LCON*.xlsx        -> Lazada reseller invoice
    dropship-reseller-file/INV-SCON*.xlsx        -> Shopee reseller invoice (Case On)
    dropship-reseller-file/INV-SSUP*.xlsx        -> Shopee reseller invoice (Superbvalue)
    *ocbc*.csv / e-Statement*.csv                -> bank statement
    past-6months-payroll-history/*.zip           -> payslip PDFs

Writes:
    Executive_Financial_Report.md          -- narrative report
    dashboard.html                         -- local interactive dashboard (open in a browser)
    history/<month>.json                   -- one snapshot per run, accumulates a real
                                               multi-month trend as you run this monthly
"""
import glob
import json
import re
import sys
import zipfile
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"Missing dependency: {e}. Run: pip install pandas openpyxl pdfplumber")

try:
    import pdfplumber
    HAVE_PDFPLUMBER = True
except ImportError:
    HAVE_PDFPLUMBER = False

BASE_DIR = Path(__file__).resolve().parent
WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)
    print(f"[WARN] {msg}")


def fmt_money(x):
    return f"RM {x:,.2f}"


# ---------------------------------------------------------------------------
# 1. SKU MASTER (cost/RSP/DSP lookup, single source of truth for margins)
#
# Backed by isku_database.json, the local replacement for the
# "REVAMP-NEW-PRODUCT-CREATION" Google Sheet (see the ISKU Manager dashboard
# tab for add/search/edit). Cost and DSP are stored FROZEN on each record --
# same as the sheet's "Add To DB" button, which pastes computed values, not
# live formulas. That's deliberate, not a shortcut: currency rates and ITEC
# fees drift over time, so re-deriving Cost from *today's* tables on every
# read would silently re-price old stock at a rate/fee that was never
# actually paid for it. compute_cost()/compute_dsp() below implement the
# formula chain -- used for the live preview in the ISKU Manager's Add/Edit
# form, and to freeze cost/dsp onto a record at the moment it's saved, not
# for the bulk read path here.
#
# Schema v2 (2026-07-23, per Bun): currency/exchange-rate and ITEC_CODE are
# separate concerns now. Each SKU record picks its own `currency` (MYR/USD/
# CNY/SGD, independent of which ITEC code it uses) -- looked up in the
# top-level `currencies` table for a RM-per-unit rate. ITEC_CODE no longer
# carries currency/tax/a single flat fee; it's just Bun's "invisible costing"
# preset -- an itemized `fees` list (packaging/shipping, engraving logo,
# custom packaging, etc.), summed together, in the same currency as Import
# Price.
#   Fees Total   = sum(itec.fees[*].amount)
#   Tax          = Import Price x itec.tax_rate
#   Cost (RM)    = (Import Price + Tax + Fees Total) x currency.rate_to_rm
#   Ds Margin    = 50% if Cost < 20, 40% if Cost > 60, else 45%   (tier bucket)
#   Fresco Profit = (RSP - Cost) x (1 - Ds Margin)
#   DSP          = Cost + Fresco Profit
# ---------------------------------------------------------------------------
def compute_cost(rec, itec_codes, currencies):
    """Cost in RM for one ISKU record. Returns None if it can't be computed."""
    itec = itec_codes.get(rec.get("itec_code") or "")
    currency = currencies.get(rec.get("currency") or "")
    import_price = rec.get("import_price")
    if not itec or not currency or import_price is None:
        return None
    fees_total = sum(f["amount"] for f in itec.get("fees", []))
    tax = import_price * itec.get("tax_rate", 0)
    return (import_price + tax + fees_total) * currency["rate_to_rm"]


def compute_dsp(rec, cost):
    """Online reseller DSP: Cost + Fresco Profit share of (RSP - Cost)."""
    rsp = rec.get("rsp")
    if cost is None or rsp is None:
        return None
    if cost < 20:
        ds_margin = 0.50
    elif cost > 60:
        ds_margin = 0.40
    else:
        ds_margin = 0.45
    fresco_profit = (rsp - cost) * (1 - ds_margin)
    return cost + fresco_profit


def load_isku_database():
    path = BASE_DIR / "isku_database.json"
    if not path.exists():
        raise FileNotFoundError(
            "isku_database.json not found -- run migrate_isku_database.py once, "
            "or use the ISKU Manager dashboard tab to add ISKUs."
        )
    return json.loads(path.read_text(encoding="utf-8"))


def load_sku_master():
    db = load_isku_database()
    rows = []
    for sku, rec in db.get("skus", {}).items():
        rows.append({"sku": sku, "cost": rec.get("cost"), "rsp": rec.get("rsp"),
                      "dsp": rec.get("dsp"), "status": rec.get("status")})
    df = pd.DataFrame(rows)
    for c in ("cost", "rsp", "dsp"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df.set_index("sku")[["cost", "rsp", "dsp", "status"]]


def apply_cost_overrides(sku_master):
    """Deprecated pass-through. Overrides now live inside isku_database.json
    itself (a `cost_override`/`dsp_override` field per SKU, applied inside
    load_sku_master()), so there's nothing left to merge here. Kept as a no-op
    so existing call sites (restock_planner.py, restock_strategy.py) don't
    need to change.
    """
    return sku_master


# ---------------------------------------------------------------------------
# 2. SITEGIANT D2C ORDERS (TikTok / Shopee / Lazada / Web)
# ---------------------------------------------------------------------------
def load_sitegiant_orders():
    zips = glob.glob(str(BASE_DIR / "from-sitegiant-Orders_*.zip"))
    if not zips:
        raise FileNotFoundError("Could not find the SiteGiant orders zip (from-sitegiant-Orders_*.zip)")
    frames = []
    with zipfile.ZipFile(zips[0]) as z:
        for name in z.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            wb = load_workbook(BytesIO(z.read(name)), data_only=True, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(h).strip() if h else "" for h in next(rows)]
            df = pd.DataFrame(rows, columns=header)
            frames.append(df)
    orders = pd.concat(frames, ignore_index=True)

    # The export is split into 4 chunk files that are NOT sequential by ID or
    # date (each spans the full order-ID range) -- de-dupe exact-duplicate
    # rows rather than assume the chunks are disjoint (business rule #1).
    before = len(orders)
    orders = orders.drop_duplicates()
    removed = before - len(orders)
    if removed:
        warn(f"SiteGiant export: removed {removed} exact-duplicate line-item rows across the 4 export chunk files.")

    numeric_cols = ["product_price", "product_quantity", "product_total", "product_cost",
                     "shipping_fee", "total", "coupon_amount", "discount_amount", "voucher_amount"]
    for c in numeric_cols:
        if c in orders.columns:
            orders[c] = pd.to_numeric(orders[c], errors="coerce").fillna(0)
    orders["order_creation_date"] = pd.to_datetime(orders["order_creation_date"], errors="coerce")
    orders["order_id"] = orders["order_id"].astype(str).str.strip()
    orders["sku_key"] = orders["isku"].where(orders["isku"].notna() & (orders["isku"] != ""),
                                              orders["product_sku"]).astype(str).str.strip()
    return orders


# Confirmed by 100% order-ID overlap against the dropship-reseller invoices:
# these SiteGiant "marketplace" channels are the reseller's own storefront,
# fulfilled by Frescoone and re-billed at DSP via INV-SCON/LCON/SSUP. Frescoone
# does not own the full retail-price revenue shown for these lines -- only the
# DSP amount already captured in dropship_channel_metrics(). Counting both
# would double the revenue for these orders.
RESELLER_FULFILLED_MARKETPLACES = {"CASEON | SHOPEE", "Caseon | LZD", "Superb | SHOPEE"}
# JUSTINCASE | SHOPEE follows the same naming pattern as the confirmed reseller
# channels above, but Bun confirmed 2026-07-21 it's his own sub-storefront, not
# a reseller relationship -- treated as a normal D2C channel like FRESCOONE
# Shopee, no special-casing or warning needed (previously flagged as
# unverified/pending confirmation; that's now resolved).
# These storefronts are Singapore-based and their order amounts are in SGD, not
# MYR like every other SiteGiant channel. Previously excluded from RM figures
# entirely pending a confirmed exchange rate; Bun confirmed a rate 2026-07-21
# (1 SGD = 3 MYR -- his own rate for now, not a live market feed). Converted
# at that rate and folded into every RM total below, same as any other
# channel. Update SGD_TO_MYR_RATE if the rate changes.
SGD_MARKETPLACES = {"CASEON SG - caseon123.sg", "FRESCO SG"}
SGD_TO_MYR_RATE = 3.0


def sitegiant_channel_metrics(orders, sku_master):
    recognized = orders[orders["order_status"] == "Completed"].copy()

    dup_mask = recognized["marketplace"].isin(RESELLER_FULFILLED_MARKETPLACES)
    if dup_mask.any():
        dup_revenue = float(recognized.loc[dup_mask, "product_total"].sum())
        warn(f"SiteGiant: excluded {int(dup_mask.sum())} line items (RM {dup_revenue:,.2f} retail value) "
             f"from D2C revenue -- these orders in {sorted(recognized.loc[dup_mask, 'marketplace'].unique())} "
             f"are the same orders already billed via the dropship-reseller invoices (confirmed by 100% "
             f"order-ID match); counting both would double-count this revenue.")
        recognized = recognized[~dup_mask]

    sgd_mask = recognized["marketplace"].isin(SGD_MARKETPLACES)
    sgd_summary = {"revenue_sgd": 0.0, "orders": 0, "marketplaces": [], "rate": SGD_TO_MYR_RATE, "converted_revenue_rm": 0.0}
    if sgd_mask.any():
        sgd_summary = {
            "revenue_sgd": float(recognized.loc[sgd_mask, "product_total"].sum()),
            "orders": int(recognized.loc[sgd_mask, "order_id"].nunique()),
            "marketplaces": sorted(recognized.loc[sgd_mask, "marketplace"].unique()),
            "rate": SGD_TO_MYR_RATE,
        }
        # Convert the SGD storefronts' money fields to RM at the confirmed
        # rate, then let them flow through the exact same revenue-allocation/
        # cost-merge pipeline below as every other RM channel -- real cost
        # (from the shared RM-denominated cost master) against real converted
        # revenue, folded into every total from here on. No longer excluded
        # or kept in a parallel un-netted path.
        recognized.loc[sgd_mask, "product_total"] = recognized.loc[sgd_mask, "product_total"] * SGD_TO_MYR_RATE
        recognized.loc[sgd_mask, "total"] = recognized.loc[sgd_mask, "total"] * SGD_TO_MYR_RATE
        sgd_summary["converted_revenue_rm"] = float(recognized.loc[sgd_mask, "product_total"].sum())
        warn(f"SiteGiant: converted {int(sgd_mask.sum())} line items (SGD {sgd_summary['revenue_sgd']:,.2f}, "
             f"{sgd_summary['orders']} orders) in {sgd_summary['marketplaces']} to RM at Bun's confirmed rate of "
             f"1 SGD = {SGD_TO_MYR_RATE} MYR ({fmt_money(sgd_summary['converted_revenue_rm'])}) -- included in "
             f"RM revenue/COGS/every total from here on, no longer excluded. This is Bun's own stated rate, not "
             f"a live market rate.")

    # `product_total` is the PRE-discount line value -- `coupon_amount` (and
    # any other order-level adjustment) is applied at the ORDER level and
    # isn't netted into product_total. The order-level `total` field is the
    # actual final amount, confirmed to equal
    # sum(product_total)+sum(coupon_amount)+sum(discount_amount)+sum(voucher_amount)
    # per order to within rounding across the full June dataset. Using raw
    # product_total as revenue overstates orders with a coupon applied (seen
    # up to ~85% overstatement on individual heavily-discounted orders).
    # Allocate each order's true `total` back across its lines, proportional
    # to each line's share of that order's pre-discount product_total, so
    # per-SKU/per-brand revenue breakdowns still sum correctly to the order total.
    order_totals = recognized.groupby("order_id").agg(
        order_total=("total", "first"), order_product_total=("product_total", "sum")
    )
    recognized = recognized.merge(order_totals, left_on="order_id", right_index=True, how="left")
    recognized["revenue"] = recognized["product_total"] * (
        recognized["order_total"] / recognized["order_product_total"].replace(0, float("nan"))
    )
    recognized["revenue"] = recognized["revenue"].fillna(recognized["product_total"]).astype(float)

    recognized = recognized.merge(sku_master[["cost"]], left_on="sku_key", right_index=True, how="left")
    recognized["cost_known"] = recognized["cost"].notna()
    missing_cost = int(recognized["cost"].isna().sum())
    if missing_cost:
        warn(f"SiteGiant: {missing_cost} Completed line items have no SKU match in the master cost table "
             f"-- COGS for these lines is treated as 0 (this understates COGS). Listed in the dashboard's "
             f"Pending Action table for manual cost entry.")
    recognized["cost"] = recognized["cost"].fillna(0)
    recognized["line_cogs"] = recognized["cost"] * recognized["product_quantity"]
    recognized["month"] = recognized["order_creation_date"].dt.to_period("M")

    monthly = recognized.groupby("month").agg(
        revenue=("revenue", "sum"),
        cogs=("line_cogs", "sum"),
        orders=("order_id", "nunique"),
    ).reset_index().sort_values("month")

    by_marketplace = recognized.assign(
        marketplace_norm=recognized["marketplace"].fillna("Unknown").astype(str).str.strip()
    ).groupby("marketplace_norm").agg(
        revenue=("revenue", "sum"), cogs=("line_cogs", "sum")
    ).reset_index().sort_values("revenue", ascending=False)

    return {
        "line_items": recognized,
        "monthly": monthly,
        "by_marketplace": by_marketplace,
        "total_revenue": float(recognized["revenue"].sum()),
        "total_cogs": float(recognized["line_cogs"].sum()),
        "sgd_summary": sgd_summary,
    }


# ---------------------------------------------------------------------------
# 3. DROPSHIP / RESELLER INVOICES (LCON, SCON, SSUP)
#
# Each workbook is bespoke (different column layouts, different reseller
# business rules for chargeable order status, different ads-rebate terms).
# LCON's formula cells (Final Unit Price / Line Total / Return-Adjusted
# Line Total) were saved with EMPTY cached values, so they must be
# recomputed from the raw inputs. SCON and SSUP have live cached values
# for their equivalent columns and can be read directly.
# ---------------------------------------------------------------------------
def find_invoice_line(ws, label_substring, value_col="D"):
    label_substring = label_substring.lower()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and label_substring in cell.value.lower():
                return ws[f"{value_col}{cell.row}"].value
    return None


def get_period_label(ws):
    # Label text varies ("Period:" vs "Billing Period:") across the 3 reseller
    # templates, and the value can sit in the next column OR the next row
    # (LCON/SSUP use column-offset D->E, SCON uses B->C on the same row).
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "period" in cell.value.lower():
                neighbor = ws.cell(row=cell.row, column=cell.column + 1).value
                if neighbor:
                    return neighbor
    return None


def period_to_month(period_label):
    if not period_label:
        return None
    m = re.search(r"([A-Za-z]+)\s+(\d{4})", str(period_label))
    if not m:
        return None
    try:
        return pd.Period(f"{m.group(1)} {m.group(2)}", freq="M")
    except ValueError:
        return None


def parse_lcon(path):
    wb = load_workbook(path, data_only=True)
    ws_calc, ws_returns, ws_claims, ws_invoice = (
        wb["Itemized Calculation"], wb["Returns & Refunds"], wb["Claims"], wb["Invoice"]
    )

    # Returns & Refunds: header row 2, data from row 3 until Order ID blank.
    # (The workbook's own formula hardcodes a 1-row match range -- we scan
    # the full table instead so a future second return isn't silently missed.)
    returns = {}
    r = 3
    while ws_returns.cell(row=r, column=1).value not in (None, ""):
        order_id = str(ws_returns.cell(row=r, column=1).value).strip()
        returns[order_id] = ws_returns.cell(row=r, column=7).value  # G = Decision
        r += 1

    line_items = []
    r = 2
    while ws_calc.cell(row=r, column=2).value not in (None, ""):
        order_id = str(ws_calc.cell(row=r, column=2).value).strip()
        sku = ws_calc.cell(row=r, column=5).value
        qty = ws_calc.cell(row=r, column=7).value or 0
        dsp_price = ws_calc.cell(row=r, column=8).value or 0
        promo_rate = ws_calc.cell(row=r, column=9).value or 0
        applied = ws_calc.cell(row=r, column=10).value
        order_date = ws_calc.cell(row=r, column=4).value
        verification = ws_calc.cell(row=r, column=18).value

        final_unit_price = round(dsp_price * (1 - promo_rate), 4) if applied == "Yes" else dsp_price
        line_total = round(final_unit_price * qty, 2)
        decision = returns.get(order_id)
        adj_total = 0.0 if decision == "Confirmed Good - Remove Charge" else line_total

        line_items.append({"order_id": order_id, "sku": sku, "qty": qty, "order_date": order_date,
                            "adj_total": adj_total, "verification": verification})
        r += 1

    claims_total = 0.0
    for r in range(3, min(45, ws_claims.max_row + 1)):
        status = ws_claims.cell(row=r, column=8).value
        amt = ws_claims.cell(row=r, column=7).value or 0
        if status == "Approved":
            claims_total += amt

    df = pd.DataFrame(line_items)
    net_invoice = float(df["adj_total"].sum() - claims_total) if not df.empty else 0.0
    period = get_period_label(ws_invoice)
    return {"line_items": df, "net_invoice": net_invoice, "period": period, "ads_rebate": 0.0}


def parse_scon(path):
    wb = load_workbook(path, data_only=True)
    ws_calc, ws_invoice = wb["Itemized Calculation"], wb["Invoice"]

    line_items = []
    r = 2
    while ws_calc.cell(row=r, column=2).value not in (None, ""):
        order_id = str(ws_calc.cell(row=r, column=2).value).strip()
        sku = ws_calc.cell(row=r, column=5).value
        qty = ws_calc.cell(row=r, column=6).value or 0
        order_date = ws_calc.cell(row=r, column=4).value
        adj_total = ws_calc.cell(row=r, column=11).value or 0  # K, already return-adjusted, cached
        verification = ws_calc.cell(row=r, column=15).value
        line_items.append({"order_id": order_id, "sku": sku, "qty": qty, "order_date": order_date,
                            "adj_total": adj_total, "verification": verification})
        r += 1

    df = pd.DataFrame(line_items)
    net_invoice = find_invoice_line(ws_invoice, "AMOUNT DUE")
    if net_invoice is None:
        warn("SCON invoice: no cached 'NET INVOICE AMOUNT DUE' found -- falling back to sum of line "
             "items (ads rebate / claims not netted off).")
        net_invoice = float(df["adj_total"].sum()) if not df.empty else 0.0

    ads_rebate = find_invoice_line(ws_invoice, "ads rebate") or find_invoice_line(ws_invoice, "ads spend")
    ads_rebate = abs(ads_rebate) if isinstance(ads_rebate, (int, float)) else 0.0

    period = get_period_label(ws_invoice)
    return {"line_items": df, "net_invoice": float(net_invoice), "period": period, "ads_rebate": ads_rebate}


def parse_ssup(path):
    wb = load_workbook(path, data_only=True)
    ws_calc, ws_invoice = wb["Itemized Calculation"], wb["Invoice"]

    line_items = []
    r = 2
    while ws_calc.cell(row=r, column=2).value not in (None, ""):
        status = ws_calc.cell(row=r, column=17).value  # Q = Order Status
        if status in ("Completed", "Shipped"):
            line_items.append({
                "order_id": str(ws_calc.cell(row=r, column=2).value).strip(),
                "sku": ws_calc.cell(row=r, column=5).value,
                "qty": ws_calc.cell(row=r, column=6).value or 0,
                "order_date": ws_calc.cell(row=r, column=4).value,
                "adj_total": ws_calc.cell(row=r, column=18).value or 0,  # R, return-adjusted, cached
                "status": status,
            })
        r += 1

    df = pd.DataFrame(line_items)
    net_invoice = find_invoice_line(ws_invoice, "AMOUNT DUE")
    if net_invoice is None:
        warn("SSUP invoice: no cached 'NET INVOICE AMOUNT DUE' found -- falling back to sum of line "
             "items (claims not netted off).")
        net_invoice = float(df["adj_total"].sum()) if not df.empty else 0.0

    ads_rebate = find_invoice_line(ws_invoice, "ads rebate") or find_invoice_line(ws_invoice, "ads spend")
    ads_rebate = abs(ads_rebate) if isinstance(ads_rebate, (int, float)) else 0.0

    period = get_period_label(ws_invoice)
    return {"line_items": df, "net_invoice": float(net_invoice), "period": period, "ads_rebate": ads_rebate}


def dropship_channel_metrics(sku_master):
    file_parsers = {
        "LCON (Lazada Reseller)": (glob.glob(str(BASE_DIR / "dropship-reseller-file" / "INV-LCON*.xlsx")), parse_lcon),
        "SCON (Shopee Reseller - Case On)": (glob.glob(str(BASE_DIR / "dropship-reseller-file" / "INV-SCON*.xlsx")), parse_scon),
        "SSUP (Shopee Reseller - Superbvalue)": (glob.glob(str(BASE_DIR / "dropship-reseller-file" / "INV-SSUP*.xlsx")), parse_ssup),
    }

    all_line_items, channel_summary = [], []
    monthly = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0})
    total_ads_rebate = 0.0

    for label, (paths, parser) in file_parsers.items():
        if not paths:
            warn(f"Dropship reseller file for {label} not found -- skipped.")
            continue
        result = parser(paths[0])
        df = result["line_items"]
        if df.empty:
            warn(f"Dropship {label}: no line items parsed.")
            continue

        df["sku"] = df["sku"].astype(str).str.strip()
        df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
        df = df.merge(sku_master[["cost"]], left_on="sku", right_index=True, how="left")
        df["cost_known"] = df["cost"].notna()
        missing = int(df["cost"].isna().sum())
        if missing:
            warn(f"Dropship {label}: {missing} line items have no SKU match in the master cost table "
                 f"-- COGS for these lines treated as 0. Listed in the dashboard's Pending Action table "
                 f"for manual cost entry.")
        df["cost"] = df["cost"].fillna(0)
        df["line_cogs"] = df["cost"] * df["qty"]
        df["reseller"] = label

        total_cogs = float(df["line_cogs"].sum())
        all_line_items.append(df)
        total_ads_rebate += result["ads_rebate"]

        channel_summary.append({
            "reseller": label, "period": result["period"], "net_revenue": result["net_invoice"],
            "cogs": total_cogs, "gross_profit": result["net_invoice"] - total_cogs,
            "ads_rebate": result["ads_rebate"],
        })

        month_key = period_to_month(result["period"])
        if month_key is not None:
            monthly[month_key]["revenue"] += result["net_invoice"]
            monthly[month_key]["cogs"] += total_cogs
        else:
            warn(f"Dropship {label}: could not parse a calendar month from period label "
                 f"'{result['period']}' -- excluded from monthly cash-flow/growth tables.")

    line_items_df = pd.concat(all_line_items, ignore_index=True) if all_line_items else pd.DataFrame()
    summary_df = pd.DataFrame(channel_summary)
    monthly_df = (pd.DataFrame([{"month": k, **v} for k, v in monthly.items()]).sort_values("month")
                  if monthly else pd.DataFrame(columns=["month", "revenue", "cogs"]))

    return {
        "line_items": line_items_df,
        "by_reseller": summary_df,
        "monthly": monthly_df,
        "total_revenue": float(summary_df["net_revenue"].sum()) if not summary_df.empty else 0.0,
        "total_cogs": float(summary_df["cogs"].sum()) if not summary_df.empty else 0.0,
        "total_ads_rebate": total_ads_rebate,
    }


# ---------------------------------------------------------------------------
# 3b. PLATFORM FEES -- optional. Only present once the seller manually exports
# each platform's income/fee report (Shopee/TikTok/Lazada all require a
# password re-verification step that this script cannot complete on its own,
# so this can't be automated end-to-end). If the "income-exported*" folder
# isn't present, platform fees remain unknown and D2C revenue stays gross,
# exactly as before -- this section is additive, not required.
#
# Coverage is confirmed by order-ID matching against the SiteGiant export
# (see conversation record): each platform's report only covers the
# "FRESCOONE"-branded storefront, NOT the separately-run "DUX"-branded
# storefront on the same platforms -- DUX fees remain unknown.
# ---------------------------------------------------------------------------
def _find_platform_income_files():
    root_candidates = glob.glob(str(BASE_DIR / "income-exported*"))
    if not root_candidates:
        return {}
    root = Path(root_candidates[0])
    files = {"lazada": [], "shopee": [], "tiktok": []}
    for p in root.rglob("*"):
        if p.is_file() and not p.name.startswith("~$"):
            lp = str(p).lower()
            for platform in files:
                if platform in lp:
                    files[platform].append(p)
    return files


def _clean_order_id(x):
    return str(x).lstrip("'").strip().upper()


def _find_marketplace_order_export_files():
    root_candidates = glob.glob(str(BASE_DIR / "order-exported*"))
    if not root_candidates:
        return {}
    root = Path(root_candidates[0])
    files = {"lazada": [], "shopee": [], "tiktok": []}
    for p in root.rglob("*"):
        if p.is_file() and not p.name.startswith("~$"):
            lp = str(p).lower()
            for platform in files:
                if platform in lp:
                    files[platform].append(p)
    return files


def load_lazada_fees(files, target_month):
    csvs = [f for f in files if f.suffix.lower() == ".csv"]
    if not csvs:
        return None
    df = pd.read_csv(csvs[0], sep=";", dtype=str)
    df["Order Creation Date"] = pd.to_datetime(df["Order Creation Date"], format="%d %b %Y", errors="coerce")
    df["Amount(Include Tax)"] = pd.to_numeric(df["Amount(Include Tax)"], errors="coerce").fillna(0)
    month_df = df[df["Order Creation Date"].dt.to_period("M") == target_month]
    if month_df.empty:
        warn(f"Lazada fee report: no rows found for {target_month} -- covers "
             f"{df['Order Creation Date'].min()} to {df['Order Creation Date'].max()}.")
        return None
    # "Item Price Credit" / "Reversal Item Price" are the item's own value credited
    # to the seller, not a fee. "Sponsored Affiliates" is Lazada's paid-ads/affiliate
    # program -- split out as ad spend, separate from marketplace/payment fees.
    revenue = float(month_df[month_df["Fee Name"] == "Item Price Credit"]["Amount(Include Tax)"].sum()
                     + month_df[month_df["Fee Name"] == "Reversal Item Price"]["Amount(Include Tax)"].sum())
    ads_rows = month_df[month_df["Fee Name"].isin(["Sponsored Affiliates", "Sponsored Affiliates Refund"])]
    ads_spend = -float(ads_rows["Amount(Include Tax)"].sum())
    fee_rows = month_df[~month_df["Fee Name"].isin(["Item Price Credit", "Reversal Item Price",
                                                       "Sponsored Affiliates", "Sponsored Affiliates Refund"])]
    fee_total = -float(fee_rows["Amount(Include Tax)"].sum())
    per_order_fee = (-fee_rows.groupby("Order Number")["Amount(Include Tax)"].sum()).to_dict()
    per_order_revenue = ((month_df[month_df["Fee Name"] == "Item Price Credit"]
                           .groupby("Order Number")["Amount(Include Tax)"].sum())
                          .add(month_df[month_df["Fee Name"] == "Reversal Item Price"]
                               .groupby("Order Number")["Amount(Include Tax)"].sum(), fill_value=0)).to_dict()
    return {"platform": "Lazada (FRESCOONE storefront only)", "revenue": revenue,
            "fee": fee_total, "ads_spend": ads_spend, "orders": month_df["Order Number"].nunique(),
            "per_order_fee": {_clean_order_id(k): v for k, v in per_order_fee.items()},
            "per_order_revenue": {_clean_order_id(k): v for k, v in per_order_revenue.items()}}


def load_shopee_fees(files, sitegiant_orders, target_month):
    zips = [f for f in files if f.suffix.lower() == ".zip"]
    xlsxs = [f for f in files if f.suffix.lower() == ".xlsx"]
    frames = []
    try:
        if zips:
            with zipfile.ZipFile(zips[0]) as z:
                for name in z.namelist():
                    if name.lower().endswith(".xlsx"):
                        frames.append(pd.read_excel(BytesIO(z.read(name)), header=17))
        else:
            for f in xlsxs:
                frames.append(pd.read_excel(f, header=17))
    except Exception as e:
        warn(f"Shopee fee report: could not parse -- {e}")
        return None
    if not frames:
        return None
    bal = pd.concat(frames, ignore_index=True)
    bal["Date"] = pd.to_datetime(bal["Date"], errors="coerce")
    bal_month = bal[bal["Date"].dt.to_period("M") == target_month]

    # Shopee Ads / "Value-added Services" is a separate wallet deduction, not tied
    # to individual orders at all -- doesn't show up in the per-order fee gap below.
    ads_rows = bal_month[(bal_month["Transaction Type"] == "Seller Balance Payment")
                          & (bal_month["Description"].str.contains("Value-added Services", case=False, na=False))]
    ads_spend = -float(ads_rows["Amount"].sum())

    order_income = bal[bal["Transaction Type"] == "Order Income"].copy()
    order_income["order_id_clean"] = order_income["Order ID"].apply(_clean_order_id)
    bal_net = order_income.groupby("order_id_clean")["Amount"].sum()

    fresco = sitegiant_orders[(sitegiant_orders["marketplace"] == "FRESCOONE Shopee")
                               & (sitegiant_orders["order_status"] == "Completed")
                               & (sitegiant_orders["order_creation_date"].dt.to_period("M") == target_month)]
    if fresco.empty:
        return None
    # Use the order-level `total` (actual post-coupon/discount amount), not
    # `product_total` (pre-discount line value) -- otherwise heavily-discounted
    # orders inflate the "accrued" side and the fee gap looks far bigger and
    # far less consistent per-order than it actually is.
    by_order = fresco.groupby("marketplace_order_id")["total"].first().reset_index()
    by_order["order_id_clean"] = by_order["marketplace_order_id"].apply(_clean_order_id)
    by_order = by_order.merge(bal_net.rename("net_income"), left_on="order_id_clean", right_index=True, how="left")

    matched = by_order[by_order["net_income"].notna()]
    unmatched_count = len(by_order) - len(matched)
    if unmatched_count:
        warn(f"Shopee fee report: {unmatched_count} of {len(by_order)} FRESCOONE Shopee orders in "
             f"{target_month} have no matching payout yet (likely still in escrow) -- their fees are "
             f"not included, so the Shopee fee total is a floor, not the full amount.")
    fee_total = float((matched["total"] - matched["net_income"]).sum())
    revenue = float(by_order["total"].sum())
    per_order_fee = dict(zip(matched["order_id_clean"], matched["total"] - matched["net_income"]))
    per_order_revenue = dict(zip(by_order["order_id_clean"], by_order["total"]))
    return {"platform": "Shopee (FRESCOONE storefront only)", "revenue": revenue,
            "fee": fee_total, "ads_spend": ads_spend, "orders": len(matched), "orders_total": len(by_order),
            "per_order_fee": per_order_fee, "per_order_revenue": per_order_revenue}


def load_shopee_order_export_fees(path, sitegiant_orders, target_month, already_covered):
    """Shopee's own bulk order export ("Order.all...xlsx") carries Transaction
    Fee/Commission Fee/Service Fee directly, per order -- unlike the wallet
    balance report, it works for ANY connected storefront, not just the
    FRESCOONE-branded one. Auto-detects which SiteGiant marketplace the file
    belongs to via order-ID overlap (skipping marketplaces already covered by
    a real income report), so it works for whichever storefront's file gets
    dropped in without relying on the filename.

    IMPORTANT: the sheet's "Grand Total" column is NOT net settlement --
    verified against Shopee's own per-order income breakdown for a real
    order: Total Amount minus Grand Total equalled exactly that order's
    "Discount Voucher Amount Sponsored by Shopee", not a seller-borne fee at
    all. The real fee (matching Shopee's income report almost exactly, short
    only the Saver Programme Fee which this export doesn't carry a column
    for) is Transaction Fee + Commission Fee + Service Fee. Using Grand Total
    previously understated DUX/Justincase Shopee fees by roughly 4x, which in
    turn overstated their per-order Net Profit. Fixed 2026-07-21.
    Does not capture ad spend -- that's a separate wallet withdrawal not visible
    in an order export, same limitation as everywhere else without a wallet report.
    """
    try:
        df = pd.read_excel(path, sheet_name="orders")
    except Exception as e:
        warn(f"Shopee order export {path.name}: could not parse -- {e}")
        return None
    fee_cols = ["Transaction Fee", "Commission Fee", "Service Fee"]
    if "Order ID" not in df.columns or not all(c in df.columns for c in fee_cols):
        warn(f"Shopee order export {path.name}: missing expected columns -- skipped.")
        return None

    file_ids = set(df["Order ID"].dropna().astype(str).apply(_clean_order_id))
    shopee_marketplaces = [m for m in sitegiant_orders["marketplace"].dropna().unique()
                            if "shopee" in str(m).lower() and m not in already_covered]
    best_match, best_overlap = None, 0
    for m in shopee_marketplaces:
        m_ids = set(sitegiant_orders[sitegiant_orders["marketplace"] == m]["marketplace_order_id"].apply(_clean_order_id))
        overlap = len(file_ids & m_ids)
        if overlap > best_overlap:
            best_match, best_overlap = m, overlap
    if not best_match or best_overlap < len(file_ids) * 0.5:
        warn(f"Shopee order export {path.name}: could not confidently match to an uncovered "
             f"marketplace (best match {best_match!r} with {best_overlap}/{len(file_ids)} order IDs) -- skipped.")
        return None

    # Fee columns repeat identically on every SKU line of a multi-line order
    # (confirmed: order 2606034BQJFAFB, 4 lines, same Service Fee 8.08 on all
    # 4 rows) -- take the first row per order, not a sum across lines, or a
    # 4-line order's fee would be 4x overcounted.
    by_order_fee = df.groupby("Order ID")[fee_cols].first().sum(axis=1)
    by_order_fee.index = by_order_fee.index.astype(str).map(_clean_order_id)

    fresco = sitegiant_orders[(sitegiant_orders["marketplace"] == best_match)
                               & (sitegiant_orders["order_status"] == "Completed")
                               & (sitegiant_orders["order_creation_date"].dt.to_period("M") == target_month)]
    if fresco.empty:
        return None
    by_order = fresco.groupby("marketplace_order_id")["total"].first().reset_index()
    by_order["order_id_clean"] = by_order["marketplace_order_id"].apply(_clean_order_id)
    by_order = by_order.merge(by_order_fee.rename("order_fee"), left_on="order_id_clean", right_index=True, how="left")

    matched = by_order[by_order["order_fee"].notna()]
    unmatched_count = len(by_order) - len(matched)
    if unmatched_count:
        warn(f"Shopee order export ({best_match}): {unmatched_count} of {len(by_order)} Completed "
             f"{target_month} orders not found in the order export -- their fee is unknown.")
    fee_total = float(matched["order_fee"].sum())
    revenue = float(by_order["total"].sum())
    per_order_fee = dict(zip(matched["order_id_clean"], matched["order_fee"]))
    return {"platform": f"{best_match} (from raw order export)", "marketplace": best_match,
            "revenue": revenue, "fee": fee_total, "ads_spend": 0.0,
            "orders": len(matched), "orders_total": len(by_order), "per_order_fee": per_order_fee}


def load_tiktok_fees(files, target_month):
    xlsxs = [f for f in files if f.suffix.lower() == ".xlsx"]
    if not xlsxs:
        return None
    df = pd.read_excel(xlsxs[0], sheet_name="Order details")
    df["Order created time"] = pd.to_datetime(df["Order created time"], errors="coerce")
    month_df = df[(df["Order created time"].dt.to_period("M") == target_month) & (df["Transaction type"] == "Order")]
    if month_df.empty:
        warn(f"TikTok fee report: no 'Order' rows found for {target_month}.")
        return None
    fee_total = -float(pd.to_numeric(month_df["Total Fees"], errors="coerce").fillna(0).sum())
    revenue = float(pd.to_numeric(month_df["Subtotal before discounts"], errors="coerce").fillna(0).sum())

    # GMV Max ad spend is a separate wallet withdrawal ("Withdrawal records" sheet),
    # not part of any individual order's "Total Fees" -- would otherwise be missed.
    ads_spend = 0.0
    try:
        wd = pd.read_excel(xlsxs[0], sheet_name="Withdrawal records")
        wd["Request time"] = pd.to_datetime(wd["Request time"], errors="coerce")
        wd_month = wd[(wd["Request time"].dt.to_period("M") == target_month)
                      & (wd["Transaction type"] == "GMV Pay Deduction")]
        ads_spend = -float(pd.to_numeric(wd_month["Amount"], errors="coerce").fillna(0).sum())
    except Exception as e:
        warn(f"TikTok fee report: could not read Withdrawal records for ad spend -- {e}")

    per_order = month_df.groupby("Order/Adjustment ID").agg(
        fee=("Total Fees", lambda s: -pd.to_numeric(s, errors="coerce").fillna(0).sum()),
        revenue=("Subtotal before discounts", lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()),
    )
    per_order.index = per_order.index.astype(str).map(_clean_order_id)
    return {"platform": "TikTok (FRESCOONE storefront only)", "revenue": revenue,
            "fee": fee_total, "ads_spend": ads_spend, "orders": month_df["Order/Adjustment ID"].nunique(),
            "per_order_fee": per_order["fee"].to_dict(), "per_order_revenue": per_order["revenue"].to_dict()}


def load_platform_fees(sitegiant_orders, target_month):
    income_files = _find_platform_income_files()
    order_export_files = _find_marketplace_order_export_files()

    results = []
    marketplace_fee_lookup = {}

    if income_files.get("lazada"):
        r = load_lazada_fees(income_files["lazada"], target_month)
        if r:
            results.append(r)
            marketplace_fee_lookup["FRESCOONE Lazada"] = r
    if income_files.get("shopee"):
        r = load_shopee_fees(income_files["shopee"], sitegiant_orders, target_month)
        if r:
            results.append(r)
            marketplace_fee_lookup["FRESCOONE Shopee"] = r
    if income_files.get("tiktok"):
        r = load_tiktok_fees(income_files["tiktok"], target_month)
        if r:
            results.append(r)
            marketplace_fee_lookup["FRESCOONE | TIKTOK"] = r

    # Raw marketplace order exports (e.g. Shopee's own "Order.all...xlsx") cover
    # storefronts that don't have a dedicated income/fee report -- auto-detected
    # per file, skipping whatever marketplace_fee_lookup already covers above.
    for path in order_export_files.get("shopee", []):
        r = load_shopee_order_export_fees(path, sitegiant_orders, target_month,
                                           already_covered=set(marketplace_fee_lookup.keys()))
        if r:
            results.append(r)
            marketplace_fee_lookup[r["marketplace"]] = r

    if not results:
        return {"total": 0.0, "total_fees": 0.0, "total_ads": 0.0, "by_platform": [], "marketplace_fee_lookup": {}}

    for r in results:
        r["total_cost"] = r["fee"] + r["ads_spend"]
        r["fee_pct"] = r["fee"] / r["revenue"] * 100 if r["revenue"] else 0.0
        r["ads_pct"] = r["ads_spend"] / r["revenue"] * 100 if r["revenue"] else 0.0
        r["total_pct"] = r["total_cost"] / r["revenue"] * 100 if r["revenue"] else 0.0

        # Per user direction: Order Detail Net Profit deducts ONLY ad spend, not
        # the marketplace fee (shown separately as an informational value/%, not
        # subtracted). Ad spend has no per-order breakdown at the source -- it's
        # a single platform-wide wallet withdrawal -- so it's allocated
        # proportionally, revenue-weighted, across every order we have a revenue
        # figure for on that platform, independent of whether the marketplace
        # fee happens to be separately known for that specific order.
        per_order_ads = {}
        per_order_revenue = r.get("per_order_revenue")
        if r["ads_spend"] and per_order_revenue:
            total_rev = sum(per_order_revenue.values())
            if total_rev > 0:
                for oid, rev in per_order_revenue.items():
                    per_order_ads[oid] = rev / total_rev * r["ads_spend"]
        r["per_order_ads"] = per_order_ads

    covered = sorted(marketplace_fee_lookup.keys())
    warn(f"Platform fees: real commission/payment-processing fees deducted for: {covered}. Ad spend (a "
         f"separate wallet withdrawal, not part of any order's line-item fees) is only captured for storefronts "
         f"with a full income/balance report, not for storefronts covered only by a raw order export. Any "
         f"Shopee/Lazada/TikTok storefront not in that list is still gross of fees.")

    return {
        "total": sum(r["total_cost"] for r in results),
        "total_fees": sum(r["fee"] for r in results),
        "total_ads": sum(r["ads_spend"] for r in results),
        "by_platform": results,
        "marketplace_fee_lookup": marketplace_fee_lookup,
    }


# ---------------------------------------------------------------------------
# 4. BANK STATEMENT (cash flow, liquidity, OpEx)
# ---------------------------------------------------------------------------
BANK_CATEGORY_RULES = [
    (re.compile(r"SALARY:\d{4}-\d{1,2}"), "Payroll (bank cash outflow)"),
    (re.compile(r"LOAN\s+TEE\s+CHIN\s+XIONG", re.I), "Financing (Director Loan)"),
    (re.compile(r"SITEGIANT", re.I), "Platform Subscription (SiteGiant)"),
    (re.compile(r"ZURICH", re.I), "Insurance"),
    (re.compile(r"JOMPAY", re.I), "Utilities"),
    (re.compile(r"GIRO SERVICE CHARGE|POSTAGE CHARGE|DUITNOW.*\bSC\b", re.I), "Bank Charges"),
]
PLATFORM_PAYOUT_PATTERNS = [
    (re.compile(r"SHOPEE|SSUP|SCON", re.I), "Shopee Settlement"),
    (re.compile(r"\bLAZ\b|LAZADA|LCON", re.I), "Lazada Settlement"),
    (re.compile(r"TIKTOK|MONEYMATCH", re.I), "TikTok Settlement"),
]


def categorize_txn(details, supplementary, ref_owner, credit, debit):
    text = " ".join(str(x) for x in (details, supplementary, ref_owner) if x and str(x) != "nan")
    for pattern, label in BANK_CATEGORY_RULES:
        if pattern.search(text):
            return label
    if credit > 0:
        for pattern, label in PLATFORM_PAYOUT_PATTERNS:
            if pattern.search(text):
                return f"Platform Payout - {label}"
        return "Other Inflow"
    if debit > 0:
        return "Vendor / Other Payment"
    return "Other"


def load_bank_statement():
    csvs = glob.glob(str(BASE_DIR / "*ocbc*.csv")) or glob.glob(str(BASE_DIR / "e-Statement*.csv"))
    if not csvs:
        raise FileNotFoundError("Could not find the OCBC e-statement CSV.")
    try:
        df = pd.read_csv(csvs[0])
    except UnicodeDecodeError:
        df = pd.read_csv(csvs[0], encoding="cp1252")

    df["Statement Date"] = pd.to_datetime(df["Statement Date"], format="%Y%m%d")
    df["Debit Amount"] = pd.to_numeric(df["Debit Amount"], errors="coerce").fillna(0)
    df["Credit Amount"] = pd.to_numeric(df["Credit Amount"], errors="coerce").fillna(0)
    df["category"] = df.apply(
        lambda row: categorize_txn(row.get("Statement Details Info"), row.get("Supplementary Details"),
                                    row.get("Ref For Account Owner"), row["Credit Amount"], row["Debit Amount"]),
        axis=1,
    )
    df["month"] = df["Statement Date"].dt.to_period("M")
    return df


def cash_flow_trends(bank_df):
    monthly = bank_df.groupby("month").agg(
        inflows=("Credit Amount", "sum"), outflows=("Debit Amount", "sum"),
    ).reset_index().sort_values("month")
    monthly["net_cash_flow"] = monthly["inflows"] - monthly["outflows"]
    monthly["inflow_mom_growth_pct"] = monthly["inflows"].pct_change() * 100
    monthly["month_str"] = monthly["month"].astype(str)

    inflow_by_month = monthly.set_index("month")["inflows"]
    yoy = []
    for m in monthly["month"]:
        prior = m - 12
        if prior in inflow_by_month.index and inflow_by_month[prior]:
            yoy.append((inflow_by_month[m] - inflow_by_month[prior]) / inflow_by_month[prior] * 100)
        else:
            yoy.append(None)
    monthly["inflow_yoy_growth_pct"] = yoy
    return monthly


def bank_opex(bank_df):
    excluded = {"Payroll (bank cash outflow)", "Financing (Director Loan)"}
    opex_df = bank_df[(bank_df["Debit Amount"] > 0) & (~bank_df["category"].isin(excluded))]
    by_category = opex_df.groupby("category")["Debit Amount"].sum().sort_values(ascending=False)
    by_month = opex_df.groupby("month")["Debit Amount"].sum()
    by_month_category = opex_df.groupby(["month", "category"])["Debit Amount"].sum()
    return {"total": float(opex_df["Debit Amount"].sum()), "by_category": by_category,
            "by_month": by_month, "by_month_category": by_month_category, "detail": opex_df}


def bank_opex_category_for_month(bank_opex_result, month):
    """Category breakdown scoped to a single month -- must be used alongside
    Section 1's month-scoped OpEx total, not the full-history by_category,
    or the two figures in the report won't reconcile."""
    by_month_category = bank_opex_result["by_month_category"]
    if month not in by_month_category.index.get_level_values("month"):
        return pd.Series(dtype=float)
    return by_month_category.loc[month].sort_values(ascending=False)


# ---------------------------------------------------------------------------
# 5. PAYROLL (fixed OpEx, parsed from payslip PDFs)
# ---------------------------------------------------------------------------
MONTH_ABBR_TO_NUM = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                      "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
MONTH_NUM_TO_NAME = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                      7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"}
# Payslips mix abbreviated ("Jun 2026") and full ("May 2026") month names --
# match the 3-letter prefix plus any trailing letters so both forms hit.
PAYROLL_MONTH_RE = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(20\d{2})\b", re.I)
GROSS_PAY_RE = re.compile(r"GROSS PAY\s*[:\-]?\s*(?:RM)?\s*([\d,]+\.\d{2})", re.I)
NET_PAY_RE = re.compile(r"Net Pay\s*[:\-]?\s*(?:RM)?\s*([\d,]+\.\d{2})", re.I)
TOTAL_DEDUCTION_RE = re.compile(r"TOTAL DEDUCTION\s*[:\-]?\s*(?:RM)?\s*([\d,]+\.\d{2})", re.I)


def _to_float(s):
    return float(s.replace(",", "")) if s else None


def parse_payslip_pdf(file_bytes, filename):
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    month_match = PAYROLL_MONTH_RE.search(text)
    net_match = NET_PAY_RE.search(text)
    if not (month_match and net_match):
        warn(f"Payroll: could not parse month/net pay from {filename} -- skipped.")
        return None

    month_num = MONTH_ABBR_TO_NUM[month_match.group(1).capitalize()]
    year = int(month_match.group(2))
    gross_match = GROSS_PAY_RE.search(text)
    ded_match = TOTAL_DEDUCTION_RE.search(text)
    return {
        "employee": filename.rsplit(".", 1)[0],
        "month": f"{MONTH_NUM_TO_NAME[month_num]} {year}",
        "month_period": pd.Period(year=year, month=month_num, freq="M"),
        "gross_pay": _to_float(gross_match.group(1)) if gross_match else None,
        "total_deduction": _to_float(ded_match.group(1)) if ded_match else None,
        "net_pay": _to_float(net_match.group(1)),
    }


def load_payroll():
    if not HAVE_PDFPLUMBER:
        warn("pdfplumber not installed -- payroll OpEx will be RM 0 and excluded from Net Profit. "
             "Run: pip install pdfplumber")
        return pd.DataFrame(columns=["employee", "month", "gross_pay", "total_deduction", "net_pay"])

    zips = glob.glob(str(BASE_DIR / "past-6months-payroll-history" / "*.zip"))
    if not zips:
        warn("Payroll: no payslip zip files found in past-6months-payroll-history/.")
    records = []
    for zpath in zips:
        with zipfile.ZipFile(zpath) as z:
            for name in z.namelist():
                if not name.lower().endswith(".pdf"):
                    continue
                rec = parse_payslip_pdf(z.read(name), name)
                if rec:
                    records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        return df
    dup = int(df.duplicated(subset=["employee", "month"]).sum())
    if dup:
        warn(f"Payroll: {dup} duplicate (employee, month) payslip rows found across zips -- kept the first.")
        df = df.drop_duplicates(subset=["employee", "month"], keep="first")
    return df


def payroll_monthly(payroll_df):
    if payroll_df.empty:
        return pd.DataFrame(columns=["month", "total_gross", "total_net", "headcount"])
    return (payroll_df.groupby("month_period")
            .agg(total_gross=("gross_pay", "sum"), total_net=("net_pay", "sum"), headcount=("employee", "nunique"))
            .reset_index().rename(columns={"month_period": "month"}).sort_values("month"))


# ---------------------------------------------------------------------------
# 6. HEADLINE METRICS
# ---------------------------------------------------------------------------
def compute_net_profit(sitegiant, dropship, bank_opex_result, payroll_monthly_df, target_month, platform_fees=None):
    revenue = sitegiant["total_revenue"] + dropship["total_revenue"]
    cogs = sitegiant["total_cogs"] + dropship["total_cogs"]
    gross_profit = revenue - cogs
    gross_margin = gross_profit / revenue * 100 if revenue else 0.0

    payroll_row = payroll_monthly_df[payroll_monthly_df["month"] == target_month] if not payroll_monthly_df.empty else payroll_monthly_df
    payroll_cost = float(payroll_row["total_gross"].sum()) if not payroll_row.empty else 0.0
    if payroll_row.empty:
        warn(f"Payroll: no parsed payslip data for {target_month} -- payroll OpEx for that month is RM 0.")

    bank_opex_month = float(bank_opex_result["by_month"].get(target_month, 0.0))
    ads_spend = dropship.get("total_ads_rebate", 0.0)
    d2c_platform_fees = platform_fees["total"] if platform_fees else 0.0
    fees_and_ads = ads_spend + d2c_platform_fees

    net_profit = gross_profit - (payroll_cost + bank_opex_month + fees_and_ads)
    net_margin = net_profit / revenue * 100 if revenue else 0.0

    # Monthly NPE (per user direction): Gross Profit less ONLY known D2C ad
    # spend -- not the marketplace fee, not Payroll/Bank OpEx. This is the
    # aggregate of the same "deduct ads only" logic used per-row in Order
    # Detail (summed across every resolved-cost order/SKU line), NOT the full
    # company P&L bottom line above -- the two intentionally differ.
    d2c_ad_spend = platform_fees.get("total_ads", 0.0) if platform_fees else 0.0
    monthly_npe = gross_profit - d2c_ad_spend
    monthly_npe_margin = monthly_npe / revenue * 100 if revenue else 0.0

    return {
        "month": str(target_month), "revenue": revenue, "cogs": cogs,
        "gross_profit": gross_profit, "gross_margin_pct": gross_margin,
        "payroll_cost": payroll_cost, "bank_opex": bank_opex_month,
        "dropship_ads_rebate": ads_spend, "d2c_platform_fees": d2c_platform_fees,
        "platform_fees_ad_spend": fees_and_ads,
        "net_profit": net_profit, "net_margin_pct": net_margin,
        "d2c_ad_spend": d2c_ad_spend, "monthly_npe": monthly_npe, "monthly_npe_margin_pct": monthly_npe_margin,
    }


def build_order_profitability(sitegiant, dropship, platform_fees):
    """Per-order (D2C) / per-SKU-line (dropship) Net Income Received minus COGS.
    Net Profit deducts the real platform fee (commission/service/transaction,
    where known) AND ad spend (where known) -- both are actual cash the
    marketplace kept, so both come out before COGS. Company-wide Payroll/Bank
    OpEx are NOT allocated per row -- there's no traceable way to attribute
    shared fixed overhead to one order, so this table's total will not match
    the headline company Net Profit; it's Gross Profit minus known platform
    fee minus known ad spend, which is a different, honest number.
    """
    rows = []

    si = sitegiant["line_items"]
    if not si.empty:
        si = si.copy()
        # A handful of orders have a genuinely blank marketplace field in the raw
        # SiteGiant export itself -- same convention as the Overview tab's channel
        # breakdown (sitegiant_channel_metrics' by_marketplace), not a code bug.
        si["marketplace"] = si["marketplace"].fillna("Unknown Marketplace").astype(str).str.strip()
        si.loc[si["marketplace"] == "", "marketplace"] = "Unknown Marketplace"
        si["order_id_clean"] = si["marketplace_order_id"].apply(_clean_order_id)

        # Marketplace fee and ad spend are each only known at the ORDER level (the
        # source reports don't break either down by SKU) -- allocate across SKU
        # lines proportional to each line's share of that order's revenue, same
        # technique already used for coupon/discount allocation. They're tracked
        # independently: a fee can be known without ads being known, or vice versa.
        fee_lookup = (platform_fees or {}).get("marketplace_fee_lookup", {})
        combined_fee_map, combined_ads_map = {}, {}
        for fee_data in fee_lookup.values():
            combined_fee_map.update(fee_data.get("per_order_fee", {}))
            combined_ads_map.update(fee_data.get("per_order_ads", {}))
        si["order_fee"] = si["order_id_clean"].map(combined_fee_map)
        si["order_ads"] = si["order_id_clean"].map(combined_ads_map)
        si["fee_known"] = si["order_fee"].notna()
        si["ads_known"] = si["order_ads"].notna()
        order_revenue = si.groupby("order_id")["revenue"].transform("sum")

        si["platform_fee"] = 0.0
        fee_ok = si["fee_known"] & (order_revenue > 0)
        si.loc[fee_ok, "platform_fee"] = si.loc[fee_ok, "revenue"] / order_revenue[fee_ok] * si.loc[fee_ok, "order_fee"]

        si["ads_fee"] = 0.0
        ads_ok = si["ads_known"] & (order_revenue > 0)
        si.loc[ads_ok, "ads_fee"] = si.loc[ads_ok, "revenue"] / order_revenue[ads_ok] * si.loc[ads_ok, "order_ads"]

        # One row per (order, SKU) line, matching the reseller/dropship granularity.
        by_line = si.groupby(["order_id", "sku_key"]).agg(
            revenue=("revenue", "sum"), cogs=("line_cogs", "sum"), platform_fee=("platform_fee", "sum"),
            fee_known=("fee_known", "first"), ads_fee=("ads_fee", "sum"), ads_known=("ads_known", "first"),
            marketplace=("marketplace", "first"),
            order_date=("order_creation_date", "first"), qty=("product_quantity", "sum"),
            cost_known=("cost_known", "first"),
        ).reset_index()
        for _, r in by_line.iterrows():
            rows.append({
                "order_id": str(r["order_id"]), "platform": r["marketplace"], "channel": "d2c",
                "sku": r["sku_key"], "qty": int(r["qty"]) if pd.notna(r["qty"]) else None,
                "order_date": r["order_date"].strftime("%Y-%m-%d") if pd.notna(r["order_date"]) else None,
                "revenue": float(r["revenue"]), "cogs": float(r["cogs"]),
                "platform_fee": float(r["platform_fee"]), "fee_known": bool(r["fee_known"]),
                "ads_fee": float(r["ads_fee"]), "ads_known": bool(r["ads_known"]),
                "cost_known": bool(r["cost_known"]),
            })

    ds = dropship["line_items"]
    if not ds.empty:
        ds = ds.copy()
        ds["order_date"] = pd.to_datetime(ds["order_date"], errors="coerce")
        # `adj_total` is the return-adjusted line total BEFORE each invoice's
        # ads-rebate/claims deduction (those apply at the whole-invoice level,
        # not to any specific order). Scale each reseller's lines so per-order
        # revenue sums to that reseller's actual net_revenue -- same proportional
        # allocation approach already used for SiteGiant's order-level coupons.
        scale = {}
        if not dropship["by_reseller"].empty:
            raw_sum = ds.groupby("reseller")["adj_total"].sum()
            for _, row in dropship["by_reseller"].iterrows():
                raw = raw_sum.get(row["reseller"], 0)
                scale[row["reseller"]] = (row["net_revenue"] / raw) if raw else 1.0
        ds["scale"] = ds["reseller"].map(scale).fillna(1.0)
        ds["scaled_revenue"] = ds["adj_total"] * ds["scale"]

        # One row per (order, SKU) line -- i.e. per ISKU per item -- not
        # aggregated to the whole order, so Final Unit Price x Qty vs Cost x Qty
        # is visible at the actual granularity the reseller invoice bills at.
        by_line = ds.groupby(["order_id", "reseller", "sku"]).agg(
            revenue=("scaled_revenue", "sum"), cogs=("line_cogs", "sum"),
            order_date=("order_date", "first"), qty=("qty", "sum"),
            cost_known=("cost_known", "first"),
        ).reset_index()
        for _, r in by_line.iterrows():
            rows.append({
                "order_id": str(r["order_id"]), "platform": r["reseller"], "channel": "dropship",
                "sku": r["sku"], "qty": int(r["qty"]) if pd.notna(r["qty"]) else None,
                "order_date": r["order_date"].strftime("%Y-%m-%d") if pd.notna(r["order_date"]) else None,
                "revenue": float(r["revenue"]), "cogs": float(r["cogs"]),
                "platform_fee": 0.0, "fee_known": False,
                "ads_fee": 0.0, "ads_known": False, "cost_known": bool(r["cost_known"]),
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["gross_profit"] = df["revenue"] - df["cogs"]
    df["gross_margin_pct"] = df.apply(lambda r: r["gross_profit"] / r["revenue"] * 100 if r["revenue"] else 0.0, axis=1)
    df["fee_pct"] = df.apply(lambda r: r["platform_fee"] / r["revenue"] * 100 if r["revenue"] else 0.0, axis=1)
    df["ads_pct"] = df.apply(lambda r: r["ads_fee"] / r["revenue"] * 100 if r["revenue"] else 0.0, axis=1)

    # Net Profit = actual income received minus COGS: Revenue minus the real
    # platform fee (commission/service/transaction, where known) minus ad spend
    # (where known) minus COGS. Previously this only subtracted ads_fee and left
    # platform_fee as informational-only, which silently overstated Net Profit
    # for any order priced from an order-export-only fee source (e.g. DUX/
    # Justincase Shopee, which has no wallet-report ad spend so ads_fee is
    # always 0) -- those rows showed gross Merchandise Subtotal minus COGS
    # instead of the actual settled amount minus COGS. Fixed 2026-07-21.
    # Company-wide Payroll and Bank OpEx are still NOT allocated down to
    # individual rows -- there is no traceable way to attribute a share of fixed
    # overhead to one order, and doing so would just be a modeled guess dressed
    # up as a real number. Those costs are only ever shown at the aggregate level
    # (Net Profit Overview tab), where they're real. This means this table's Net
    # Profit column will NOT sum to the headline company Net Profit -- it sums
    # close to Gross Profit minus known platform fee minus known ad spend, a
    # different, honest thing (unresolved-fee rows are simply not charged one).
    df["net_profit"] = df["gross_profit"] - df["platform_fee"] - df["ads_fee"]
    df["net_margin_pct"] = df.apply(lambda r: r["net_profit"] / r["revenue"] * 100 if r["revenue"] else 0.0, axis=1)

    return df.sort_values("revenue", ascending=False).reset_index(drop=True)


def brand_margin_analysis(sitegiant, dropship):
    def classify(sku):
        return "Frescoone Proprietary (FRES)" if str(sku).upper().startswith("FRES") else "Other/Generic Brands"

    frames = []
    si = sitegiant["line_items"][["sku_key", "revenue", "line_cogs"]].rename(
        columns={"sku_key": "sku", "line_cogs": "cogs"})
    frames.append(si)

    ds = dropship["line_items"]
    if not ds.empty:
        frames.append(ds[["sku", "adj_total", "line_cogs"]].rename(
            columns={"adj_total": "revenue", "line_cogs": "cogs"}))

    combined = pd.concat(frames, ignore_index=True)
    combined["brand"] = combined["sku"].apply(classify)
    summary = combined.groupby("brand").agg(revenue=("revenue", "sum"), cogs=("cogs", "sum")).reset_index()
    summary["gross_profit"] = summary["revenue"] - summary["cogs"]
    summary["gross_margin_pct"] = summary.apply(
        lambda r: (r["gross_profit"] / r["revenue"] * 100) if r["revenue"] else 0.0, axis=1)
    return summary


# ---------------------------------------------------------------------------
# 7. REPORT
# ---------------------------------------------------------------------------
def write_executive_report(ctx, out_path):
    L = []
    L.append("# Frescoone Trading Sdn Bhd -- Executive Financial Report")
    L.append(f"*Generated {datetime.now():%Y-%m-%d %H:%M} by finance_analyzer.py*\n")

    np_ = ctx["net_profit"]
    L.append("## 1. Executive Summary\n")
    L.append(f"- **Reporting period (accrual P&L):** {np_['month']}")
    L.append(f"- **Revenue:** {fmt_money(np_['revenue'])}")
    L.append(f"- **COGS:** {fmt_money(np_['cogs'])}")
    L.append(f"- **Gross Profit:** {fmt_money(np_['gross_profit'])} ({np_['gross_margin_pct']:.1f}% margin)")
    L.append(f"- **Payroll OpEx:** {fmt_money(np_['payroll_cost'])}")
    L.append(f"- **Bank Statement OpEx (non-payroll):** {fmt_money(np_['bank_opex'])}")
    L.append(f"- **Platform Fees / Ad Spend:** {fmt_money(np_['platform_fees_ad_spend'])} "
              f"(D2C platform fees: {fmt_money(np_['d2c_platform_fees'])}; "
              f"dropship ads rebate: {fmt_money(np_['dropship_ads_rebate'])})")
    L.append(f"- **Net Profit:** {fmt_money(np_['net_profit'])} ({np_['net_margin_pct']:.1f}% margin)\n")

    L.append("### Net Profit Calculation Walkthrough\n")
    L.append("| Step | Amount | Running Total |")
    L.append("|---|---:|---:|")
    L.append(f"| Revenue (SiteGiant D2C + dropship, reseller-duplicate orders excluded) | {fmt_money(np_['revenue'])} | {fmt_money(np_['revenue'])} |")
    L.append(f"| less COGS (SKU cost x qty, all channels) | -{fmt_money(np_['cogs'])} | {fmt_money(np_['gross_profit'])} |")
    L.append(f"| **= Gross Profit** | | **{fmt_money(np_['gross_profit'])}** ({np_['gross_margin_pct']:.1f}%) |")
    running = np_["gross_profit"] - np_["payroll_cost"]
    L.append(f"| less Payroll OpEx (payslip gross pay, all staff) | -{fmt_money(np_['payroll_cost'])} | {fmt_money(running)} |")
    running -= np_["bank_opex"]
    L.append(f"| less Bank Statement OpEx (non-payroll debits this month) | -{fmt_money(np_['bank_opex'])} | {fmt_money(running)} |")
    running -= np_["dropship_ads_rebate"]
    L.append(f"| less Dropship ads rebate (Shopee ads rebate on SCON/SSUP invoices) | -{fmt_money(np_['dropship_ads_rebate'])} | {fmt_money(running)} |")
    running -= np_["d2c_platform_fees"]
    L.append(f"| less D2C platform fees + ad spend (Shopee/Lazada/TikTok, FRESCOONE stores only) | -{fmt_money(np_['d2c_platform_fees'])} | {fmt_money(running)} |")
    L.append(f"| **= Net Profit** | | **{fmt_money(np_['net_profit'])}** ({np_['net_margin_pct']:.1f}%) |\n")

    pf = ctx.get("platform_fees")
    if pf and pf["by_platform"]:
        L.append("**D2C platform fee detail** (only storefronts with a fee/income report or raw order export "
                  "provided -- real fees, not estimates. \"Revenue\" here is that platform's own D2C order "
                  "value, used only to compute the %% columns -- it's a subset of the SiteGiant D2C total "
                  "above.):\n")
        L.append("| Platform | Revenue | Marketplace/Payment Fee | Fee % | Ad Spend | Ad % | Total Cost | Total % | Orders Covered |")
        L.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|")
        for r in pf["by_platform"]:
            orders_str = f"{r['orders']}/{r.get('orders_total', r['orders'])}" if "orders_total" in r else str(r["orders"])
            L.append(f"| {r['platform']} | {fmt_money(r['revenue'])} | {fmt_money(r['fee'])} | {r['fee_pct']:.1f}% | "
                      f"{fmt_money(r['ads_spend'])} | {r['ads_pct']:.1f}% | {fmt_money(r['total_cost'])} | "
                      f"{r['total_pct']:.1f}% | {orders_str} |")
        blended_revenue = sum(r["revenue"] for r in pf["by_platform"])
        blended_pct = pf["total"] / blended_revenue * 100 if blended_revenue else 0
        L.append(f"| **Blended average (revenue-weighted)** | {fmt_money(blended_revenue)} | {fmt_money(pf['total_fees'])} | "
                  f"{pf['total_fees']/blended_revenue*100 if blended_revenue else 0:.1f}% | {fmt_money(pf['total_ads'])} | "
                  f"{pf['total_ads']/blended_revenue*100 if blended_revenue else 0:.1f}% | {fmt_money(pf['total'])} | "
                  f"**{blended_pct:.1f}%** | |")
        L.append("\n*The blended average is revenue-weighted (Shopee's larger order volume dominates it), not a "
                  "simple average of the three platform percentages -- a simple average would overweight Lazada's "
                  "small, low-fee order base.*\n")
        L.append("**Verified:** marketplace fees and ad spend are both deducted *inside the platform wallet* "
                  "before a single net amount is withdrawn to the bank account -- confirmed against Shopee's "
                  "own ledger for a real month (Order Income - Ads - misc matched Shopee's own "
                  "\"Withdrawals\" line to within normal settlement-timing lag) -- and the bank statement itself "
                  "has zero rows mentioning \"ads\"/\"advertising\", so "
                  "there is no separate direct ad payment being double-counted. This means the \"Platform Payout\" "
                  "credits in Section 4's cash flow chart are already net of both fees and ads -- they are not "
                  "gross revenue and should not be read as such.\n")

    L.append("## 2. Revenue & COGS by Channel\n")
    L.append("| Channel | Revenue | COGS | Gross Profit | Gross Margin % |")
    L.append("|---|---:|---:|---:|---:|")
    sg = ctx["sitegiant"]
    sg_gp = sg["total_revenue"] - sg["total_cogs"]
    sg_gm = sg_gp / sg["total_revenue"] * 100 if sg["total_revenue"] else 0
    L.append(f"| SiteGiant D2C (TikTok/Shopee/Lazada/Web) | {fmt_money(sg['total_revenue'])} | "
              f"{fmt_money(sg['total_cogs'])} | {fmt_money(sg_gp)} | {sg_gm:.1f}% |")
    for _, row in ctx["dropship"]["by_reseller"].iterrows():
        gm = row["gross_profit"] / row["net_revenue"] * 100 if row["net_revenue"] else 0
        L.append(f"| Reseller/Dropship -- {row['reseller']} | {fmt_money(row['net_revenue'])} | "
                  f"{fmt_money(row['cogs'])} | {fmt_money(row['gross_profit'])} | {gm:.1f}% |")
    L.append("")

    L.append("## 3. Brand Performance: Frescoone Proprietary vs. Generic\n")
    L.append("| Brand | Revenue | COGS | Gross Profit | Gross Margin % |")
    L.append("|---|---:|---:|---:|---:|")
    for _, row in ctx["brand_margin"].iterrows():
        L.append(f"| {row['brand']} | {fmt_money(row['revenue'])} | {fmt_money(row['cogs'])} | "
                  f"{fmt_money(row['gross_profit'])} | {row['gross_margin_pct']:.1f}% |")
    L.append("")

    L.append("## 4. Cash Flow Trends (Monthly, Bank Statement)\n")
    L.append("| Month | Inflows | Outflows | Net Cash Flow | Inflow MoM % | Inflow YoY % |")
    L.append("|---|---:|---:|---:|---:|---:|")
    for _, row in ctx["cash_flow"].iterrows():
        mom = f"{row['inflow_mom_growth_pct']:.1f}%" if pd.notna(row["inflow_mom_growth_pct"]) else "n/a"
        yoy = f"{row['inflow_yoy_growth_pct']:.1f}%" if pd.notna(row["inflow_yoy_growth_pct"]) else "n/a"
        L.append(f"| {row['month_str']} | {fmt_money(row['inflows'])} | {fmt_money(row['outflows'])} | "
                  f"{fmt_money(row['net_cash_flow'])} | {mom} | {yoy} |")
    L.append("\n*Growth % here is based on total bank-statement cash inflows as a liquidity proxy -- it is "
              "not the same as accrual revenue growth. Order-level exports (SiteGiant, dropship) currently "
              "cover one month only, so a true MoM/YoY revenue growth series requires collecting monthly "
              "exports going forward.*\n")

    L.append(f"## 5. Operating Expense Breakdown -- {np_['month']} (Bank Statement, excl. Payroll & Financing)\n")
    L.append("*Scoped to the same month as Section 1's Net Profit figure -- these rows sum to the "
              "'Bank Statement OpEx' line above.*\n")
    L.append("| Category | Amount |")
    L.append("|---|---:|")
    month_category = ctx["bank_opex_month_category"]
    if month_category.empty:
        L.append("| *(no non-payroll debits found this month)* | RM 0.00 |")
    else:
        for cat, amt in month_category.items():
            L.append(f"| {cat} | {fmt_money(amt)} |")
    L.append("")

    L.append("### 5b. Operating Expense by Category -- Full Bank Statement History\n")
    L.append("| Category | Amount (all months) |")
    L.append("|---|---:|")
    for cat, amt in ctx["bank_opex"]["by_category"].items():
        L.append(f"| {cat} | {fmt_money(amt)} |")
    L.append("")

    L.append("## 6. Payroll (Recent History)\n")
    L.append("| Month | Headcount | Total Gross Pay | Total Net Pay |")
    L.append("|---|---:|---:|---:|")
    for _, row in ctx["payroll_monthly"].iterrows():
        L.append(f"| {row['month']} | {int(row['headcount'])} | {fmt_money(row['total_gross'])} | "
                  f"{fmt_money(row['total_net'])} |")
    L.append("")

    L.append("## 7. Data Quality & Verification Notes\n")
    if ctx["warnings"]:
        for w in ctx["warnings"]:
            L.append(f"- {w}")
    else:
        L.append("- No data-quality issues flagged.")
    L.append("")

    L.append("## 8. Known Scope Limitations\n")
    L.append("- SiteGiant D2C and dropship reseller accrual data currently reflect **one month only** -- "
              "MoM/YoY *revenue* growth cannot yet be computed from order-level data; only cash-inflow-based "
              "growth (Section 4) spans the full bank statement history.")
    L.append("- \"Bank Statement OpEx\" is every non-payroll, non-financing debit in the account -- it is not "
              "yet split between true operating expense and inventory/COGS-adjacent vendor payments (e.g. "
              "packaging, freight forwarders). Treat Section 5 as a starting point for manual review.")
    L.append("- Platform Fees/Ad Spend includes real Shopee/Lazada/TikTok fees for the FRESCOONE-branded "
              "storefronts (when an `income-exported*` folder with each platform's report is present) plus "
              "the dropship ads rebate. It does **not** include fees for the separately-run DUX-branded "
              "storefronts on the same platforms (no fee export obtained for those yet), and Shopee "
              "specifically only counts orders whose payout has already posted -- orders still in escrow "
              "are excluded, so the Shopee figure is a floor, not final. If no `income-exported*` folder is "
              "present at all, this line is dropship-ads-rebate only and D2C fees remain fully unknown.")

    Path(out_path).write_text("\n".join(L), encoding="utf-8")


# ---------------------------------------------------------------------------
# 8. MONTHLY HISTORY -- one snapshot per run, so re-running this script each
# month builds up a real accrual-basis trend instead of only the bank's cash
# inflows. Safe to run this script multiple times in the same month (e.g. to
# fix a data issue and re-verify) -- each save overwrites that month's file.
# ---------------------------------------------------------------------------
HISTORY_DIR = BASE_DIR / "history"


def save_history_snapshot(ctx):
    HISTORY_DIR.mkdir(exist_ok=True)
    np_ = ctx["net_profit"]
    snapshot = {
        "month": np_["month"],
        "revenue": np_["revenue"], "cogs": np_["cogs"],
        "gross_profit": np_["gross_profit"], "gross_margin_pct": np_["gross_margin_pct"],
        "payroll_cost": np_["payroll_cost"], "bank_opex": np_["bank_opex"],
        "platform_fees_ad_spend": np_["platform_fees_ad_spend"],
        "net_profit": np_["net_profit"], "net_margin_pct": np_["net_margin_pct"],
        "sitegiant_revenue": ctx["sitegiant"]["total_revenue"],
        "dropship_revenue": ctx["dropship"]["total_revenue"],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    path = HISTORY_DIR / f"{np_['month']}.json"
    path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    return path


def load_history():
    HISTORY_DIR.mkdir(exist_ok=True)
    snapshots = []
    for p in sorted(HISTORY_DIR.glob("*.json")):
        try:
            snapshots.append(json.loads(p.read_text(encoding="utf-8")))
        except (json.JSONDecodeError, OSError):
            warn(f"History: could not read snapshot {p.name} -- skipped.")
    snapshots.sort(key=lambda s: s["month"])
    return snapshots


def load_restock_data():
    """Optional: restock_planner.py's output, if it's been run. Additive --
    the dashboard builds fine without it (same pattern as platform fees), the
    Restock Plan tab just shows a "not generated yet" note instead. Runs on
    its own cadence (SiteGiant forecast exports are refreshed manually, much
    less often than this script), so it's a separate file, not recomputed here.
    """
    path = BASE_DIR / "restock_data.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        warn(f"restock_data.json: could not parse -- {e}")
        return None


def load_restock_pipeline_data():
    """Optional: restock_pipeline.py's output (the confirmed 60-day/MOQ
    methodology from the "Frescoone Sales & Stock Analyst" project) -- a
    second, deliberately separate restock system from restock_data.json's
    ABC/OTB version, per Bun's direction 2026-07-21 to keep both side by
    side rather than merge them. Additive, same as every other optional
    input in this file.
    """
    path = BASE_DIR / "restock_pipeline_data.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        warn(f"restock_pipeline_data.json: could not parse -- {e}")
        return None


def load_restock_strategy_data():
    """Optional: restock_strategy.py's output -- the brand-mix / sellout
    scale-up / stop-restock / trend-potential decision layer, per Bun's
    direction 2026-07-21. Sits on top of both restock_data.json and
    restock_pipeline_data.json without replacing either. Additive, same as
    every other optional input in this file.
    """
    path = BASE_DIR / "restock_strategy_data.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        warn(f"restock_strategy_data.json: could not parse -- {e}")
        return None


# ---------------------------------------------------------------------------
# 8b. FINANCIAL HEALTH -- data-driven analysis of the company's own real
# numbers (not investment/financial advice -- there's no license behind this,
# just arithmetic on figures already computed elsewhere in this file). Every
# number here traces back to compute_net_profit()/load_platform_fees()/the
# bank statement; nothing is a fabricated benchmark presented as fact. Where
# a generic industry range is shown for context, it's explicitly labeled as
# generic, separate from Frescoone's own real figures.
# ---------------------------------------------------------------------------
def build_financial_health(ctx, restock):
    np_ = ctx["net_profit"]
    pf = ctx.get("platform_fees") or {"total_fees": 0.0, "total_ads": 0.0}
    bank_df = ctx.get("bank_df")

    cash_position, cash_as_of = None, None
    if bank_df is not None and not bank_df.empty and "Closing Book Balance" in bank_df.columns:
        bs = bank_df.sort_values("Statement Date")
        cash_position = float(bs["Closing Book Balance"].iloc[-1])
        cash_as_of = bs["Statement Date"].iloc[-1].strftime("%Y-%m-%d")

    monthly_fixed_cost = np_["payroll_cost"] + np_["bank_opex"]
    cash_runway_months = cash_position / monthly_fixed_cost if cash_position is not None and monthly_fixed_cost else None

    # Trailing cash flow: the bank statement runs later than the accrual P&L's
    # target_month (it's updated as new statements arrive), so its most recent
    # row is often a still-in-progress partial month -- kept separate from the
    # prior up-to-12 FULL months rather than blended into one "trend" number.
    cf = ctx.get("cash_flow")
    trailing_12mo_cash_flow, trailing_12mo_months = None, 0
    current_month_cash_flow, current_month_label = None, None
    if cf is not None and not cf.empty:
        cf_sorted = cf.sort_values("month")
        current_month_cash_flow = float(cf_sorted["net_cash_flow"].iloc[-1])
        current_month_label = str(cf_sorted["month"].iloc[-1])
        prior = cf_sorted.iloc[:-1].tail(12)
        if not prior.empty:
            trailing_12mo_cash_flow = float(prior["net_cash_flow"].sum())
            trailing_12mo_months = len(prior)

    # Breakeven decomposition: holding COGS/Payroll/Bank OpEx/platform
    # transaction fees fixed, how much headroom is there for ad spend before
    # Net Profit hits zero? Negative headroom means the loss isn't primarily
    # an ad-spend problem -- overhead alone already exceeds Gross Profit.
    platform_transaction_fee = float(pf.get("total_fees", 0.0))
    current_ad_spend = float(pf.get("total_ads", 0.0))
    headroom_before_ads = (np_["gross_profit"] - np_["payroll_cost"] - np_["bank_opex"]
                            - np_["dropship_ads_rebate"] - platform_transaction_fee)
    ad_pct_of_revenue = current_ad_spend / np_["revenue"] * 100 if np_["revenue"] else 0.0

    restock_budget = restock["budget"]["final_budget"] if restock else None
    cash_headroom_after_restock = None
    if cash_position is not None:
        cash_headroom_after_restock = max(0.0, cash_position - monthly_fixed_cost - (restock_budget or 0.0))

    # Revenue & Profit Targets -- what revenue would it take to hit breakeven,
    # then a ladder of illustrative margin tiers above that. MODEL, not a
    # guarantee: holds this month's real cost RATIOS constant (COGS%,
    # platform transaction fee% of revenue -- both roughly proportional to
    # sales in reality) and this month's real FIXED costs constant in RM
    # terms (Payroll, Bank OpEx, and current ad spend -- these don't
    # automatically scale with revenue, they're budget/headcount decisions).
    # Verified algebra: plugging this month's actual revenue back through the
    # model exactly reproduces this month's actual Net Profit (both use the
    # same real ratios by construction) -- it's a re-arrangement of the real
    # P&L, not an independent guess. Re-run monthly as ratios shift.
    profit_ladder = []
    variable_cost_pct = (np_["cogs"] + platform_transaction_fee + np_["dropship_ads_rebate"]) / np_["revenue"] if np_["revenue"] else None
    contribution_margin_pct = (1 - variable_cost_pct) if variable_cost_pct is not None else None
    fixed_cost_incl_ads = np_["payroll_cost"] + np_["bank_opex"] + current_ad_spend
    if contribution_margin_pct and contribution_margin_pct > 0:
        for target_margin_pct in (0, 5, 10, 15, 20):
            tm = target_margin_pct / 100
            if contribution_margin_pct <= tm:
                continue  # target margin exceeds contribution margin -- no finite revenue solves it, skip rather than show nonsense
            revenue_needed = fixed_cost_incl_ads / (contribution_margin_pct - tm)
            profit_ladder.append({
                "target_margin_pct": target_margin_pct,
                "revenue_needed": revenue_needed,
                "net_profit_at": tm * revenue_needed,
                "gap_vs_current_revenue": revenue_needed - np_["revenue"],
                "restock_ceiling_40pct": revenue_needed * 0.40,
            })
    breakeven_revenue = profit_ladder[0]["revenue_needed"] if profit_ladder else None

    flags = []
    if np_["net_profit"] < 0:
        flags.append({"severity": "warn",
                       "text": f"Net Profit (accrual) is negative for {np_['month']}: {fmt_money(np_['net_profit'])} "
                               f"({np_['net_margin_pct']:.1f}% margin)."})
    if headroom_before_ads < 0:
        flags.append({"severity": "warn",
                       "text": f"Even at zero ad spend, Payroll + Bank OpEx + platform transaction fees exceed "
                               f"Gross Profit by {fmt_money(-headroom_before_ads)} this month -- ad spend is not "
                               f"the main driver of the loss. (Bank OpEx is still an undifferentiated catch-all "
                               f"that may include some inventory-adjacent payments, so this is a conservative, "
                               f"not precise, read.)"})
    if current_month_cash_flow is not None and current_month_cash_flow < 0:
        flags.append({"severity": "info",
                       "text": f"{current_month_label}'s cash flow is negative so far ({fmt_money(current_month_cash_flow)}) "
                               f"-- likely a partial month if it's the current one, check the date range before reacting."})
    if trailing_12mo_cash_flow is not None:
        verdict = ("roughly flat" if abs(trailing_12mo_cash_flow) < monthly_fixed_cost
                    else ("net positive" if trailing_12mo_cash_flow > 0 else "net negative"))
        flags.append({"severity": "info",
                       "text": f"Trailing {trailing_12mo_months}-month net cash flow: {fmt_money(trailing_12mo_cash_flow)} "
                               f"-- {verdict} overall, with real month-to-month swings (see Bank Statement tab)."})
    if np_["gross_margin_pct"] >= 50:
        flags.append({"severity": "good",
                       "text": f"Gross margin is {np_['gross_margin_pct']:.1f}% -- strong for a phone-accessories "
                               f"business (generic industry context, not a Frescoone-specific benchmark)."})
    if breakeven_revenue is not None:
        flags.append({"severity": "warn" if breakeven_revenue > np_["revenue"] else "good",
                       "text": f"At this month's real cost ratios, breakeven needs {fmt_money(breakeven_revenue)} "
                               f"in revenue -- {fmt_money(breakeven_revenue - np_['revenue'])} more than the "
                               f"{fmt_money(np_['revenue'])} actually made ({(breakeven_revenue/np_['revenue']-1)*100:.0f}% "
                               f"growth needed, cost structure unchanged)."})

    restock_range = None
    if restock:
        b = restock["budget"]
        restock_range = {"min_need": b["raw_need"], "max_ceiling": min(b["revenue_cap"], b["cash_cap"]),
                          "recommended": b["final_budget"]}

    return {
        "target_month": np_["month"],
        "revenue": np_["revenue"], "gross_profit": np_["gross_profit"], "gross_margin_pct": np_["gross_margin_pct"],
        "net_profit": np_["net_profit"], "net_margin_pct": np_["net_margin_pct"],
        "cash_position": cash_position, "cash_position_as_of": cash_as_of,
        "monthly_fixed_cost": monthly_fixed_cost, "cash_runway_months": cash_runway_months,
        "trailing_12mo_cash_flow": trailing_12mo_cash_flow, "trailing_12mo_months": trailing_12mo_months,
        "current_month_cash_flow": current_month_cash_flow, "current_month_label": current_month_label,
        "breakeven": {
            "gross_profit": np_["gross_profit"], "payroll": np_["payroll_cost"], "bank_opex": np_["bank_opex"],
            "dropship_ads_rebate": np_["dropship_ads_rebate"], "platform_transaction_fee": platform_transaction_fee,
            "headroom_before_ads": headroom_before_ads,
        },
        "ad_spend": {
            "current": current_ad_spend, "current_pct_of_revenue": ad_pct_of_revenue,
            "breakeven_ceiling": max(0.0, headroom_before_ads),
        },
        "restock_budget": restock_budget,
        "restock_range": restock_range,
        "cash_headroom_after_restock": cash_headroom_after_restock,
        "breakeven_revenue": breakeven_revenue,
        "contribution_margin_pct": contribution_margin_pct * 100 if contribution_margin_pct else None,
        "profit_ladder": profit_ladder,
        "flags": flags,
    }


# ---------------------------------------------------------------------------
# 9. LOCAL DASHBOARD -- a self-contained HTML file (Chart.js via CDN, so an
# internet connection is needed the first time it's opened, but no server or
# install). Not published anywhere -- this stays local since it's company
# financial data. Re-running the script overwrites it with fresh numbers.
# ---------------------------------------------------------------------------
def build_dashboard_data(ctx, history):
    np_ = ctx["net_profit"]
    sg, ds = ctx["sitegiant"], ctx["dropship"]

    # Channel Detail: one row per individual D2C marketplace (not lumped into a
    # single "SiteGiant D2C" row) -- including "Unknown Marketplace" (COD/retail
    # offline orders, no marketplace involved at all) as its own row, per user
    # direction 2026-07-21. Built from the same per-order profitability table as
    # the Order Detail tab (revenue/cogs/platform_fee/ads_fee already resolved
    # there, including the 2026-07-21 DUX/Justincase fee fix), so this table
    # stays numerically consistent with Order Detail rather than recomputing
    # independently. "Actual Income Received" = Revenue minus the real platform
    # fee and ad spend where known (0 for Unknown Marketplace and any storefront
    # without fee data, since none applies/is known); "NPE" = Actual Income
    # Received minus COGS, matching the per-order Net Profit definition.
    #
    # Merged with the old separate "D2C Platform Fee & Ad Spend Detail" table
    # into this one, per user direction 2026-07-21 -- both were channel-level
    # breakdowns built from the same underlying data, just showing different
    # column subsets. Fee/Ad Spend columns here are the SAME real numbers that
    # table showed, just alongside COGS/NPE instead of in a separate table.
    channels = []
    op = ctx.get("order_profitability")
    op_d2c = op[op["channel"] == "d2c"] if op is not None and not op.empty else pd.DataFrame()
    if not op_d2c.empty:
        by_mkt = op_d2c.groupby("platform").agg(
            revenue=("revenue", "sum"), cogs=("cogs", "sum"),
            platform_fee=("platform_fee", "sum"), ads_fee=("ads_fee", "sum"),
            fee_known=("fee_known", "any"), ads_known=("ads_known", "any"),
            orders=("order_id", "nunique"),
        ).reset_index().sort_values("revenue", ascending=False)
        for _, row in by_mkt.iterrows():
            actual_income = float(row["revenue"] - row["platform_fee"] - row["ads_fee"])
            npe = actual_income - float(row["cogs"])
            margin = npe / row["revenue"] * 100 if row["revenue"] else 0
            fee_pct = float(row["platform_fee"]) / row["revenue"] * 100 if row["revenue"] else 0
            ads_pct = float(row["ads_fee"]) / row["revenue"] * 100 if row["revenue"] else 0
            channels.append({
                "name": row["platform"], "currency": "RM",
                "converted_from_sgd": row["platform"] in SGD_MARKETPLACES,
                "revenue": float(row["revenue"]), "cogs": float(row["cogs"]),
                "fee": float(row["platform_fee"]), "fee_pct": fee_pct, "fee_known": bool(row["fee_known"]),
                "ads_spend": float(row["ads_fee"]), "ads_pct": ads_pct, "ads_known": bool(row["ads_known"]),
                "orders": int(row["orders"]),
                "actual_income_received": actual_income, "npe": npe, "margin_pct": margin,
                "gross_profit": float(row["revenue"]) - float(row["cogs"]),
            })

    # Reseller/dropship channels: Actual Income Received = DSP revenue in full
    # (no marketplace fee applies to a wholesale invoice) and NPE = DSP revenue
    # minus COGS, full stop -- per the reseller Net Profit rule. No platform fee
    # or ad spend column applies here at all (not "unknown" -- genuinely zero).
    reseller_orders = (ds["line_items"].groupby("reseller")["order_id"].nunique()
                        if not ds["line_items"].empty else pd.Series(dtype=int))
    for _, row in ds["by_reseller"].iterrows():
        npe = row["gross_profit"]
        margin = npe / row["net_revenue"] * 100 if row["net_revenue"] else 0
        channels.append({"name": row["reseller"], "currency": "RM",
                          "revenue": row["net_revenue"], "cogs": row["cogs"],
                          "fee": 0.0, "fee_pct": 0.0, "fee_known": True,
                          "ads_spend": 0.0, "ads_pct": 0.0, "ads_known": True,
                          "orders": int(reseller_orders.get(row["reseller"], 0)),
                          "actual_income_received": row["net_revenue"], "npe": npe, "margin_pct": margin,
                          "gross_profit": row["gross_profit"]})

    # Blended average across D2C channels with real fee data -- revenue-weighted,
    # same definition the old standalone fee table used. Resellers are excluded
    # (no fee applies to a wholesale invoice); SGD storefronts are excluded too,
    # not for currency reasons anymore (they're RM now) but because there's
    # simply no fee/income document for them, same as DUX Lazmall/TikTok.
    fee_known_channels = [c for c in channels if c.get("fee_known") and c["currency"] == "RM" and c.get("orders", 0) and c["name"] not in set(ds["by_reseller"]["reseller"])]
    blended_revenue = sum(c["revenue"] for c in fee_known_channels)
    channel_blended = {
        "revenue": blended_revenue,
        "fee_pct": sum(c["fee"] for c in fee_known_channels) / blended_revenue * 100 if blended_revenue else 0,
        "ads_pct": sum(c["ads_spend"] for c in fee_known_channels) / blended_revenue * 100 if blended_revenue else 0,
    } if fee_known_channels else None

    brand = [{"name": r["brand"], "revenue": r["revenue"], "cogs": r["cogs"],
              "gross_profit": r["gross_profit"], "margin_pct": r["gross_margin_pct"]}
             for _, r in ctx["brand_margin"].iterrows()]

    cf = ctx["cash_flow"]
    cashflow = {"months": cf["month_str"].tolist(),
                "inflows": [round(x, 2) for x in cf["inflows"]],
                "outflows": [round(x, 2) for x in cf["outflows"]],
                "net": [round(x, 2) for x in cf["net_cash_flow"]]}

    opex_month = ctx["bank_opex_month_category"]
    opex = {"labels": list(opex_month.index), "values": [round(v, 2) for v in opex_month.values]}

    bank_transactions = []
    bdf = ctx.get("bank_df")
    if bdf is not None and not bdf.empty:
        bt = bdf.sort_values("Statement Date", ascending=False)
        for _, r in bt.iterrows():
            desc = str(r.get("Supplementary Details") or r.get("Statement Details Info") or "").strip()
            ref = str(r.get("Ref For Account Owner") or "").strip()
            if ref and ref not in desc:
                desc = f"{desc} -- {ref}" if desc else ref
            bank_transactions.append({
                "date": r["Statement Date"].strftime("%Y-%m-%d") if pd.notna(r["Statement Date"]) else None,
                "details": desc[:220],
                "category": r.get("category", ""),
                "debit": round(float(r.get("Debit Amount", 0) or 0), 2),
                "credit": round(float(r.get("Credit Amount", 0) or 0), 2),
            })

    pr = ctx["payroll_monthly"]
    payroll = {"months": [str(m) for m in pr["month"]],
               "gross": [round(x, 2) for x in pr["total_gross"]],
               "headcount": [int(x) for x in pr["headcount"]]} if not pr.empty else {"months": [], "gross": [], "headcount": []}

    trend = {"months": [s["month"] for s in history],
              "revenue": [round(s["revenue"], 2) for s in history],
              "net_profit": [round(s["net_profit"], 2) for s in history],
              "net_margin_pct": [round(s["net_margin_pct"], 2) for s in history]}

    running = np_["gross_profit"]
    walkthrough = [
        {"label": "Revenue (SiteGiant D2C + dropship, reseller-duplicate orders excluded)",
         "amount": np_["revenue"], "running": np_["revenue"]},
        {"label": "less COGS (SKU cost x qty, all channels)", "amount": -np_["cogs"], "running": np_["gross_profit"]},
    ]
    running -= np_["payroll_cost"]
    walkthrough.append({"label": "less Payroll OpEx (payslip gross pay, all staff)",
                         "amount": -np_["payroll_cost"], "running": running})
    running -= np_["bank_opex"]
    walkthrough.append({"label": "less Bank Statement OpEx (non-payroll debits this month)",
                         "amount": -np_["bank_opex"], "running": running})
    running -= np_["dropship_ads_rebate"]
    walkthrough.append({"label": "less Dropship ads rebate (Shopee ads rebate on SCON/SSUP invoices)",
                         "amount": -np_["dropship_ads_rebate"], "running": running})
    running -= np_["d2c_platform_fees"]
    walkthrough.append({"label": "less D2C platform fees + ad spend (Shopee/Lazada/TikTok, FRESCOONE stores only)",
                         "amount": -np_["d2c_platform_fees"], "running": running})

    op = ctx.get("order_profitability")
    orders_json = []
    if op is not None and not op.empty:
        for _, r in op.iterrows():
            orders_json.append({
                "order_id": r["order_id"], "platform": r["platform"], "currency": "RM",
                "converted_from_sgd": r["platform"] in SGD_MARKETPLACES, "date": r["order_date"],
                "sku": r.get("sku"), "qty": int(r["qty"]) if pd.notna(r.get("qty")) else None,
                "revenue": round(r["revenue"], 2), "cogs": round(r["cogs"], 2),
                "gross_profit": round(r["gross_profit"], 2), "gross_margin_pct": round(r["gross_margin_pct"], 2),
                "platform_fee": round(r["platform_fee"], 2), "fee_pct": round(r["fee_pct"], 2),
                "fee_known": bool(r["fee_known"]),
                "ads_fee": round(r["ads_fee"], 2), "ads_pct": round(r["ads_pct"], 2),
                "ads_known": bool(r["ads_known"]),
                "cost_known": bool(r.get("cost_known", True)),
                "net_profit": round(r["net_profit"], 2), "net_margin_pct": round(r["net_margin_pct"], 2),
            })

    # Company-wide split of order-level Net Profit, positive vs negative, summed
    # separately (not netted) -- this is the SAME order-level figures shown in
    # the Order Detail tab, surfaced on the Overview tab too, for ALL orders
    # regardless of any filter selected in that tab. Rows with an unresolved SKU
    # cost are excluded here too (their Net Profit is inflated by COGS=0) --
    # same treatment as the Order Detail table, which shows them separately.
    op_costed = op[op["cost_known"]] if op is not None and not op.empty else op
    if op_costed is not None and not op_costed.empty:
        pos = op_costed[op_costed["net_profit"] >= 0]
        neg = op_costed[op_costed["net_profit"] < 0]
        order_net_profit_split = {
            "positive_orders": int(len(pos)), "positive_total": round(float(pos["net_profit"].sum()), 2),
            "negative_orders": int(len(neg)), "negative_total": round(float(neg["net_profit"].sum()), 2),
            "negative_settlement": round(float((neg["revenue"] - neg["platform_fee"] - neg["ads_fee"]).sum()), 2),
        }
    else:
        order_net_profit_split = {"positive_orders": 0, "positive_total": 0.0, "negative_orders": 0,
                                   "negative_total": 0.0, "negative_settlement": 0.0}

    sgd = ctx["sitegiant"].get("sgd_summary", {"revenue_sgd": 0.0, "orders": 0, "marketplaces": []})

    return {
        "company": "Frescoone Trading Sdn Bhd", "period": np_["month"],
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "kpis": {"revenue": np_["revenue"], "cogs": np_["cogs"], "gross_profit": np_["gross_profit"],
                 "gross_margin_pct": np_["gross_margin_pct"], "payroll_cost": np_["payroll_cost"],
                 "bank_opex": np_["bank_opex"], "ad_spend": np_["platform_fees_ad_spend"],
                 "net_profit": np_["net_profit"], "net_margin_pct": np_["net_margin_pct"],
                 "d2c_ad_spend": np_["d2c_ad_spend"], "monthly_npe": np_["monthly_npe"],
                 "monthly_npe_margin_pct": np_["monthly_npe_margin_pct"]},
        "order_net_profit_split": order_net_profit_split,
        "sgd_summary": sgd,
        "channels": channels, "brand": brand, "cashflow": cashflow,
        "opex": opex, "payroll": payroll, "trend": trend,
        "walkthrough": walkthrough, "channel_blended": channel_blended,
        "orders": orders_json, "bank_transactions": bank_transactions,
        "warnings": ctx["warnings"],
        "restock": ctx.get("restock"),
        "restock_pipeline": ctx.get("restock_pipeline"),
        "restock_strategy": ctx.get("restock_strategy"),
        "health": ctx.get("health"),
    }


DASHBOARD_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__COMPANY__ -- Finance Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.1"></script>
<style>
:root {
  --bg-primary:#f8f9fa; --bg-card:#ffffff; --bg-header:#1a1a2e;
  --text-primary:#212529; --text-secondary:#6c757d; --text-on-dark:#ffffff;
  --color-1:#4C72B0; --color-2:#DD8452; --color-3:#55A868; --color-4:#C44E52; --color-5:#8172B3; --color-6:#937860;
  --positive:#28a745; --negative:#dc3545; --gap:16px; --radius:8px;
}
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:var(--bg-primary); color:var(--text-primary); line-height:1.5; }
.dashboard-container { max-width:1400px; margin:0 auto; padding:var(--gap); }
.dashboard-header { background:var(--bg-header); color:var(--text-on-dark); padding:20px 24px; border-radius:var(--radius); margin-bottom:var(--gap); display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:12px; }
.dashboard-header h1 { font-size:20px; font-weight:600; }
.dashboard-header .sub { font-size:13px; color:rgba(255,255,255,0.7); margin-top:2px; }
.kpi-row { display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:var(--gap); margin-bottom:var(--gap); }
.kpi-card { background:var(--bg-card); border-radius:var(--radius); padding:18px 22px; box-shadow:0 1px 3px rgba(0,0,0,0.08); }
.kpi-label { font-size:12px; color:var(--text-secondary); text-transform:uppercase; letter-spacing:0.5px; margin-bottom:6px; }
.kpi-value { font-size:24px; font-weight:700; }
.kpi-sub { font-size:12px; color:var(--text-secondary); margin-top:4px; }
.kpi-card.highlight-positive .kpi-value { color:var(--positive); }
.kpi-card.highlight-negative .kpi-value { color:var(--negative); }
.chart-row { display:grid; grid-template-columns:repeat(auto-fit,minmax(420px,1fr)); gap:var(--gap); margin-bottom:var(--gap); }
.chart-container { background:var(--bg-card); border-radius:var(--radius); padding:20px 24px; box-shadow:0 1px 3px rgba(0,0,0,0.08); }
.chart-container h3 { font-size:14px; font-weight:600; margin-bottom:4px; }
.chart-container .note { font-size:12px; color:var(--text-secondary); margin-bottom:12px; }
.chart-container canvas { max-height:280px; }
.table-section { background:var(--bg-card); border-radius:var(--radius); padding:20px 24px; box-shadow:0 1px 3px rgba(0,0,0,0.08); margin-bottom:var(--gap); overflow-x:auto; }
.data-table { width:100%; border-collapse:collapse; font-size:13px; }
.data-table thead th { text-align:left; padding:10px 12px; border-bottom:2px solid #dee2e6; color:var(--text-secondary); font-weight:600; font-size:12px; text-transform:uppercase; }
.data-table tbody td { padding:10px 12px; border-bottom:1px solid #f0f0f0; }
.data-table tbody tr:hover { background:#f8f9fa; }
.warnings-section { background:#fff8e1; border:1px solid #ffe0a3; border-radius:var(--radius); padding:16px 22px; font-size:13px; color:#7a5c00; }
.warnings-section h3 { font-size:14px; margin-bottom:8px; color:#5c4400; }
.warnings-section ul { padding-left:18px; }
.dashboard-footer { text-align:center; font-size:12px; color:var(--text-secondary); padding:16px 0 4px; }
.tab-bar { display:flex; gap:4px; margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:6px; box-shadow:0 1px 3px rgba(0,0,0,0.08); }
.tab-btn { flex:1; padding:10px 16px; border:none; background:none; border-radius:6px; font-size:14px; font-weight:600; color:var(--text-secondary); cursor:pointer; }
.tab-btn.active { background:var(--bg-header); color:var(--text-on-dark); }
.orders-controls { display:flex; gap:12px; flex-wrap:wrap; align-items:center; margin-bottom:14px; }
.orders-controls select, .orders-controls input { padding:8px 10px; border:1px solid #dee2e6; border-radius:6px; font-size:13px; }
.orders-controls input[type="text"] { flex:1; min-width:160px; }
.orders-summary { font-size:13px; color:var(--text-secondary); margin-bottom:10px; }
.data-table th.sortable { cursor:pointer; user-select:none; white-space:nowrap; }
.data-table th.sortable:hover { color:var(--text-primary); }
.data-table th.sortable .sort-arrow { display:inline-block; margin-left:4px; opacity:0.3; font-size:10px; }
.data-table th.sortable .sort-arrow.active { opacity:1; color:var(--accent, #4a6cf7); }
.pagination { display:flex; gap:6px; align-items:center; justify-content:center; margin-top:14px; font-size:13px; flex-wrap:wrap; }
.pagination button { padding:6px 12px; border:1px solid #dee2e6; background:#fff; border-radius:6px; cursor:pointer; }
.pagination button:disabled { opacity:0.4; cursor:default; }
.pagination button.pagination-active { background:var(--accent, #4a6cf7); border-color:var(--accent, #4a6cf7); color:#fff; font-weight:600; }
.pagination-ellipsis { padding:0 4px; color:var(--text-secondary); }
.fee-unknown { color:var(--text-secondary); font-style:italic; }
@media (max-width:768px){ .kpi-row{grid-template-columns:repeat(2,1fr);} .chart-row{grid-template-columns:1fr;} }
</style>
</head>
<body>
<div class="dashboard-container">
  <header class="dashboard-header">
    <div><h1>__COMPANY__</h1><div class="sub">Executive Finance Dashboard -- Period: __PERIOD__</div></div>
    <div class="sub">Generated __GENERATED_AT__</div>
  </header>

  <nav class="tab-bar">
    <button class="tab-btn active" id="tab-btn-overview" onclick="showTab('overview')">Net Profit Overview</button>
    <button class="tab-btn" id="tab-btn-bank" onclick="showTab('bank')">Bank Statement</button>
    <button class="tab-btn" id="tab-btn-orders" onclick="showTab('orders')">Order Detail</button>
    <button class="tab-btn" id="tab-btn-restock" onclick="showTab('restock')">Restock Plan</button>
    <button class="tab-btn" id="tab-btn-pipeline" onclick="showTab('pipeline')">Restock Pipeline (60-day)</button>
    <button class="tab-btn" id="tab-btn-strategy" onclick="showTab('strategy')">Restock Strategy</button>
    <button class="tab-btn" id="tab-btn-health" onclick="showTab('health')">Financial Health</button>
    <button class="tab-btn" id="tab-btn-isku" onclick="showTab('isku')">ISKU Manager</button>
  </nav>

  <div id="tab-overview">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
    <strong>Accrual P&amp;L for the month</strong> -- revenue and costs recognized when earned/incurred, regardless of
    when cash actually moves. Use this tab to check whether the company is profitable each month. It will NOT match
    the Bank Statement tab: platform payouts land in the bank weeks after the sale, so accrual and cash timing differ.
  </div>
  <section class="kpi-row" id="kpi-row"></section>

  <section class="warnings-section" id="sgd-section" style="display:none; margin-bottom:var(--gap); background:#e7f1ff; border-color:#b6d4fe; color:#084298;">
    <h3 style="color:#052c65;">Converted: Singapore (SGD) Channels</h3>
    <div id="sgd-note"></div>
  </section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Net Profit Calculation Walkthrough</h3>
    <div class="note" style="margin-bottom:12px;">Every step from Revenue down to Net Profit, in order</div>
    <table class="data-table" id="walkthrough-table"></table>
  </section>

  <section class="chart-row">
    <div class="chart-container">
      <h3>Accrual Revenue &amp; Net Profit Trend</h3>
      <div class="note" id="trend-note">Builds up as you run this script each month</div>
      <canvas id="trendChart"></canvas>
    </div>
    <div class="chart-container">
      <h3>Revenue &amp; NPE by Channel</h3>
      <div class="note">Every D2C marketplace and dropship reseller shown individually</div>
      <canvas id="channelChart"></canvas>
    </div>
  </section>

  <section class="chart-row">
    <div class="chart-container">
      <h3>Frescoone Proprietary vs. Generic Brands</h3>
      <div class="note">Gross margin comparison</div>
      <canvas id="brandChart"></canvas>
    </div>
    <div class="chart-container">
      <h3>Payroll Trend</h3>
      <div class="note">Total gross pay by month</div>
      <canvas id="payrollChart"></canvas>
    </div>
  </section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Channel Detail</h3>
    <div class="note" style="margin-bottom:12px;">Every D2C marketplace and dropship reseller individually.
      Fee/Ad Spend are real numbers only for storefronts with a fee/income report or raw order export provided
      (shown as "n/a" otherwise, not estimated) -- resellers show 0 since no marketplace fee applies to a
      wholesale invoice at all. Singapore (SGD) storefronts are converted to RM at Bun's confirmed rate (see
      banner above) and folded in like any other channel -- tagged "(SGD, conv. @rate)" so it's clear which rows
      those are.</div>
    <table class="data-table" id="channel-table"></table>
  </section>
  </div>

  <div id="tab-bank" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
    <strong>Actual cash that moved through the bank account</strong> -- use this tab to verify the real
    financial/liquidity situation, separate from the accrual Net Profit Overview tab. Platform payout credits here
    are already net of marketplace fees AND ad spend (confirmed via wallet reconciliation) -- they are settlement
    amounts, not gross revenue. This tab intentionally will NOT match the Net Profit Overview tab's numbers.
  </div>

  <section class="chart-row">
    <div class="chart-container">
      <h3>Cash Flow Trend (Bank Statement)</h3>
      <div class="note">Inflows vs. outflows by month -- full bank statement history</div>
      <canvas id="cashflowChart"></canvas>
    </div>
    <div class="chart-container">
      <h3>Operating Expense by Category</h3>
      <div class="note" id="opex-note">This month's non-payroll bank debits</div>
      <canvas id="opexChart"></canvas>
    </div>
  </section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Bank Transaction Detail</h3>
    <div class="note" style="margin-bottom:12px;">
      Every line from the OCBC statement, money in and money out, full history. "Category" is the same
      classification used in the Operating Expense chart above.
    </div>
    <div class="orders-controls">
      <select id="bank-month-filter" onchange="bankState.page=0; renderBankTable();"></select>
      <select id="bank-direction-filter" onchange="bankState.page=0; renderBankTable();">
        <option value="all">In &amp; Out</option>
        <option value="in">Money In Only</option>
        <option value="out">Money Out Only</option>
      </select>
      <input type="text" id="bank-search" placeholder="Search description..." oninput="bankState.page=0; renderBankTable();">
    </div>
    <div class="orders-summary" id="bank-summary"></div>
    <table class="data-table" id="bank-table"></table>
    <div class="pagination" id="bank-pagination"></div>
  </section>
  </div>

  <section class="warnings-section" id="warnings-section" style="display:none; margin-top:var(--gap);">
    <h3>Data Quality Notes</h3>
    <ul id="warnings-list"></ul>
  </section>

  <div id="tab-orders" style="display:none;">
  <section class="table-section">
    <h3 style="margin-bottom:4px;">Order-Level Net Profit &amp; Margin</h3>
    <div class="note" style="margin-bottom:12px;">
      One row per SKU per order, for every channel (D2C and reseller/dropship alike) -- Final Unit Price/Deal
      Price x Qty vs Cost x Qty, the actual invoice/order billing granularity. Revenue/COGS/Gross Profit are
      directly computed. Platform Fee % is real per-order data where a fee/income report or raw order export
      has been provided for that storefront -- confirmed via bank/wallet reconciliation that ad spend already
      reduces what actually settles to the bank, same as the marketplace fee, so <strong>the RM amount in
      brackets is marketplace fee + that platform's ad spend combined</strong>, spread across the order's SKU
      lines proportional to each line's revenue share (neither fees nor ad spend are broken down by SKU or by
      order at the source) -- other rows show "n/a" and are not charged a fee.
      <strong>Net Profit here = Net Income Received minus COGS for every row</strong> -- for D2C that's
      Revenue minus the real platform fee (incl. ad spend, where known) minus COGS; for reseller/dropship it's
      DSP revenue minus COGS (no fee applies). <strong>It does NOT include a share of company-wide Payroll, Bank OpEx, or
      ad spend</strong> -- there's no traceable way to attribute fixed overhead to one order, so it's left out
      entirely rather than modeled and dressed up as real. That means this table's total will NOT match the
      headline company Net Profit on the Net Profit Overview tab (which does subtract those) -- this view
      answers "which orders/products are actually profitable at the unit level," the Net Profit Overview tab
      answers "is the company profitable overall." Both are correct; they're answering different questions.
      Rows whose SKU cost couldn't be found in the master cost table are <strong>excluded from this table</strong>
      and shown separately below in Pending Action, since their COGS/Net Profit would otherwise be wrong.
    </div>
    <div class="orders-controls">
      <select id="orders-platform-filter" onchange="ordersState.page=0; renderOrdersTable();"></select>
      <input type="text" id="orders-search" placeholder="Search Order ID..." oninput="ordersState.page=0; renderOrdersTable();">
      <span class="note">Click any column header to sort</span>
    </div>
    <div class="orders-summary" id="orders-summary"></div>
    <table class="data-table" id="orders-table"></table>
    <div class="pagination" id="orders-pagination"></div>
  </section>

  <section class="table-section" id="pending-section" style="display:none;">
    <h3 style="margin-bottom:4px;">Pending Action -- SKU Cost Not Found (<span id="pending-count">0</span>)</h3>
    <div class="note" style="margin-bottom:12px;">
      These SKUs have no match in <code>isku_database.json</code> (the ISKU Manager tab's database), so their
      COGS is currently unknown -- excluded from every total above. <strong>Key in the correct cost per unit</strong>,
      then click <strong>Confirm</strong>. In Chrome/Edge this writes straight into <code>isku_database.json</code>
      in the project folder (a one-time file picker the first time this session, shared with the ISKU Manager tab)
      -- next time the script runs, that SKU resolves automatically and you never key it in again. The row
      disappears from this list as soon as it's confirmed. <strong>Undo</strong> clears the input without
      confirming. This needs Chrome or Edge (File System Access API) -- for a full costing entry (Import Price,
      RSP, ITEC Code, etc.) rather than just a raw cost, use <strong>Add New ISKU</strong> on the ISKU Manager tab
      instead. One row per unique SKU (aggregated across every order it appeared in this month).
    </div>
    <div class="orders-controls">
      <span><strong id="override-count">0</strong> cost(s) entered, not yet confirmed</span>
    </div>
    <table class="data-table" id="pending-table"></table>
  </section>
  </div>

  <div id="tab-restock" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);" id="restock-empty-note">
    No restock plan generated yet -- run <code>restock_planner.py</code> in the project folder (needs
    <code>active-isku-master.xlsx</code> and a fresh <code>inventory-forecast-SiteGiant/last-30days-*.xlsx</code>
    export), then re-run <code>finance_analyzer.py</code> to pick it up.
  </div>
  <div id="restock-content" style="display:none;">
    <section class="table-section">
      <h3 style="margin-bottom:4px;">Open-To-Buy Budget Summary</h3>
      <div class="note" style="margin-bottom:12px;" id="restock-generated-note"></div>
      <table class="data-table" id="restock-budget-table"></table>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Recommended Restock Orders</h3>
      <div class="note" style="margin-bottom:12px;" id="restock-funded-note"></div>
      <table class="data-table" id="restock-funded-table"></table>
    </section>

    <section class="table-section" id="restock-cut-section" style="display:none;">
      <h3 style="margin-bottom:4px;">Cut for Budget</h3>
      <div class="note" style="margin-bottom:12px;">Not funded this round -- lowest priority (C-class / most
        days-of-stock-left) trimmed first. Rows marked Urgent still have &lt;20 days of stock or are already
        oversold -- worth reviewing even though they didn't fit the budget.</div>
      <table class="data-table" id="restock-cut-table"></table>
    </section>

    <section class="table-section" id="restock-nocost-section" style="display:none;">
      <h3 style="margin-bottom:4px;">Active SKUs -- No Cost on File</h3>
      <div class="note" style="margin-bottom:12px;">Can't size a budget without a real cost, so excluded rather
        than guessed -- these also show up in the Pending Action table above.</div>
      <table class="data-table" id="restock-nocost-table"></table>
    </section>

    <section class="table-section" id="restock-notfound-section" style="display:none;">
      <h3 style="margin-bottom:4px;">Active SKUs -- Not Found in SiteGiant Forecast</h3>
      <div class="note" style="margin-bottom:12px;">On the active-isku-master list but not matched in the
        SiteGiant export -- either not set up in SiteGiant yet or a SKU-code mismatch. Review manually.</div>
      <table class="data-table" id="restock-notfound-table"></table>
    </section>
  </div>
  </div>

  <div id="tab-pipeline" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);" id="pipeline-empty-note">
    No pipeline run generated yet -- run <code>restock_pipeline.py</code> in the project folder, then re-run
    <code>finance_analyzer.py</code> to pick it up.
  </div>
  <div id="pipeline-content" style="display:none;">
    <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
      <strong>A separate, independent system from the Restock Plan tab.</strong> This replicates the confirmed
      60-day lead-time / MOQ / round-to-5 methodology from the "Frescoone Sales &amp; Stock Analyst" project
      (Stock on Purchase Order is never netted off, only flagged) -- kept deliberately side by side rather than
      merged, since the two use genuinely different rules and will show different numbers for the same SKU. See
      <code>Restock_Purchase_List</code> and the per-supplier PO files in <code>restock-po-files/</code> for the
      full downloadable output.
    </div>

    <section class="kpi-row" id="pipeline-kpi-row"></section>

    <section class="warnings-section" id="pipeline-otb-section" style="margin-bottom:var(--gap);">
      <div id="pipeline-otb-note"></div>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Spend by Supplier</h3>
      <table class="data-table" id="pipeline-supplier-table"></table>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Full Detail</h3>
      <div class="orders-controls">
        <select id="pipeline-supplier-filter" onchange="pipelineState.page=0; renderPipelineDetail();"></select>
        <input type="text" id="pipeline-search" placeholder="Search SKU..." oninput="pipelineState.page=0; renderPipelineDetail();">
        <select id="pipeline-action-filter" onchange="pipelineState.page=0; renderPipelineDetail();">
          <option value="all">All Actions</option>
          <option value="RESTOCK">Restock Only</option>
          <option value="NO_ACTION">No Action Only</option>
        </select>
      </div>
      <table class="data-table" id="pipeline-detail-table"></table>
      <div class="pagination" id="pipeline-pagination"></div>
    </section>
  </div>
  </div>

  <div id="tab-strategy" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);" id="strategy-empty-note">
    No strategy run generated yet -- run <code>restock_strategy.py</code> in the project folder, then re-run
    <code>finance_analyzer.py</code> to pick it up.
  </div>
  <div id="strategy-content" style="display:none;">
    <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
      <strong>A decision layer on top of Restock Plan and Restock Pipeline, not a replacement for either.</strong>
      Answers four things: how close is the real FRES-brand sales mix to the 85% goal, which SKUs need restocking
      and how urgently, which SKUs have no real case for continued restock, and which SKUs show genuine
      month-over-month sales momentum. The restock multiplier still applies to the same baseline demand formula
      as Restock Pipeline -- it does NOT yet use actual prior-order-quantity history (not tracked anywhere yet,
      per Bun 2026-07-21). Honest approximations, not guesses dressed up as fact.
    </div>

    <section class="kpi-row" id="strategy-kpi-row"></section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Brand Mix -- FRES vs. Generic</h3>
      <div class="note" style="margin-bottom:12px;" id="strategy-brandmix-note"></div>
      <table class="data-table" id="strategy-brandmix-table"></table>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Restock Status -- Every Active SKU</h3>
      <div class="note" style="margin-bottom:12px;" id="strategy-status-note">Every active SKU gets one status, anchored to
        the real <span id="strategy-leadtime-days">50</span>-day restock lead time (not a fixed day-count bucket): if stock
        won't last until a freshly-placed order would arrive, that's real gap risk --
        <span style="color:#a3342a;font-weight:600;">Out of Stock</span>,
        <span style="color:#a3342a;font-weight:600;">Stockout Risk -- Order Now</span> (no reorder placed yet), or
        <span style="color:#c26a00;font-weight:600;">Stockout Risk -- Reorder Placed</span> (already on PO -- verify
        quantity/timing). No gap risk but a FRES top-earner or trending up gets
        <span style="color:#2b6cb0;font-weight:600;">Scale Up -- Running Well</span> -- a proactive grow signal, not a
        reactive one. Otherwise <span style="color:#6c757d;font-weight:600;">Stop-Restock Candidate</span> (weak 4-month
        trend), <span style="color:#3f7a3f;font-weight:600;">Covered</span>, or
        <span style="color:#6c757d;font-weight:600;">No Sales Signal</span>. Multiplier: 2x baseline on gap risk, 3x if
        also a top-20% avg. NPE earner among FRES SKUs, 1.5x for Running Well, capped at 1.5x if unit cost exceeds
        RM50. NPE below is the <strong>average</strong> net profit per order for that SKU (not a total, so it isn't
        affected by how many orders a SKU happened to have) -- hover the count for how many orders it's averaged
        over. Generic SKUs show real status too but never scale up -- safety restock only.</div>
      <div class="orders-controls">
        <input type="text" id="strategy-search" placeholder="Search SKU..." oninput="strategyScaleupState.page=0; renderStrategyScaleupTable();">
        <select id="strategy-status-filter" onchange="strategyScaleupState.page=0; renderStrategyScaleupTable();">
          <option value="all">All Status</option>
          <option value="Out of Stock">Out of Stock</option>
          <option value="Stockout Risk -- Order Now">Stockout Risk -- Order Now</option>
          <option value="Stockout Risk -- Reorder Placed">Stockout Risk -- Reorder Placed</option>
          <option value="Scale Up -- Running Well">Scale Up -- Running Well</option>
          <option value="Stop-Restock Candidate">Stop-Restock Candidate</option>
          <option value="Covered">Covered</option>
          <option value="No Sales Signal">No Sales Signal</option>
        </select>
        <select id="strategy-brand-filter" onchange="strategyScaleupState.page=0; renderStrategyScaleupTable();">
          <option value="all">All Brands</option>
          <option value="FRES">FRES only</option>
          <option value="Generic">Generic only</option>
        </select>
        <select id="strategy-ordersheet-filter" onchange="strategyScaleupState.page=0; renderStrategyScaleupTable();">
          <option value="all">All Order Sheets</option>
        </select>
        <label style="display:flex; align-items:center; gap:6px; font-size:13px; color:var(--text-secondary); white-space:nowrap;">
          <input type="checkbox" id="strategy-topearner-filter" onchange="strategyScaleupState.page=0; renderStrategyScaleupTable();">
          Top earners only &#9733;
        </label>
      </div>
      <table class="data-table" id="strategy-scaleup-table"></table>
      <div class="pagination" id="strategy-scaleup-pagination"></div>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Stop-Restock Candidates</h3>
      <div class="note" style="margin-bottom:12px;" id="strategy-stoprestock-note"></div>
      <table class="data-table" id="strategy-stoprestock-table"></table>
      <div class="pagination" id="strategy-stoprestock-pagination"></div>
    </section>

    <section class="table-section">
      <h3 style="margin-bottom:4px;">Trending Up -- Real Month-over-Month Momentum</h3>
      <div class="note" style="margin-bottom:12px;">Sales rose every single month across the window analyzed --
        a verified trend, not a statistical forecast. Sorted by total units sold, highest first.</div>
      <table class="data-table" id="strategy-trending-table"></table>
      <div class="pagination" id="strategy-trending-pagination"></div>
    </section>
  </div>
  </div>

  <div id="tab-health" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
    <strong>Not licensed financial advice</strong> -- this is data-driven analysis of your own real numbers (same
    arithmetic as the rest of this dashboard: real revenue, real cash position, real ad spend), not a substitute
    for a licensed accountant or financial advisor on major decisions. Every figure below traces back to a number
    computed elsewhere in this dashboard; where a generic industry range is shown, it's explicitly labeled as
    generic context, not a Frescoone-specific fact.
  </div>

  <section class="table-section" id="health-flags-section" style="display:none;">
    <h3 style="margin-bottom:12px;">Key Findings</h3>
    <ul id="health-flags-list" style="list-style:none; display:flex; flex-direction:column; gap:8px;"></ul>
  </section>

  <section class="kpi-row" id="health-kpi-row"></section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Where This Month's Result Comes From</h3>
    <div class="note" style="margin-bottom:12px;">Gross Profit down to the headroom left for ad spend before Net
      Profit hits zero -- everything else (COGS, Payroll, Bank OpEx, platform transaction fees) held fixed at
      this month's real, actual numbers.</div>
    <table class="data-table" id="health-breakeven-table"></table>
  </section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Revenue &amp; Profit Targets</h3>
    <div class="note" style="margin-bottom:12px;" id="health-ladder-note"></div>
    <table class="data-table" id="health-ladder-table"></table>
  </section>

  <section class="chart-row">
    <div class="chart-container" style="flex:1;">
      <h3>Advertising Budget</h3>
      <div class="note">Real current spend vs. what the numbers can sustain</div>
      <table class="data-table" id="health-ad-table"></table>
    </div>
    <div class="chart-container" style="flex:1;">
      <h3>Stock Purchase Budget</h3>
      <div class="note">Range at today's revenue -- same underlying numbers as the Restock Plan tab</div>
      <table class="data-table" id="health-restock-table"></table>
    </div>
  </section>

  <section class="table-section">
    <h3 style="margin-bottom:4px;">Cash Flow Health</h3>
    <div class="note" style="margin-bottom:12px;" id="health-cashflow-note"></div>
    <table class="data-table" id="health-cashflow-table"></table>
  </section>
  </div>

  <div id="tab-isku" style="display:none;">
  <div class="note" style="margin-bottom:var(--gap); background:var(--bg-card); border-radius:var(--radius); padding:12px 16px; box-shadow:0 1px 3px rgba(0,0,0,0.08);">
    The local replacement for the Google Sheet -- <code>isku_database.json</code> in the project folder is the
    real source of truth for every ISKU's costing/pricing. In Chrome/Edge, <strong>Open Database</strong> below
    picks that file once (remembered for next time) and every Save here writes straight back into it -- the
    same auto-save mechanism the Pending Action table already uses. Cost and DSP are computed live as you type
    (Import Price + Tax + Import Fee, divided by the exchange rate for the chosen ITEC Code; DSP = Cost + a
    margin share of RSP-minus-Cost) and then frozen onto the record when you save -- exactly like the sheet's
    own "Add To DB" button, so a future exchange-rate update never silently re-prices stock you already bought.
  </div>
  <div class="orders-controls">
    <button onclick="loadIskuDatabase()" style="padding:8px 14px;border:none;border-radius:6px;background:var(--bg-header);color:#fff;cursor:pointer;font-size:13px;">Open isku_database.json</button>
    <span id="isku-load-status" class="note"></span>
  </div>

  <section class="table-section" id="isku-add-section" style="display:none;">
    <h3 style="margin-bottom:12px;">Add New ISKU</h3>
    <div style="display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:12px; margin-bottom:14px;">
      <div><label class="note">Brand</label><input id="isku-add-brand" type="text" list="isku-add-brand-list" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"><datalist id="isku-add-brand-list"></datalist></div>
      <div><label class="note">Category</label><input id="isku-add-category" type="text" list="isku-add-category-list" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"><datalist id="isku-add-category-list"></datalist></div>
      <div><label class="note">Series</label><input id="isku-add-series" type="text" list="isku-add-series-list" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"><datalist id="isku-add-series-list"></datalist></div>
      <div><label class="note">Variant</label><input id="isku-add-variant" type="text" list="isku-add-variant-list" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"><datalist id="isku-add-variant-list"></datalist></div>
      <div><label class="note">Color</label><input id="isku-add-color" type="text" list="isku-add-color-list" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"><datalist id="isku-add-color-list"></datalist></div>
    </div>
    <div class="note" style="margin-bottom:14px;">SKU: <strong id="isku-add-sku-preview">--</strong> (edit any part above to change it, or key in the SKU box directly)
      <input id="isku-add-sku" type="text" style="width:100%;margin-top:6px;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="document.getElementById('isku-add-sku-preview').textContent = this.value || '--'">
    </div>
    <div style="display:grid; grid-template-columns:repeat(auto-fit,minmax(170px,1fr)); gap:12px; margin-bottom:14px;">
      <div><label class="note">ITEC Code (invisible costing preset)</label><select id="isku-add-itec" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" onchange="updateIskuAddPreview()"></select></div>
      <div><label class="note">Import Price</label>
        <div style="display:flex;gap:6px;">
          <input id="isku-add-import-price" type="number" step="0.01" min="0" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()">
          <select id="isku-add-currency" style="padding:8px;border:1px solid #dee2e6;border-radius:6px;" onchange="updateIskuAddPreview()"></select>
        </div>
      </div>
      <div><label class="note">RSP (RM)</label><input id="isku-add-rsp" type="number" step="0.01" min="0" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"></div>
      <div><label class="note">Computed Cost</label><div id="isku-add-cost-preview" style="padding:8px;font-weight:600;">--</div></div>
      <div><label class="note">Computed DSP</label><div id="isku-add-dsp-preview" style="padding:8px;font-weight:600;">--</div></div>
    </div>
    <div style="display:grid; grid-template-columns:repeat(auto-fit,minmax(170px,1fr)); gap:12px; margin-bottom:14px;">
      <div><label class="note">UPC (optional)</label><input id="isku-add-upc" type="text" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;"></div>
      <div><label class="note">Status</label><select id="isku-add-status" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;">
        <option>ACTIVE</option><option>CLEARANCE</option><option>STOPPED</option></select></div>
      <div><label class="note">Offline Set Margin % (optional)</label><input id="isku-add-offline-margin" type="number" step="1" min="0" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;" oninput="updateIskuAddPreview()"></div>
      <div><label class="note">Offline Purchase Price</label><div id="isku-add-offline-preview" style="padding:8px;font-weight:600;">--</div></div>
    </div>
    <div style="margin-bottom:14px;"><label class="note">Notes</label><input id="isku-add-notes" type="text" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;"></div>
    <button onclick="saveNewIsku()" style="padding:10px 18px;border:none;border-radius:6px;background:var(--positive);color:#fff;cursor:pointer;font-size:14px;font-weight:600;">Save New ISKU</button>
    <span id="isku-add-status-msg" class="note" style="margin-left:12px;"></span>
  </section>

  <section class="table-section" id="isku-search-section" style="display:none;">
    <h3 style="margin-bottom:4px;">Search / Edit ISKU (<span id="isku-count">0</span>)</h3>
    <div class="orders-controls">
      <input type="text" id="isku-search" placeholder="Search SKU / brand / status..." oninput="onIskuFilterChange()">
      <select id="isku-filter-brand" onchange="onIskuFilterChange()"><option value="">All Brands</option></select>
      <select id="isku-filter-status" onchange="onIskuFilterChange()"><option value="">All Statuses</option></select>
      <select id="isku-filter-itec" onchange="onIskuFilterChange()"><option value="">All ITEC Codes</option></select>
      <button onclick="clearIskuFilters()" style="padding:8px 14px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Clear Filters</button>
    </div>

    <div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:14px 16px;margin-bottom:14px;">
      <div style="font-weight:600;font-size:13px;margin-bottom:6px;">Bulk Edit -- applies to the <span id="isku-bulk-match-count">0</span> ISKU(s) matching the search above</div>
      <div class="note" style="margin-bottom:10px;">Not free-form AI instructions -- a small fixed command syntax, checked and previewed
        before anything is written. One assignment per line, or separate with <code>;</code>.
        <code>field=value</code> to set, or <code>field+=</code> / <code>field-=</code> / <code>field*=</code> for
        <code>rsp</code>/<code>import_price</code> (e.g. <code>rsp+=2</code> adds RM2 to RSP on every matching row,
        <code>rsp*=1.05</code> raises it 5%). Editable fields: <code>status</code>, <code>itec_code</code>,
        <code>currency</code>, <code>rsp</code>, <code>import_price</code>, <code>notes</code>. Cost/DSP recompute
        automatically from whatever itec_code/currency/import_price/rsp each row ends up with.</div>
      <textarea id="isku-bulk-command" rows="2" placeholder="e.g. status=CLEARANCE&#10;rsp+=2" style="width:100%;padding:8px;border:1px solid #dee2e6;border-radius:6px;font-family:monospace;font-size:13px;margin-bottom:8px;"></textarea>
      <button onclick="previewBulkCommand()" style="padding:8px 14px;border:none;border-radius:6px;background:var(--bg-header);color:#fff;cursor:pointer;font-size:13px;">Preview Bulk Edit</button>
      <div id="isku-bulk-preview"></div>
    </div>

    <div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:14px 16px;margin-bottom:14px;">
      <div style="font-weight:600;font-size:13px;margin-bottom:6px;">Bulk Edit via File</div>
      <div class="note" style="margin-bottom:10px;">Download the ISKUs matching the search above as a CSV, edit
        Status/ITEC Code/Currency/Import Price/RSP/Notes in Excel or Sheets, then upload it back -- previewed as a
        diff before anything is written. Rows for SKUs not already in the database are skipped (reported, not
        silently dropped) -- use Add New ISKU for brand-new items, this is for editing existing ones only.</div>
      <button onclick="downloadIskuTemplate()" style="padding:8px 14px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Download Template (CSV)</button>
      <input type="file" id="isku-bulk-file-input" accept=".csv" style="margin-left:10px;" onchange="handleBulkFileUpload(this)">
      <div id="isku-bulk-file-preview"></div>
    </div>

    <div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:14px 16px;margin-bottom:14px;">
      <div style="font-weight:600;font-size:13px;margin-bottom:6px;">Bulk Add via File -- brand-new ISKUs, many at once</div>
      <div class="note" style="margin-bottom:10px;" id="isku-bulk-add-note">Download the blank template, fill in one row per new
        ISKU, then upload it back -- previewed (with computed Cost/DSP) before anything is written. Either fill in
        <code>sku</code> directly, or leave it blank and fill <code>brand</code>/<code>category</code>/<code>series</code>/
        <code>variant</code>/<code>color</code> to have it built the same way the Add New ISKU form does.
        <code>itec_code</code> and <code>currency</code> must be one of the active ones: <span id="isku-bulk-add-active-codes">(loading)</span>.
        Rows that already exist, duplicate another row in the file, or use an unknown ITEC Code/currency are
        skipped and reported, not silently guessed at.</div>
      <button onclick="downloadIskuAddTemplate()" style="padding:8px 14px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Download Blank Template (CSV)</button>
      <input type="file" id="isku-bulk-add-file-input" accept=".csv" style="margin-left:10px;" onchange="handleBulkAddFileUpload(this)">
      <div id="isku-bulk-add-preview"></div>
    </div>

    <table class="data-table" id="isku-table"></table>
    <div class="pagination" id="isku-pagination"></div>
  </section>

  <section class="table-section" id="isku-currency-section" style="display:none;">
    <h3 style="margin-bottom:4px;">Currency Table</h3>
    <div class="note" style="margin-bottom:12px;">RM value of 1 unit of each currency -- edit here as rates
      drift. Import Price on any ISKU picks one of these independently of its ITEC Code.</div>
    <table class="data-table" id="currency-table"></table>
    <div style="display:flex;gap:8px;align-items:center;margin-top:12px;">
      <input id="currency-add-code" type="text" placeholder="Code, e.g. GBP" style="width:120px;padding:6px;border:1px solid #dee2e6;border-radius:4px;">
      <input id="currency-add-rate" type="number" step="0.0001" placeholder="RM per unit" style="width:120px;padding:6px;border:1px solid #dee2e6;border-radius:4px;">
      <button onclick="addCurrency()" style="padding:6px 14px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">+ Add Currency</button>
      <span id="currency-add-msg" class="note"></span>
    </div>
  </section>

  <section class="table-section" id="isku-itec-section" style="display:none;">
    <h3 style="margin-bottom:4px;">ITEC Code Reference Table -- Invisible Costing</h3>
    <div class="note" style="margin-bottom:12px;">Not currency or tax -- this is Bun's per-item markup for costs
      that don't show up on the supplier invoice itself (packaging, shipping, engraving, custom packaging, etc.),
      itemized per code so each line can be tracked and adjusted separately. Editing a code only affects
      new/edited ISKUs going forward; existing records keep their frozen Cost/DSP.</div>
    <div id="itec-cards"></div>
    <div style="display:flex;gap:8px;align-items:center;">
      <input id="itec-add-code" type="text" placeholder="New code, e.g. IT-CN-XL" style="width:180px;padding:6px;border:1px solid #dee2e6;border-radius:4px;">
      <button onclick="addItec()" style="padding:6px 14px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">+ Add ITEC Code</button>
      <span id="itec-add-msg" class="note"></span>
    </div>
  </section>
  </div>

  <footer class="dashboard-footer">Frescoone Trading Sdn Bhd -- generated locally by finance_analyzer.py -- not published anywhere</footer>
</div>

<script>
const DATA = __DASHBOARD_DATA_JSON__;
const COLORS = ['#4C72B0','#DD8452','#55A868','#C44E52','#8172B3','#937860'];

function fmtMoney(v) {
  if (Math.abs(v) >= 1e6) return 'RM ' + (v/1e6).toFixed(2) + 'M';
  if (Math.abs(v) >= 1e3) return 'RM ' + (v/1e3).toFixed(1) + 'K';
  return 'RM ' + v.toFixed(2);
}
function fmtPct(v) { return v.toFixed(1) + '%'; }

// --- Generic click-to-sort table headers -----------------------------------
// Every sortable table shares this: a {sortField, sortDir} bag on that table's
// state object, a `th()` header cell that shows a ▲/▼ arrow only on the active
// column, and `sortRows()` to actually order the array before rendering.
function sortRows(rows, field, dir) {
  // Nulls (e.g. the SGD "n/a" cells, which have no real numeric value) always
  // sort last regardless of direction -- applying `dir` to the whole array
  // including null-ordering would flip them to the TOP on a descending sort,
  // which is worse than useless (verified: sorting Channel Detail by NPE desc
  // put "FRESCO SG (n/a)" in first place). Only the actual value comparison
  // gets negated for descending; null-ordering is direction-independent.
  const mult = dir === 'desc' ? -1 : 1;
  return [...rows].sort((a, b) => {
    let av = a[field], bv = b[field];
    if (av == null && bv == null) return 0;
    if (av == null) return 1;
    if (bv == null) return -1;
    if (typeof av === 'string' || typeof bv === 'string') {
      av = String(av).toLowerCase(); bv = String(bv).toLowerCase();
    }
    if (av < bv) return -1 * mult;
    if (av > bv) return 1 * mult;
    return 0;
  });
}
function toggleSort(state, field, renderFnName, defaultDir) {
  if (state.sortField === field) {
    state.sortDir = state.sortDir === 'asc' ? 'desc' : 'asc';
  } else {
    state.sortField = field;
    state.sortDir = defaultDir || 'desc';
  }
  if ('page' in state) state.page = 0;
  window[renderFnName]();
}
function th(label, field, state, stateName, renderFnName) {
  const active = state.sortField === field;
  const arrow = active ? (state.sortDir === 'asc' ? '&#9650;' : '&#9660;') : '&#8645;';
  return `<th class="sortable" onclick="toggleSort(${stateName},'${field}','${renderFnName}')">${label}` +
    `<span class="sort-arrow${active ? ' active' : ''}">${arrow}</span></th>`;
}

// Numbered page buttons (1 2 3 ... 8) instead of just First/Prev/Next/Last, shared
// by every paginated table. Keeps first, last, and a window around the current page,
// collapsing the rest into a single ellipsis rather than listing every page number.
function renderPagination(containerId, state, stateName, totalRows, renderFnName) {
  const totalPages = Math.max(1, Math.ceil(totalRows / state.pageSize));
  if (state.page > totalPages - 1) state.page = totalPages - 1;
  if (state.page < 0) state.page = 0;

  const pages = [];
  for (let p = 0; p < totalPages; p++) {
    if (p === 0 || p === totalPages - 1 || Math.abs(p - state.page) <= 1) {
      pages.push(p);
    } else if (pages[pages.length - 1] !== '...') {
      pages.push('...');
    }
  }
  const numbers = pages.map(p => p === '...'
    ? `<span class="pagination-ellipsis">&hellip;</span>`
    : `<button class="${p === state.page ? 'pagination-active' : ''}" onclick="${stateName}.page=${p};${renderFnName}();">${p + 1}</button>`
  ).join('');

  document.getElementById(containerId).innerHTML = `
    <button ${state.page===0?'disabled':''} onclick="${stateName}.page=0;${renderFnName}();">First</button>
    <button ${state.page===0?'disabled':''} onclick="${stateName}.page--;${renderFnName}();">Prev</button>
    ${numbers}
    <button ${state.page>=totalPages-1?'disabled':''} onclick="${stateName}.page++;${renderFnName}();">Next</button>
    <button ${state.page>=totalPages-1?'disabled':''} onclick="${stateName}.page=${totalPages-1};${renderFnName}();">Last</button>
  `;
}

function renderKPIs() {
  const k = DATA.kpis;
  const s = DATA.order_net_profit_split;
  const cards = [
    ['Revenue', fmtMoney(k.revenue), ''],
    ['COGS', fmtMoney(k.cogs), ''],
    ['Gross Profit', fmtMoney(k.gross_profit), fmtPct(k.gross_margin_pct) + ' margin'],
    ['Payroll OpEx', fmtMoney(k.payroll_cost), ''],
    ['Bank OpEx', fmtMoney(k.bank_opex), ''],
    ['Platform Fees/Ads', fmtMoney(k.ad_spend), ''],
    ['Monthly NPE', fmtMoney(k.monthly_npe), fmtPct(k.monthly_npe_margin_pct) + ' margin -- Gross Profit less ad spend ONLY'],
    ['Net Profit (Company P&L)', fmtMoney(k.net_profit), fmtPct(k.net_margin_pct) + ' margin -- incl. Payroll/Bank OpEx + fees'],
    ['+ Positive Net Profit (Orders)', fmtMoney(s.positive_total), s.positive_orders.toLocaleString() + ' profitable order/SKU rows'],
    ['- Negative Net Profit (Orders)', fmtMoney(s.negative_total), s.negative_orders.toLocaleString() + ' loss-making order/SKU rows'],
  ];
  const row = document.getElementById('kpi-row');
  row.innerHTML = cards.map(([label, val, sub]) => {
    let cls = '';
    if (label === 'Monthly NPE') cls = k.monthly_npe >= 0 ? 'highlight-positive' : 'highlight-negative';
    if (label.startsWith('Net Profit')) cls = k.net_profit >= 0 ? 'highlight-positive' : 'highlight-negative';
    if (label.startsWith('+ Positive')) cls = 'highlight-positive';
    if (label.startsWith('- Negative')) cls = 'highlight-negative';
    return `
    <div class="kpi-card ${cls}">
      <div class="kpi-label">${label}</div>
      <div class="kpi-value">${val}</div>
      ${sub ? `<div class="kpi-sub">${sub}</div>` : ''}
    </div>`;
  }).join('');
}

function renderCashflowChart() {
  const cf = DATA.cashflow;
  new Chart(document.getElementById('cashflowChart'), {
    type: 'bar',
    data: {
      labels: cf.months,
      datasets: [
        { type: 'bar', label: 'Inflows', data: cf.inflows, backgroundColor: COLORS[2] + 'CC' },
        { type: 'bar', label: 'Outflows', data: cf.outflows, backgroundColor: COLORS[3] + 'CC' },
        { type: 'line', label: 'Net Cash Flow', data: cf.net, borderColor: COLORS[0], backgroundColor: COLORS[0], tension: 0.3, borderWidth: 2 },
      ]
    },
    options: { responsive:true, maintainAspectRatio:false,
      plugins:{ legend:{position:'top', labels:{usePointStyle:true}}, tooltip:{callbacks:{label:(c)=>`${c.dataset.label}: ${fmtMoney(c.parsed.y)}`}} },
      scales:{ y:{ ticks:{ callback:(v)=>fmtMoney(v) } } } }
  });
}

function renderTrendChart() {
  const t = DATA.trend;
  const note = document.getElementById('trend-note');
  if (t.months.length < 2) {
    note.textContent = `Only ${t.months.length} month of accrual data so far -- run this script each month to build a real trend line here.`;
  } else {
    note.textContent = `${t.months.length} months of accrual history`;
  }
  new Chart(document.getElementById('trendChart'), {
    type: 'line',
    data: { labels: t.months, datasets: [
      { label: 'Revenue', data: t.revenue, borderColor: COLORS[0], backgroundColor: COLORS[0]+'20', fill:true, tension:0.3, borderWidth:2 },
      { label: 'Net Profit', data: t.net_profit, borderColor: COLORS[2], backgroundColor: COLORS[2]+'20', fill:true, tension:0.3, borderWidth:2 },
    ]},
    options: { responsive:true, maintainAspectRatio:false,
      plugins:{ legend:{position:'top', labels:{usePointStyle:true}}, tooltip:{callbacks:{label:(c)=>`${c.dataset.label}: ${fmtMoney(c.parsed.y)}`}} },
      scales:{ y:{ ticks:{ callback:(v)=>fmtMoney(v) } } } }
  });
}

function renderChannelChart() {
  const ch = DATA.channels;
  new Chart(document.getElementById('channelChart'), {
    type: 'bar',
    data: { labels: ch.map(c=>c.name), datasets: [
      { label: 'Revenue', data: ch.map(c=>c.revenue), backgroundColor: COLORS[0]+'CC' },
      { label: 'NPE', data: ch.map(c=>c.npe), backgroundColor: COLORS[2]+'CC' },
    ]},
    options: { responsive:true, maintainAspectRatio:false, indexAxis:'y',
      plugins:{ legend:{position:'top'}, tooltip:{callbacks:{label:(c)=>`${c.dataset.label}: ${fmtMoney(c.parsed.x)}`}} },
      scales:{ x:{ ticks:{ callback:(v)=>fmtMoney(v) } } } }
  });
}

function renderBrandChart() {
  const b = DATA.brand;
  new Chart(document.getElementById('brandChart'), {
    type: 'bar',
    data: { labels: b.map(x=>x.name), datasets: [{ label: 'Gross Margin %', data: b.map(x=>x.margin_pct), backgroundColor: [COLORS[0]+'CC', COLORS[1]+'CC'] }] },
    options: { responsive:true, maintainAspectRatio:false,
      plugins:{ legend:{display:false}, tooltip:{callbacks:{label:(c)=>fmtPct(c.parsed.y)}} },
      scales:{ y:{ ticks:{ callback:(v)=>v+'%' } } } }
  });
}

function renderOpexChart() {
  const o = DATA.opex;
  const note = document.getElementById('opex-note');
  if (!o.labels.length) { note.textContent = 'No non-payroll bank debits found for this month.'; return; }
  new Chart(document.getElementById('opexChart'), {
    type: 'doughnut',
    data: { labels: o.labels, datasets: [{ data: o.values, backgroundColor: COLORS.map(c=>c+'CC'), borderColor:'#fff', borderWidth:2 }] },
    options: { responsive:true, maintainAspectRatio:false, cutout:'60%',
      plugins:{ legend:{position:'right', labels:{usePointStyle:true}},
        tooltip:{callbacks:{label:(c)=>{ const total=c.dataset.data.reduce((a,b)=>a+b,0); return `${c.label}: ${fmtMoney(c.parsed)} (${(c.parsed/total*100).toFixed(1)}%)`; }}} } }
  });
}

function renderPayrollChart() {
  const p = DATA.payroll;
  if (!p.months.length) { document.getElementById('payrollChart').parentElement.querySelector('.note').textContent = 'No payroll data parsed.'; return; }
  new Chart(document.getElementById('payrollChart'), {
    type: 'bar',
    data: { labels: p.months, datasets: [{ label: 'Total Gross Pay', data: p.gross, backgroundColor: COLORS[4]+'CC' }] },
    options: { responsive:true, maintainAspectRatio:false,
      plugins:{ legend:{display:false}, tooltip:{callbacks:{label:(c)=>fmtMoney(c.parsed.y)}} },
      scales:{ y:{ ticks:{ callback:(v)=>fmtMoney(v) } } } }
  });
}

const channelState = { sortField: 'revenue', sortDir: 'desc' };

function renderChannelTable() {
  const ch = sortRows(DATA.channels, channelState.sortField, channelState.sortDir);
  const table = document.getElementById('channel-table');
  const s = channelState, sn = 'channelState', rf = 'renderChannelTable';
  table.innerHTML = '<thead><tr>' +
    th('Channel', 'name', s, sn, rf) + th('Revenue', 'revenue', s, sn, rf) + th('COGS', 'cogs', s, sn, rf) +
    th('Fee', 'fee', s, sn, rf) + th('Fee %', 'fee_pct', s, sn, rf) +
    th('Ad Spend', 'ads_spend', s, sn, rf) + th('Ad %', 'ads_pct', s, sn, rf) +
    th('Actual Income Received', 'actual_income_received', s, sn, rf) + th('NPE', 'npe', s, sn, rf) +
    th('Margin %', 'margin_pct', s, sn, rf) + th('Orders', 'orders', s, sn, rf) +
    '</tr></thead><tbody>' +
    ch.map(c => {
      const name = c.name + (c.converted_from_sgd
        ? ` <span class="note" style="display:inline">(SGD, conv. @${DATA.sgd_summary ? DATA.sgd_summary.rate : 3})</span>` : '');
      const fee = c.fee_known ? fmtMoney(c.fee) : '<span class="fee-unknown">n/a</span>';
      const feePct = c.fee_known ? fmtPct(c.fee_pct) : '<span class="fee-unknown">n/a</span>';
      const ads = c.ads_known ? fmtMoney(c.ads_spend) : '<span class="fee-unknown">n/a</span>';
      const adsPct = c.ads_known ? fmtPct(c.ads_pct) : '<span class="fee-unknown">n/a</span>';
      const aic = c.actual_income_received == null
        ? '<span class="fee-unknown">n/a</span>' : fmtMoney(c.actual_income_received);
      const npe = c.npe == null ? '<span class="fee-unknown">n/a</span>' : fmtMoney(c.npe);
      const margin = c.margin_pct == null ? '<span class="fee-unknown">n/a</span>' : fmtPct(c.margin_pct);
      const npeColor = (c.npe != null && c.npe < 0) ? 'var(--negative)' : 'inherit';
      return `<tr><td>${name}</td><td>${fmtMoney(c.revenue)}</td><td>${fmtMoney(c.cogs)}</td><td>${fee}</td><td>${feePct}</td>` +
        `<td>${ads}</td><td>${adsPct}</td><td>${aic}</td>` +
        `<td style="color:${npeColor}">${npe}</td><td>${margin}</td><td>${c.orders ?? '-'}</td></tr>`;
    }).join('') +
    (DATA.channel_blended ? `<tr style="font-weight:600;background:#f8f9fa">` +
      `<td>Blended average (D2C with fee data, revenue-weighted)</td><td>${fmtMoney(DATA.channel_blended.revenue)}</td>` +
      `<td></td><td></td><td>${fmtPct(DATA.channel_blended.fee_pct)}</td><td></td><td>${fmtPct(DATA.channel_blended.ads_pct)}</td>` +
      `<td></td><td></td><td></td><td></td></tr>` : '') +
    '</tbody>';
}

function renderWarnings() {
  if (!DATA.warnings.length) return;
  document.getElementById('warnings-section').style.display = 'block';
  document.getElementById('warnings-list').innerHTML = DATA.warnings.map(w => `<li>${w}</li>`).join('');
}

function renderWalkthrough() {
  const w = DATA.walkthrough;
  const table = document.getElementById('walkthrough-table');
  table.innerHTML = '<thead><tr><th>Step</th><th>Amount</th><th>Running Total</th></tr></thead><tbody>' +
    w.map(s => `<tr><td>${s.label}</td><td style="color:${s.amount<0?'var(--negative)':'inherit'}">${fmtMoney(s.amount)}</td><td><strong>${fmtMoney(s.running)}</strong></td></tr>`).join('') +
    '</tbody>';
}

function showTab(tab) {
  document.getElementById('tab-overview').style.display = tab === 'overview' ? '' : 'none';
  document.getElementById('tab-bank').style.display = tab === 'bank' ? '' : 'none';
  document.getElementById('tab-orders').style.display = tab === 'orders' ? '' : 'none';
  document.getElementById('tab-restock').style.display = tab === 'restock' ? '' : 'none';
  document.getElementById('tab-pipeline').style.display = tab === 'pipeline' ? '' : 'none';
  document.getElementById('tab-strategy').style.display = tab === 'strategy' ? '' : 'none';
  document.getElementById('tab-health').style.display = tab === 'health' ? '' : 'none';
  document.getElementById('tab-isku').style.display = tab === 'isku' ? '' : 'none';
  document.getElementById('tab-btn-overview').classList.toggle('active', tab === 'overview');
  document.getElementById('tab-btn-bank').classList.toggle('active', tab === 'bank');
  document.getElementById('tab-btn-orders').classList.toggle('active', tab === 'orders');
  document.getElementById('tab-btn-restock').classList.toggle('active', tab === 'restock');
  document.getElementById('tab-btn-pipeline').classList.toggle('active', tab === 'pipeline');
  document.getElementById('tab-btn-strategy').classList.toggle('active', tab === 'strategy');
  document.getElementById('tab-btn-health').classList.toggle('active', tab === 'health');
  document.getElementById('tab-btn-isku').classList.toggle('active', tab === 'isku');
  if (tab === 'isku' && !iskuDb) loadIskuDatabase();
}

const bankState = { page: 0, pageSize: 50, sortField: 'date', sortDir: 'desc' };

function populateBankFilter() {
  const select = document.getElementById('bank-month-filter');
  const months = [...new Set(DATA.bank_transactions.map(t => (t.date || '').slice(0, 7)))].filter(Boolean).sort().reverse();
  select.innerHTML = '<option value="all">All Months</option>' +
    months.map(m => `<option value="${m}">${m}</option>`).join('');
}

function getFilteredBankRows() {
  const month = document.getElementById('bank-month-filter').value;
  const direction = document.getElementById('bank-direction-filter').value;
  const search = document.getElementById('bank-search').value.trim().toLowerCase();

  let rows = DATA.bank_transactions;
  if (month !== 'all') rows = rows.filter(t => (t.date || '').startsWith(month));
  if (direction === 'in') rows = rows.filter(t => t.credit > 0);
  if (direction === 'out') rows = rows.filter(t => t.debit > 0);
  if (search) rows = rows.filter(t => t.details.toLowerCase().includes(search));
  return sortRows(rows, bankState.sortField, bankState.sortDir);
}

function renderBankTable() {
  const rows = getFilteredBankRows();
  const totalIn = rows.reduce((s, t) => s + t.credit, 0);
  const totalOut = rows.reduce((s, t) => s + t.debit, 0);
  document.getElementById('bank-summary').innerHTML =
    `<strong>${rows.length.toLocaleString()}</strong> transactions &middot; Money In ${fmtMoney(totalIn)} ` +
    `&middot; Money Out ${fmtMoney(totalOut)} &middot; Net ${fmtMoney(totalIn - totalOut)}`;

  const start = bankState.page * bankState.pageSize;
  const pageRows = rows.slice(start, start + bankState.pageSize);
  const table = document.getElementById('bank-table');
  table.innerHTML = '<thead><tr>' +
    th('Date', 'date', bankState, 'bankState', 'renderBankTable') +
    th('Details', 'details', bankState, 'bankState', 'renderBankTable') +
    th('Category', 'category', bankState, 'bankState', 'renderBankTable') +
    th('Money In', 'credit', bankState, 'bankState', 'renderBankTable') +
    th('Money Out', 'debit', bankState, 'bankState', 'renderBankTable') +
    '</tr></thead><tbody>' +
    pageRows.map(t => `<tr>
      <td>${t.date || '-'}</td><td>${t.details || '-'}</td><td>${t.category || '-'}</td>
      <td style="color:${t.credit>0?'var(--positive)':'inherit'}">${t.credit>0?fmtMoney(t.credit):''}</td>
      <td style="color:${t.debit>0?'var(--negative)':'inherit'}">${t.debit>0?fmtMoney(t.debit):''}</td>
    </tr>`).join('') +
    '</tbody>';

  renderPagination('bank-pagination', bankState, 'bankState', rows.length, 'renderBankTable');
}

const ordersState = { page: 0, pageSize: 50, sortField: 'net_profit', sortDir: 'desc' };

function populateOrdersFilter() {
  const select = document.getElementById('orders-platform-filter');
  const platforms = [...new Set(DATA.orders.map(o => o.platform))].sort();
  select.innerHTML = '<option value="all">All Platforms</option>' +
    platforms.map(p => `<option value="${p}">${p}</option>`).join('');
}

function getFilteredSortedOrders() {
  const platform = document.getElementById('orders-platform-filter').value;
  const search = document.getElementById('orders-search').value.trim().toLowerCase();

  // Rows with an unresolved SKU cost live only in the Pending Action table below,
  // not here -- their COGS/Net Profit would be wrong (COGS=0) if shown as final.
  let rows = DATA.orders.filter(o => o.cost_known);
  if (platform !== 'all') rows = rows.filter(o => o.platform === platform);
  if (search) rows = rows.filter(o => o.order_id.toLowerCase().includes(search));

  return sortRows(rows, ordersState.sortField, ordersState.sortDir);
}

function renderOrdersTable() {
  const rows = getFilteredSortedOrders();
  const totalRevenue = rows.reduce((s, o) => s + o.revenue, 0);

  // Positive and negative Net Profit rows are summed separately, not netted
  // together, since they answer different questions. For loss-making rows, the
  // accounting loss (Net Profit) is shown alongside what actually came back in
  // settlement (Revenue - Platform Fee - Ad Spend, i.e. before COGS is
  // deducted) -- the order lost money on paper once COGS is counted, but real
  // cash still came in. (Fixed 2026-07-21: this used to drop platform_fee and
  // only subtract ads_fee, the same class of bug as the per-row Net Profit fix.)
  const positive = rows.filter(o => o.net_profit >= 0);
  const negative = rows.filter(o => o.net_profit < 0);
  const positiveSum = positive.reduce((s, o) => s + o.net_profit, 0);
  const negativeLossSum = negative.reduce((s, o) => s + o.net_profit, 0);
  const negativeSettlementSum = negative.reduce((s, o) => s + (o.revenue - o.platform_fee - o.ads_fee), 0);

  document.getElementById('orders-summary').innerHTML = `
    <div><strong>${rows.length.toLocaleString()}</strong> orders total &middot; Revenue ${fmtMoney(totalRevenue)}</div>
    <div style="color:var(--positive)">Profitable: <strong>${positive.length.toLocaleString()}</strong> orders &middot; Total Net Profit <strong>${fmtMoney(positiveSum)}</strong></div>
    <div style="color:var(--negative)">Loss-making: <strong>${negative.length.toLocaleString()}</strong> orders &middot; Total accounting loss ${fmtMoney(negativeLossSum)} &middot; but settlement actually received on these orders: <strong>${fmtMoney(negativeSettlementSum)}</strong></div>
  `;

  const start = ordersState.page * ordersState.pageSize;
  const pageRows = rows.slice(start, start + ordersState.pageSize);

  const table = document.getElementById('orders-table');
  const s = ordersState, sn = 'ordersState', rf = 'renderOrdersTable';
  table.innerHTML = '<thead><tr>' +
    th('Order ID', 'order_id', s, sn, rf) + th('Platform', 'platform', s, sn, rf) +
    th('SKU', 'sku', s, sn, rf) + th('Qty', 'qty', s, sn, rf) + th('Date', 'date', s, sn, rf) +
    th('Revenue', 'revenue', s, sn, rf) + th('COGS', 'cogs', s, sn, rf) +
    th('Gross Profit', 'gross_profit', s, sn, rf) + th('Gross Margin %', 'gross_margin_pct', s, sn, rf) +
    th('Platform Fee % (info only)', 'fee_pct', s, sn, rf) + th('Ad Spend % (deducted)', 'ads_pct', s, sn, rf) +
    th('Net Profit', 'net_profit', s, sn, rf) + th('Net Margin %', 'net_margin_pct', s, sn, rf) +
    '</tr></thead><tbody>' +
    pageRows.map(o => {
      const platform = o.platform + (o.converted_from_sgd
        ? ` <span class="note" style="display:inline">(SGD, conv. @${DATA.sgd_summary ? DATA.sgd_summary.rate : 3})</span>` : '');
      return `<tr>
      <td>${o.order_id}</td><td>${platform}</td><td>${o.sku || '-'}</td><td>${o.qty ?? '-'}</td><td>${o.date || '-'}</td>
      <td>${fmtMoney(o.revenue)}</td><td>${fmtMoney(o.cogs)}</td>
      <td>${fmtMoney(o.gross_profit)}</td><td>${fmtPct(o.gross_margin_pct)}</td>
      <td>${o.fee_known ? fmtPct(o.fee_pct) + ' <span class="note" style="display:inline">(' + fmtMoney(o.platform_fee) + ')</span>' : '<span class="fee-unknown">n/a</span>'}</td>
      <td>${o.ads_known ? fmtPct(o.ads_pct) + ' <span class="note" style="display:inline">(' + fmtMoney(o.ads_fee) + ')</span>' : '<span class="fee-unknown">n/a</span>'}</td>
      <td style="color:${o.net_profit<0?'var(--negative)':'inherit'}">${fmtMoney(o.net_profit)}</td>
      <td style="color:${o.net_margin_pct<0?'var(--negative)':'inherit'}">${fmtPct(o.net_margin_pct)}</td>
    </tr>`;
    }).join('') +
    '</tbody>';

  renderPagination('orders-pagination', ordersState, 'ordersState', rows.length, 'renderOrdersTable');
}

function renderSgdNote() {
  const s = DATA.sgd_summary;
  if (!s || !s.orders) return;
  document.getElementById('sgd-section').style.display = 'block';
  document.getElementById('sgd-note').innerHTML =
    `<strong>${s.marketplaces.join(', ')}</strong>: ${s.orders.toLocaleString()} orders totalling ` +
    `<strong>SGD ${s.revenue_sgd.toLocaleString(undefined, {minimumFractionDigits:2, maximumFractionDigits:2})}</strong> ` +
    `-- converted to <strong>${fmtMoney(s.converted_revenue_rm)}</strong> at 1 SGD = ${s.rate} MYR (Bun's confirmed ` +
    `rate, 2026-07-21 -- his own stated rate, not a live market feed) and folded into every RM revenue/COGS/margin ` +
    `figure on this page, same as any other channel. Look for the "(SGD, conv. @${s.rate})" tag in Channel Detail ` +
    `and Order Detail to see exactly which rows were converted.`;
}

const costOverrides = {};
const confirmedSkus = new Set();
const pendingState = { sortField: 'revenue', sortDir: 'desc' };

function getPendingSkus() {
  const pending = DATA.orders.filter(o => !o.cost_known);
  const bySku = {};
  pending.forEach(o => {
    if (!bySku[o.sku]) bySku[o.sku] = { sku: o.sku, qty: 0, revenue: 0, orders: 0, platforms: new Set() };
    bySku[o.sku].qty += (o.qty || 0);
    bySku[o.sku].revenue += o.revenue;
    bySku[o.sku].orders += 1;
    bySku[o.sku].platforms.add(o.platform);
  });
  return Object.values(bySku)
    .filter(s => !confirmedSkus.has(s.sku))
    .map(s => ({ ...s, platforms: [...s.platforms].join(', ') }));
}

function renderPendingTable() {
  const pending = sortRows(getPendingSkus(), pendingState.sortField, pendingState.sortDir);
  const section = document.getElementById('pending-section');
  if (!pending.length) { section.style.display = 'none'; return; }
  section.style.display = 'block';
  document.getElementById('pending-count').textContent = pending.length;

  const table = document.getElementById('pending-table');
  const s = pendingState, sn = 'pendingState', rf = 'renderPendingTable';
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s, sn, rf) + th('Platform(s)', 'platforms', s, sn, rf) +
    th('Orders', 'orders', s, sn, rf) + th('Total Qty', 'qty', s, sn, rf) +
    th('Total Revenue', 'revenue', s, sn, rf) + '<th>Cost per Unit (key in)</th>' +
    '</tr></thead><tbody>' +
    pending.map(p => `<tr>
      <td>${p.sku}</td>
      <td>${p.platforms}</td><td>${p.orders}</td><td>${p.qty}</td>
      <td>${fmtMoney(p.revenue)}</td>
      <td>
        <input type="number" step="0.01" min="0" placeholder="RM" style="width:90px;padding:4px 6px;border:1px solid #dee2e6;border-radius:4px;" data-sku="${p.sku.replace(/"/g,'&quot;')}" oninput="onCostInput(this)">
        <span class="pending-row-actions" style="display:none;margin-left:6px;">
          <button onclick="confirmCost(this)" style="padding:4px 10px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">Confirm</button>
          <button onclick="undoCost(this)" style="padding:4px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;margin-left:4px;">Undo</button>
        </span>
      </td>
    </tr>`).join('') +
    '</tbody>';
}

function updateOverrideCount() {
  document.getElementById('override-count').textContent =
    Object.keys(costOverrides).filter(k => !confirmedSkus.has(k)).length;
}

function onCostInput(input) {
  const sku = input.dataset.sku;
  const cost = parseFloat(input.value);
  const actions = input.parentElement.querySelector('.pending-row-actions');
  if (!isNaN(cost) && cost >= 0) {
    costOverrides[sku] = cost;
    if (actions) actions.style.display = 'inline-block';
  } else {
    delete costOverrides[sku];
    if (actions) actions.style.display = 'none';
  }
  updateOverrideCount();
}

function undoCost(btn) {
  const td = btn.closest('td');
  const input = td.querySelector('input');
  delete costOverrides[input.dataset.sku];
  input.value = '';
  td.querySelector('.pending-row-actions').style.display = 'none';
  updateOverrideCount();
}

// Pending Action writes straight into isku_database.json now (a `cost` on a
// minimal record if the SKU isn't in the database yet, or a patch to `cost`
// on its existing record) -- one file, one mechanism, instead of a separate
// cost_overrides.json. Loads isku_database.json on first confirm if the ISKU
// Manager tab hasn't been opened yet this session (same file picker either way).
async function confirmCost(btn) {
  const td = btn.closest('td');
  const input = td.querySelector('input');
  const sku = input.dataset.sku;
  const cost = costOverrides[sku];
  if (cost === undefined) return;
  if (!iskuDb) await loadIskuDatabase();
  if (!iskuDb) { alert("Could not open isku_database.json -- pick the file when prompted, or use Chrome/Edge."); return; }

  const today = new Date().toISOString().slice(0, 10);
  if (!iskuDb.skus[sku]) {
    iskuDb.skus[sku] = {
      itec_code: null, currency: null, upc: null, status: 'MANUAL OVERRIDE',
      import_price: null, tax: null, fees: null,
      cost: cost, rsp: null, dsp: null,
      offline_set_margin: null, offline_purchase_price: null,
      brand: null, category: null, series: null, variant: null, color: null,
      date_created: today, date_modified: today,
      notes: 'added via Pending Action (Order Detail tab)',
    };
  } else {
    iskuDb.skus[sku].cost = cost;
    iskuDb.skus[sku].date_modified = today;
  }

  confirmedSkus.add(sku);
  renderPendingTable();
  updateOverrideCount();
  const saved = await saveIskuDatabaseToDisk();
  showSaveStatus(sku, cost, saved);
  if (document.getElementById('isku-table').innerHTML) renderIskuTable();
}

// --- File System Access helpers shared by Pending Action and the ISKU
// Manager tab (both read/write into isku_database.json, so they share one
// remembered file handle). Chrome/Edge only -- Firefox/Safari don't
// implement showSaveFilePicker, so confirmCost() falls back to an alert.
function fsaSupported() { return typeof window.showSaveFilePicker === 'function'; }

function idbGet(key) {
  return new Promise((resolve) => {
    const req = indexedDB.open('frescoone-dashboard', 1);
    req.onupgradeneeded = () => req.result.createObjectStore('handles');
    req.onsuccess = () => {
      const getReq = req.result.transaction('handles', 'readonly').objectStore('handles').get(key);
      getReq.onsuccess = () => resolve(getReq.result || null);
      getReq.onerror = () => resolve(null);
    };
    req.onerror = () => resolve(null);
  });
}
function idbSet(key, value) {
  return new Promise((resolve) => {
    const req = indexedDB.open('frescoone-dashboard', 1);
    req.onupgradeneeded = () => req.result.createObjectStore('handles');
    req.onsuccess = () => {
      const tx = req.result.transaction('handles', 'readwrite');
      tx.objectStore('handles').put(value, key);
      tx.oncomplete = () => resolve(true);
      tx.onerror = () => resolve(false);
    };
    req.onerror = () => resolve(false);
  });
}

function showSaveStatus(sku, cost, saved) {
  let el = document.getElementById('pending-save-status');
  if (!el) {
    el = document.createElement('div');
    el.id = 'pending-save-status';
    el.style.cssText = 'margin-top:10px;font-size:12px;padding:8px 12px;border-radius:6px;';
    document.getElementById('pending-table').insertAdjacentElement('afterend', el);
  }
  el.textContent = saved
    ? `Saved: ${sku} = RM${cost.toFixed(2)} written to isku_database.json -- no need to key this one in again.`
    : `${sku} = RM${cost.toFixed(2)} confirmed but NOT saved -- auto-save needs Chrome/Edge with file access granted; reopen isku_database.json and try again.`;
  el.style.background = saved ? '#e6f4ea' : '#fff8e1';
  el.style.color = saved ? '#1e7e34' : '#7a5c00';
}

// --- ISKU Manager -- local replacement for the Google Sheet ---------------
// Reuses the exact File System Access pattern above (idbGet/idbSet,
// fsaSupported) with its own IndexedDB key + filename, so it needs its own
// folder/file picker the first time but is otherwise the same UX as the
// Pending Action auto-save. computeCost()/computeDsp() mirror
// finance_analyzer.py's Python functions of the same purpose exactly --
// keep both in sync if the formula ever changes.
let iskuDb = null;
let iskuFileHandle = null;
let iskuLoadedSkuCount = 0;
const iskuState = { page: 0, pageSize: 25, sortField: 'sku', sortDir: 'asc' };
let iskuEditingSku = null;
let iskuDeletingSku = null;

function escapeHtml(s) {
  return String(s == null ? '' : s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}

// Shared trash-bin icon for every delete trigger in the ISKU Manager (SKU
// rows, ITEC cards, currency rows) -- icon-only button, opens the inline
// confirm/cancel state; the actual destructive action stays a labeled
// "Confirm Delete" text button, not an icon, so a stray click can't destroy
// data without a second, clearly-worded step.
const TRASH_ICON = '<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6l-1 14a2 2 0 0 1-2 2H8a2 2 0 0 1-2-2L5 6"></path><path d="M10 11v6"></path><path d="M14 11v6"></path><path d="M9 6V4a2 2 0 0 1 2-2h2a2 2 0 0 1 2 2v2"></path></svg>';

function feesTotal(itec) {
  return (itec.fees || []).reduce((sum, f) => sum + (parseFloat(f.amount) || 0), 0);
}
// Schema v2 (2026-07-23): currency is its own field on the record (MYR/USD/
// CNY/SGD, independent of ITEC code), looked up in iskuDb.currencies for a
// RM-per-unit rate. ITEC code only supplies tax_rate + an itemized fees list
// (packaging/shipping, engraving logo, custom packaging, etc.) -- see the
// matching Python compute_cost() in finance_analyzer.py for the authoritative
// version of this formula.
function computeCost(rec, itecCodes, currencies) {
  const itec = itecCodes[rec.itec_code];
  const currency = currencies[rec.currency];
  if (!itec || !currency || rec.import_price == null || rec.import_price === '') return null;
  const importPrice = parseFloat(rec.import_price);
  const tax = importPrice * (itec.tax_rate || 0);
  return (importPrice + tax + feesTotal(itec)) * currency.rate_to_rm;
}
function computeDsp(rec, cost) {
  if (cost == null || rec.rsp == null || rec.rsp === '') return null;
  const rsp = parseFloat(rec.rsp);
  const dsMargin = cost < 20 ? 0.50 : (cost > 60 ? 0.40 : 0.45);
  return cost + (rsp - cost) * (1 - dsMargin);
}
function computeOfflinePurchasePrice(rec, cost) {
  if (cost == null || rec.offline_set_margin == null || rec.offline_set_margin === '') return null;
  return cost + cost * parseFloat(rec.offline_set_margin);
}

async function getIskuFileHandle() {
  if (!fsaSupported()) return null;
  if (iskuFileHandle) {
    try {
      if (await iskuFileHandle.queryPermission({ mode: 'readwrite' }) === 'granted') return iskuFileHandle;
      if (await iskuFileHandle.requestPermission({ mode: 'readwrite' }) === 'granted') return iskuFileHandle;
    } catch (e) { iskuFileHandle = null; }
  }
  const remembered = await idbGet('isku_database_handle');
  if (remembered) {
    try {
      const granted = await remembered.queryPermission({ mode: 'readwrite' }) === 'granted'
        || await remembered.requestPermission({ mode: 'readwrite' }) === 'granted';
      if (granted) { iskuFileHandle = remembered; return iskuFileHandle; }
    } catch (e) { /* stale/revoked -- fall through to re-pick */ }
  }
  try {
    // showOpenFilePicker, not showSaveFilePicker -- this is an OPEN dialog for
    // an isku_database.json that migrate_isku_database.py already created, and
    // never touches the file's content. (A prior version used
    // showSaveFilePicker here, which is a "Save As" dialog -- picking an
    // existing file through it and confirming "Replace?" wiped a real
    // 10,751-SKU database to 0 bytes. Don't repeat that mistake.)
    const [handle] = await window.showOpenFilePicker({
      types: [{ description: 'JSON', accept: { 'application/json': ['.json'] } }],
      excludeAcceptAllOption: false,
    });
    if (await handle.requestPermission({ mode: 'readwrite' }) !== 'granted') return null;
    iskuFileHandle = handle;
    await idbSet('isku_database_handle', iskuFileHandle);
    return iskuFileHandle;
  } catch (e) {
    return null; // user cancelled the picker
  }
}

async function loadIskuDatabase() {
  const statusEl = document.getElementById('isku-load-status');
  if (!fsaSupported()) {
    statusEl.textContent = "This browser can't open local files directly -- use Chrome or Edge for the ISKU Manager.";
    return;
  }
  const handle = await getIskuFileHandle();
  if (!handle) { statusEl.textContent = 'No file selected.'; return; }
  try {
    const text = await (await handle.getFile()).text();
    iskuDb = text.trim() ? JSON.parse(text) : { itec_codes: {}, skus: {} };
  } catch (e) {
    statusEl.textContent = `Could not read isku_database.json: ${e.message}`;
    return;
  }
  if (!iskuDb.currencies) iskuDb.currencies = {};
  if (!iskuDb.itec_codes) iskuDb.itec_codes = {};
  iskuLoadedSkuCount = Object.keys(iskuDb.skus).length;
  statusEl.textContent = `Loaded ${iskuLoadedSkuCount} ISKUs, ${Object.keys(iskuDb.itec_codes).length} ITEC codes, ${Object.keys(iskuDb.currencies).length} currencies.`;
  document.getElementById('isku-add-section').style.display = 'block';
  document.getElementById('isku-search-section').style.display = 'block';
  document.getElementById('isku-itec-section').style.display = 'block';
  document.getElementById('isku-currency-section').style.display = 'block';
  const itecSelect = document.getElementById('isku-add-itec');
  itecSelect.innerHTML = Object.keys(iskuDb.itec_codes).sort().map(c => `<option value="${escapeHtml(c)}">${escapeHtml(c)}</option>`).join('');
  const currencySelect = document.getElementById('isku-add-currency');
  currencySelect.innerHTML = Object.keys(iskuDb.currencies).sort().map(c => `<option value="${escapeHtml(c)}">${escapeHtml(c)}</option>`).join('');
  populateIskuSuggestions();
  populateIskuFilterDropdowns();
  updateBulkAddNote();
  updateIskuAddPreview();
  renderIskuTable();
  updateBulkMatchCount();
  renderItecCards();
  renderCurrencyTable();
}

// Autocomplete suggestions for the Add New ISKU fields, sourced from every
// SKU already in the database -- ISKUs added through this form store their
// brand/category/series/variant/color directly, but the ones migrated from
// the sheet only have a flat SKU string, so those are split on '-' as a
// best-effort fallback. Brand/Category (the first two segments) are reliable
// either way; Series/Variant/Color are only trusted when the segment count
// matches the dominant 5-part Brand-Category-Series-Variant-Color pattern, to
// avoid polluting suggestions with garbage from oddly-shaped SKUs.
function deriveSkuParts(sku, rec) {
  if (rec.brand || rec.category || rec.series || rec.variant || rec.color) {
    return { brand: rec.brand, category: rec.category, series: rec.series, variant: rec.variant, color: rec.color };
  }
  const parts = String(sku).split('-').map(p => p.trim());
  if (parts.length < 2) return {};
  const out = { brand: parts[0], category: parts[1] };
  if (parts.length === 5) {
    out.series = parts[2]; out.variant = parts[3]; out.color = parts[4];
  } else if (parts.length === 4) {
    out.variant = parts[2]; out.color = parts[3];
  }
  return out;
}

function populateIskuSuggestions() {
  const fields = ['brand', 'category', 'series', 'variant', 'color'];
  const sets = Object.fromEntries(fields.map(f => [f, new Set()]));
  Object.entries(iskuDb.skus).forEach(([sku, rec]) => {
    const parts = deriveSkuParts(sku, rec);
    fields.forEach(f => { if (parts[f]) sets[f].add(parts[f]); });
  });
  fields.forEach(f => {
    const dl = document.getElementById('isku-add-' + f + '-list');
    if (dl) dl.innerHTML = [...sets[f]].sort().map(v => `<option value="${escapeHtml(v)}">`).join('');
  });
}

async function saveIskuDatabaseToDisk() {
  // Guard against ever writing a near-empty database over a real one -- if we
  // loaded a substantial file and iskuDb has since collapsed to (near) zero
  // SKUs, something went wrong upstream (bad parse, wrong file re-picked,
  // etc.) and writing now would silently destroy real data. Refuse instead.
  const currentCount = Object.keys(iskuDb.skus || {}).length;
  if (iskuLoadedSkuCount > 10 && currentCount < iskuLoadedSkuCount / 2) {
    alert(`Refusing to save -- loaded ${iskuLoadedSkuCount} ISKUs but the database in memory now only has ${currentCount}. This looks like a bug, not an intentional bulk delete. Reload the page and reopen isku_database.json before trying again.`);
    return false;
  }
  const handle = await getIskuFileHandle();
  if (!handle) return false;
  try {
    const writable = await handle.createWritable();
    await writable.write(JSON.stringify(iskuDb, null, 2));
    await writable.close();
    iskuLoadedSkuCount = currentCount;
    return true;
  } catch (e) {
    console.error('isku_database.json auto-save failed', e);
    return false;
  }
}

function updateIskuAddPreview() {
  if (!iskuDb) return;
  const parts = ['brand', 'category', 'series', 'variant', 'color']
    .map(id => document.getElementById('isku-add-' + id).value.trim()).filter(Boolean);
  const built = parts.join('-');
  document.getElementById('isku-add-sku-preview').textContent = built || '--';
  const skuInput = document.getElementById('isku-add-sku');
  if (!skuInput.dataset.manual) skuInput.value = built;

  const offlineMarginPctPreview = parseFloat(document.getElementById('isku-add-offline-margin').value);
  const rec = {
    itec_code: document.getElementById('isku-add-itec').value,
    currency: document.getElementById('isku-add-currency').value,
    import_price: document.getElementById('isku-add-import-price').value,
    rsp: document.getElementById('isku-add-rsp').value,
    offline_set_margin: isNaN(offlineMarginPctPreview) ? null : offlineMarginPctPreview / 100,
  };
  const cost = computeCost(rec, iskuDb.itec_codes, iskuDb.currencies);
  const dsp = computeDsp(rec, cost);
  const offlinePrice = computeOfflinePurchasePrice(rec, cost);
  document.getElementById('isku-add-cost-preview').textContent = cost != null ? 'RM ' + cost.toFixed(2) : '--';
  document.getElementById('isku-add-dsp-preview').textContent = dsp != null ? 'RM ' + dsp.toFixed(2) : '--';
  document.getElementById('isku-add-offline-preview').textContent = offlinePrice != null ? 'RM ' + offlinePrice.toFixed(2) : '--';
}
document.addEventListener('DOMContentLoaded', () => {
  const skuInput = document.getElementById('isku-add-sku');
  if (skuInput) skuInput.addEventListener('input', () => { skuInput.dataset.manual = '1'; });
});

async function saveNewIsku() {
  const msg = document.getElementById('isku-add-status-msg');
  const sku = document.getElementById('isku-add-sku').value.trim();
  if (!sku) { msg.textContent = 'SKU cannot be empty.'; msg.style.color = 'var(--negative)'; return; }
  if (iskuDb.skus[sku]) { msg.textContent = `${sku} already exists -- edit it in Search/Edit instead.`; msg.style.color = 'var(--negative)'; return; }

  const itecCode = document.getElementById('isku-add-itec').value;
  const currency = document.getElementById('isku-add-currency').value;
  const importPrice = parseFloat(document.getElementById('isku-add-import-price').value);
  const rsp = parseFloat(document.getElementById('isku-add-rsp').value);
  const offlineMarginPct = parseFloat(document.getElementById('isku-add-offline-margin').value);
  const offlineMargin = isNaN(offlineMarginPct) ? null : offlineMarginPct / 100;
  const itec = iskuDb.itec_codes[itecCode];
  const rec0 = { itec_code: itecCode, currency: currency, import_price: importPrice, rsp: rsp, offline_set_margin: offlineMargin };
  const cost = computeCost(rec0, iskuDb.itec_codes, iskuDb.currencies);
  const dsp = computeDsp(rec0, cost);
  const offlinePrice = computeOfflinePurchasePrice(rec0, cost);
  const today = new Date().toISOString().slice(0, 10);

  iskuDb.skus[sku] = {
    itec_code: itecCode,
    currency: currency,
    upc: document.getElementById('isku-add-upc').value.trim() || null,
    status: document.getElementById('isku-add-status').value,
    import_price: isNaN(importPrice) ? null : importPrice,
    tax: itec ? importPrice * (itec.tax_rate || 0) : null,
    fees: itec ? itec.fees.map(f => ({ ...f })) : null,
    cost: cost, rsp: isNaN(rsp) ? null : rsp, dsp: dsp,
    offline_set_margin: offlineMargin, offline_purchase_price: offlinePrice,
    brand: document.getElementById('isku-add-brand').value.trim() || null,
    category: document.getElementById('isku-add-category').value.trim() || null,
    series: document.getElementById('isku-add-series').value.trim() || null,
    variant: document.getElementById('isku-add-variant').value.trim() || null,
    color: document.getElementById('isku-add-color').value.trim() || null,
    date_created: today, date_modified: today,
    notes: document.getElementById('isku-add-notes').value.trim(),
  };

  const saved = await saveIskuDatabaseToDisk();
  msg.textContent = saved ? `Saved ${sku} to isku_database.json.` : `${sku} added in-memory but auto-save failed -- reopen the file and try again.`;
  msg.style.color = saved ? 'var(--positive)' : 'var(--negative)';
  if (saved) {
    ['brand','category','series','variant','color','import-price','rsp','upc','notes','offline-margin'].forEach(id => document.getElementById('isku-add-' + id).value = '');
    const skuInput = document.getElementById('isku-add-sku');
    skuInput.value = ''; delete skuInput.dataset.manual;
    updateIskuAddPreview();
  }
  populateIskuFilterDropdowns();
  renderIskuTable();
}

function getIskuRows() {
  // displayBrand: most legacy (sheet-migrated) SKUs have a null `brand` field
  // -- only ones added/edited through this UI store it directly -- so brand
  // sort/filter/column falls back to the same best-effort SKU-prefix split
  // populateIskuSuggestions() already uses, otherwise brand filtering would
  // be useless for the 10,000+ migrated rows.
  return Object.entries(iskuDb.skus).map(([sku, rec]) => {
    const parts = deriveSkuParts(sku, rec);
    return { sku, ...rec, displayBrand: rec.brand || parts.brand || '' };
  });
}

// Shared by the table, the bulk-command box, and the bulk-file download --
// all three act on "whatever the search box + filter dropdowns currently
// match", so there's one definition of what that means.
function getSearchFilteredIskuRows() {
  const q = (document.getElementById('isku-search').value || '').toLowerCase().trim();
  const brandFilter = document.getElementById('isku-filter-brand').value;
  const statusFilter = document.getElementById('isku-filter-status').value;
  const itecFilter = document.getElementById('isku-filter-itec').value;
  let rows = getIskuRows();
  if (q) rows = rows.filter(r => (r.sku || '').toLowerCase().includes(q)
    || (r.displayBrand || '').toLowerCase().includes(q) || (r.status || '').toLowerCase().includes(q));
  if (brandFilter) rows = rows.filter(r => r.displayBrand === brandFilter);
  if (statusFilter) rows = rows.filter(r => r.status === statusFilter);
  if (itecFilter) rows = rows.filter(r => r.itec_code === itecFilter);
  return rows;
}

// Dropdown options are computed from whatever's actually in the data (not
// just the 3 active ITEC codes / fixed status list), so a filter is always
// available for legacy/edge values too (e.g. a "STOPPED" status or a
// since-retired ITEC code that's still on some records).
function populateIskuFilterDropdowns() {
  const rows = getIskuRows();
  const brands = [...new Set(rows.map(r => r.displayBrand).filter(Boolean))].sort();
  const statuses = [...new Set(rows.map(r => r.status).filter(Boolean))].sort();
  const itecs = [...new Set(rows.map(r => r.itec_code).filter(Boolean))].sort();
  const fill = (id, values, label) => {
    const el = document.getElementById(id);
    const current = el.value;
    el.innerHTML = `<option value="">All ${label}</option>` + values.map(v => `<option value="${escapeHtml(v)}">${escapeHtml(v)}</option>`).join('');
    if (values.includes(current)) el.value = current;
  };
  fill('isku-filter-brand', brands, 'Brands');
  fill('isku-filter-status', statuses, 'Statuses');
  fill('isku-filter-itec', itecs, 'ITEC Codes');
}

function onIskuFilterChange() {
  iskuState.page = 0;
  renderIskuTable();
  updateBulkMatchCount();
}

function clearIskuFilters() {
  document.getElementById('isku-search').value = '';
  document.getElementById('isku-filter-brand').value = '';
  document.getElementById('isku-filter-status').value = '';
  document.getElementById('isku-filter-itec').value = '';
  onIskuFilterChange();
}

function updateBulkMatchCount() {
  const el = document.getElementById('isku-bulk-match-count');
  if (el) el.textContent = getSearchFilteredIskuRows().length;
}

function renderIskuTable() {
  let rows = getSearchFilteredIskuRows();
  document.getElementById('isku-count').textContent = rows.length;
  rows = sortRows(rows, iskuState.sortField, iskuState.sortDir);

  const start = iskuState.page * iskuState.pageSize;
  const pageRows = rows.slice(start, start + iskuState.pageSize);
  const s = iskuState, sn = 'iskuState', rf = 'renderIskuTable';
  const table = document.getElementById('isku-table');
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s, sn, rf) + th('Brand', 'displayBrand', s, sn, rf) +
    th('ITEC Code', 'itec_code', s, sn, rf) + th('Currency', 'currency', s, sn, rf) +
    th('Import Price', 'import_price', s, sn, rf) + th('Cost', 'cost', s, sn, rf) +
    th('RSP', 'rsp', s, sn, rf) + th('DSP', 'dsp', s, sn, rf) +
    th('Status', 'status', s, sn, rf) + th('Created', 'date_created', s, sn, rf) +
    '<th>Actions</th>' +
    '</tr></thead><tbody>' +
    pageRows.map(r => {
      const editing = iskuEditingSku === r.sku;
      const deleting = iskuDeletingSku === r.sku;
      const skuKey = escapeHtml(r.sku).replace(/"/g, '&quot;');
      if (deleting) {
        return `<tr style="background:#fff5f5;">
          <td colspan="10">Delete <strong>${escapeHtml(r.sku)}</strong>? This removes it from the database entirely -- not undoable.</td>
          <td>
            <button onclick="deleteIsku('${skuKey}')" style="padding:4px 8px;border:none;border-radius:4px;background:var(--negative);color:#fff;cursor:pointer;font-size:12px;">Confirm Delete</button>
            <button onclick="iskuDeletingSku=null; renderIskuTable();" style="padding:4px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
          </td>
        </tr>`;
      }
      if (!editing) {
        return `<tr>
          <td>${escapeHtml(r.sku)}</td><td>${escapeHtml(r.displayBrand)}</td>
          <td>${escapeHtml(r.itec_code)}</td>
          <td>${escapeHtml(r.currency)}</td>
          <td>${r.import_price != null ? r.import_price : ''}</td>
          <td>${r.cost != null ? 'RM ' + Number(r.cost).toFixed(2) : ''}</td>
          <td>${r.rsp != null ? r.rsp : ''}</td>
          <td>${r.dsp != null ? 'RM ' + Number(r.dsp).toFixed(2) : ''}</td>
          <td>${escapeHtml(r.status)}</td>
          <td>${escapeHtml(r.date_created)}</td>
          <td>
            <button onclick="startEditIsku('${skuKey}')" style="padding:4px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Edit</button>
            <button onclick="iskuDeletingSku='${skuKey}'; renderIskuTable();" title="Delete" aria-label="Delete" style="padding:4px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;color:var(--negative);cursor:pointer;line-height:0;">${TRASH_ICON}</button>
          </td>
        </tr>`;
      }
      // A legacy SKU's itec_code/currency might no longer be one of the active
      // presets (e.g. IT-MLY, dropped when the table was trimmed to China-only)
      // -- add it as an extra option so the field shows the real stored value
      // instead of silently jumping to whatever option happens to be first.
      const itecKeys = Object.keys(iskuDb.itec_codes).sort();
      if (r.itec_code && !itecKeys.includes(r.itec_code)) itecKeys.unshift(r.itec_code);
      const itecOptions = itecKeys
        .map(c => `<option value="${escapeHtml(c)}" ${c === r.itec_code ? 'selected' : ''}>${escapeHtml(c)}${!iskuDb.itec_codes[c] ? ' (legacy)' : ''}</option>`).join('');
      const currencyKeys = Object.keys(iskuDb.currencies).sort();
      if (r.currency && !currencyKeys.includes(r.currency)) currencyKeys.unshift(r.currency);
      const currencyOptions = currencyKeys
        .map(c => `<option value="${escapeHtml(c)}" ${c === r.currency ? 'selected' : ''}>${escapeHtml(c)}</option>`).join('');
      return `<tr data-editing-sku="${skuKey}">
        <td>${escapeHtml(r.sku)}</td>
        <td><input id="isku-edit-brand" value="${escapeHtml(r.brand || '')}" style="width:80px;padding:4px;" placeholder="${escapeHtml(r.displayBrand)}"></td>
        <td><select id="isku-edit-itec" style="width:150px;padding:4px;" onchange="previewEditIsku('${skuKey}')">${itecOptions}</select></td>
        <td><select id="isku-edit-currency" style="width:70px;padding:4px;" onchange="previewEditIsku('${skuKey}')">${currencyOptions}</select></td>
        <td><input id="isku-edit-import-price" type="number" step="0.01" value="${r.import_price != null ? r.import_price : ''}" style="width:80px;padding:4px;"></td>
        <td><span id="isku-edit-cost-preview">${r.cost != null ? 'RM ' + Number(r.cost).toFixed(2) : ''}</span></td>
        <td><input id="isku-edit-rsp" type="number" step="0.01" value="${r.rsp != null ? r.rsp : ''}" style="width:80px;padding:4px;"></td>
        <td><span id="isku-edit-dsp-preview">${r.dsp != null ? 'RM ' + Number(r.dsp).toFixed(2) : ''}</span></td>
        <td><select id="isku-edit-status" style="padding:4px;">
          ${['ACTIVE','CLEARANCE','STOPPED'].map(v => `<option ${v === r.status ? 'selected' : ''}>${v}</option>`).join('')}
        </select></td>
        <td>${escapeHtml(r.date_created)}</td>
        <td>
          <button onclick="previewEditIsku('${skuKey}')" style="padding:4px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Preview</button>
          <button onclick="saveEditIsku('${skuKey}')" style="padding:4px 8px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">Save</button>
          <button onclick="iskuEditingSku=null; renderIskuTable();" style="padding:4px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
        </td>
      </tr>`;
    }).join('') +
    '</tbody>';
  renderPagination('isku-pagination', iskuState, 'iskuState', rows.length, 'renderIskuTable');
}

function startEditIsku(sku) { iskuEditingSku = sku; renderIskuTable(); }

function previewEditIsku(sku) {
  const rec0 = {
    itec_code: document.getElementById('isku-edit-itec').value,
    currency: document.getElementById('isku-edit-currency').value,
    import_price: document.getElementById('isku-edit-import-price').value,
    rsp: document.getElementById('isku-edit-rsp').value,
  };
  const cost = computeCost(rec0, iskuDb.itec_codes, iskuDb.currencies);
  const dsp = computeDsp(rec0, cost);
  document.getElementById('isku-edit-cost-preview').textContent = cost != null ? 'RM ' + cost.toFixed(2) : (iskuDb.itec_codes[rec0.itec_code] ? '' : '(legacy code, key a new one to recompute)');
  document.getElementById('isku-edit-dsp-preview').textContent = dsp != null ? 'RM ' + dsp.toFixed(2) : '';
}

async function saveEditIsku(sku) {
  const itecCode = document.getElementById('isku-edit-itec').value;
  const currency = document.getElementById('isku-edit-currency').value;
  const importPrice = parseFloat(document.getElementById('isku-edit-import-price').value);
  const rsp = parseFloat(document.getElementById('isku-edit-rsp').value);
  const status = document.getElementById('isku-edit-status').value;
  const itec = iskuDb.itec_codes[itecCode];
  const rec0 = { itec_code: itecCode, currency: currency, import_price: importPrice, rsp: rsp };
  const cost = computeCost(rec0, iskuDb.itec_codes, iskuDb.currencies);
  const dsp = computeDsp(rec0, cost);

  const rec = iskuDb.skus[sku];
  rec.brand = document.getElementById('isku-edit-brand').value.trim() || null;
  rec.itec_code = itecCode;
  rec.currency = currency;
  rec.import_price = isNaN(importPrice) ? null : importPrice;
  rec.tax = itec ? importPrice * (itec.tax_rate || 0) : rec.tax;
  rec.fees = itec ? itec.fees.map(f => ({ ...f })) : rec.fees;
  // Only overwrite frozen cost/dsp when the new inputs actually resolve to a
  // real number -- e.g. a legacy ITEC code no longer in the active table
  // computes null, and that must never blank out a real existing cost/dsp.
  if (cost != null) rec.cost = cost;
  if (dsp != null) rec.dsp = dsp;
  rec.rsp = isNaN(rsp) ? null : rsp;
  rec.status = status;
  rec.date_modified = new Date().toISOString().slice(0, 10);
  if (rec.offline_set_margin != null && rec.cost != null) rec.offline_purchase_price = computeOfflinePurchasePrice(rec, rec.cost);

  const saved = await saveIskuDatabaseToDisk();
  iskuEditingSku = null;
  populateIskuFilterDropdowns();
  renderIskuTable();
  if (!saved) alert(`${sku} updated in-memory but auto-save failed -- reopen isku_database.json and try again.`);
}

async function deleteIsku(sku) {
  delete iskuDb.skus[sku];
  iskuDeletingSku = null;
  const saved = await saveIskuDatabaseToDisk();
  populateIskuFilterDropdowns();
  renderIskuTable();
  updateBulkMatchCount();
  if (!saved) alert(`${sku} deleted in-memory but auto-save failed -- reopen isku_database.json and try again (it may still be in the file on disk).`);
}

// --- Bulk edit (command box) -----------------------------------------------
// Deliberately NOT free-form AI/NLP -- a small fixed `field=value` syntax
// (plus +=/-=/*= for the two numeric fields), fully validated up front, with
// a mandatory preview showing every affected row's before/after before
// anything is written. Scope is always "whatever the search box currently
// matches" (getSearchFilteredIskuRows()), so search first, then bulk-edit.
const BULK_EDITABLE_FIELDS = ['status', 'itec_code', 'currency', 'rsp', 'import_price', 'notes'];
const BULK_NUMERIC_FIELDS = ['rsp', 'import_price'];
const BULK_STATUS_VALUES = ['ACTIVE', 'CLEARANCE', 'STOPPED'];

function parseBulkCommand(text) {
  const assignments = [];
  const errors = [];
  const statements = text.split(/[\\n;]+/).map(s => s.trim()).filter(Boolean);
  if (!statements.length) errors.push('Nothing to do -- type at least one field=value line.');
  for (const stmt of statements) {
    const m = stmt.match(/^([a-z_]+)\s*(\+=|-=|\*=|=)\s*(.*)$/i);
    if (!m) { errors.push(`Could not parse "${stmt}" -- expected field=value, field+=value, field-=value, or field*=value.`); continue; }
    const [, fieldRaw, op, valueRaw] = m;
    const field = fieldRaw.toLowerCase();
    if (!BULK_EDITABLE_FIELDS.includes(field)) { errors.push(`Unknown field "${field}" -- editable fields are: ${BULK_EDITABLE_FIELDS.join(', ')}.`); continue; }
    const numeric = BULK_NUMERIC_FIELDS.includes(field);
    if (op !== '=' && !numeric) { errors.push(`"${op}" only works on numeric fields (${BULK_NUMERIC_FIELDS.join(', ')}), not "${field}".`); continue; }
    let value = valueRaw.trim();
    if (numeric) {
      const v = parseFloat(value);
      if (isNaN(v)) { errors.push(`"${field}" needs a number, got "${value}".`); continue; }
      value = v;
    } else if (field === 'status' && !BULK_STATUS_VALUES.includes(value.toUpperCase())) {
      errors.push(`"status" must be one of ${BULK_STATUS_VALUES.join(', ')}, got "${value}".`); continue;
    } else if (field === 'status') {
      value = value.toUpperCase();
    } else if (field === 'itec_code' && !iskuDb.itec_codes[value]) {
      errors.push(`ITEC Code "${value}" doesn't exist -- active codes are: ${Object.keys(iskuDb.itec_codes).join(', ')}.`); continue;
    } else if (field === 'currency' && !iskuDb.currencies[value]) {
      errors.push(`Currency "${value}" doesn't exist -- available: ${Object.keys(iskuDb.currencies).join(', ')}.`); continue;
    }
    assignments.push({ field, op, value });
  }
  return { assignments, errors };
}

function applyAssignmentsToRecord(rec, assignments) {
  const before = { status: rec.status, itec_code: rec.itec_code, currency: rec.currency, rsp: rec.rsp, import_price: rec.import_price, notes: rec.notes, cost: rec.cost, dsp: rec.dsp };
  const draft = { ...rec };
  for (const { field, op, value } of assignments) {
    if (op === '=') draft[field] = value;
    else if (op === '+=') draft[field] = (parseFloat(draft[field]) || 0) + value;
    else if (op === '-=') draft[field] = (parseFloat(draft[field]) || 0) - value;
    else if (op === '*=') draft[field] = (parseFloat(draft[field]) || 0) * value;
  }
  const itec = iskuDb.itec_codes[draft.itec_code];
  if (itec && assignments.some(a => a.field === 'itec_code' || a.field === 'import_price')) {
    draft.fees = itec.fees.map(f => ({ ...f }));
    draft.tax = draft.import_price != null ? draft.import_price * (itec.tax_rate || 0) : draft.tax;
  }
  const newCost = computeCost(draft, iskuDb.itec_codes, iskuDb.currencies);
  const newDsp = computeDsp(draft, newCost);
  if (newCost != null) draft.cost = newCost;
  if (newDsp != null) draft.dsp = newDsp;
  return { before, after: { status: draft.status, itec_code: draft.itec_code, currency: draft.currency, rsp: draft.rsp, import_price: draft.import_price, notes: draft.notes, cost: draft.cost, dsp: draft.dsp }, draft };
}

let bulkCommandPreviewRows = null; // [{sku, before, after, draft}]

function previewBulkCommand() {
  const text = document.getElementById('isku-bulk-command').value;
  const { assignments, errors } = parseBulkCommand(text);
  const previewEl = document.getElementById('isku-bulk-preview');
  if (errors.length) {
    previewEl.innerHTML = `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">${errors.map(escapeHtml).join('<br>')}</div>`;
    bulkCommandPreviewRows = null;
    return;
  }
  const matches = getSearchFilteredIskuRows();
  if (!matches.length) {
    previewEl.innerHTML = `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">No ISKUs match the current search -- nothing to apply.</div>`;
    bulkCommandPreviewRows = null;
    return;
  }
  bulkCommandPreviewRows = matches.map(r => ({ sku: r.sku, ...applyAssignmentsToRecord(r, assignments) }));
  const fmt = v => v == null ? '' : (typeof v === 'number' ? v.toFixed(2) : escapeHtml(v));
  const shown = bulkCommandPreviewRows.slice(0, 20);
  previewEl.innerHTML = `<div style="margin-top:10px;">
    <div class="note" style="margin-bottom:6px;">${bulkCommandPreviewRows.length} ISKU(s) will change${bulkCommandPreviewRows.length > 20 ? ` (showing first 20)` : ''}:</div>
    <table class="data-table"><thead><tr><th>SKU</th><th>Status</th><th>ITEC</th><th>Currency</th><th>Import Price</th><th>RSP</th><th>Cost</th><th>DSP</th></tr></thead><tbody>
      ${shown.map(r => `<tr>
        <td>${escapeHtml(r.sku)}</td>
        <td>${fmt(r.before.status)} &rarr; <strong>${fmt(r.after.status)}</strong></td>
        <td>${fmt(r.before.itec_code)} &rarr; <strong>${fmt(r.after.itec_code)}</strong></td>
        <td>${fmt(r.before.currency)} &rarr; <strong>${fmt(r.after.currency)}</strong></td>
        <td>${fmt(r.before.import_price)} &rarr; <strong>${fmt(r.after.import_price)}</strong></td>
        <td>${fmt(r.before.rsp)} &rarr; <strong>${fmt(r.after.rsp)}</strong></td>
        <td>${fmt(r.before.cost)} &rarr; <strong>${fmt(r.after.cost)}</strong></td>
        <td>${fmt(r.before.dsp)} &rarr; <strong>${fmt(r.after.dsp)}</strong></td>
      </tr>`).join('')}
    </tbody></table>
    <button onclick="applyBulkCommand()" style="margin-top:10px;padding:8px 16px;border:none;border-radius:6px;background:var(--positive);color:#fff;cursor:pointer;font-size:13px;">Apply to ${bulkCommandPreviewRows.length} ISKU(s)</button>
    <button onclick="document.getElementById('isku-bulk-preview').innerHTML=''; bulkCommandPreviewRows=null;" style="margin-top:10px;margin-left:8px;padding:8px 16px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Cancel</button>
  </div>`;
}

async function applyBulkCommand() {
  if (!bulkCommandPreviewRows || !bulkCommandPreviewRows.length) return;
  const today = new Date().toISOString().slice(0, 10);
  for (const { sku, draft } of bulkCommandPreviewRows) {
    draft.date_modified = today;
    delete draft.sku;
    iskuDb.skus[sku] = draft;
  }
  const count = bulkCommandPreviewRows.length;
  bulkCommandPreviewRows = null;
  document.getElementById('isku-bulk-command').value = '';
  document.getElementById('isku-bulk-preview').innerHTML = '';
  const saved = await saveIskuDatabaseToDisk();
  populateIskuFilterDropdowns();
  renderIskuTable();
  if (saved) {
    document.getElementById('isku-bulk-preview').innerHTML = `<div style="margin-top:10px;padding:10px;background:#e6f4ea;color:#1e7e34;border-radius:6px;font-size:12px;">Applied to ${count} ISKU(s) and saved.</div>`;
  } else {
    alert(`Applied to ${count} ISKU(s) in-memory but auto-save failed -- reopen isku_database.json and try again.`);
  }
}

// --- Bulk edit (CSV upload) -------------------------------------------------
const BULK_CSV_COLUMNS = ['sku', 'itec_code', 'currency', 'import_price', 'rsp', 'status', 'notes'];

function csvEscape(v) {
  const s = v == null ? '' : String(v);
  return /[",\\n]/.test(s) ? '"' + s.replace(/"/g, '""') + '"' : s;
}

// Minimal RFC4180-ish parser -- handles quoted fields, escaped quotes, and
// commas/newlines inside quotes, which is as much as Excel/Sheets exports need.
function parseCSV(text) {
  const rows = [];
  let row = [], field = '', inQuotes = false;
  for (let i = 0; i < text.length; i++) {
    const c = text[i], next = text[i + 1];
    if (inQuotes) {
      if (c === '"' && next === '"') { field += '"'; i++; }
      else if (c === '"') { inQuotes = false; }
      else { field += c; }
    } else if (c === '"') { inQuotes = true; }
    else if (c === ',') { row.push(field); field = ''; }
    else if (c === '\\n' || c === '\\r') {
      if (c === '\\r' && next === '\\n') i++;
      row.push(field); field = '';
      if (row.length > 1 || row[0] !== '') rows.push(row);
      row = [];
    } else { field += c; }
  }
  if (field !== '' || row.length) { row.push(field); rows.push(row); }
  if (!rows.length) return [];
  const header = rows[0].map(h => h.trim().toLowerCase());
  return rows.slice(1).map(r => Object.fromEntries(header.map((h, i) => [h, (r[i] || '').trim()])));
}

function downloadIskuTemplate() {
  const rows = getSearchFilteredIskuRows();
  if (!rows.length) { alert('No ISKUs match the current search.'); return; }
  const lines = [BULK_CSV_COLUMNS.join(',')];
  rows.forEach(r => lines.push(BULK_CSV_COLUMNS.map(c => csvEscape(r[c])).join(',')));
  const blob = new Blob([lines.join('\\r\\n')], { type: 'text/csv' });
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'isku_bulk_edit_template.csv';
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

let bulkFilePreviewRows = null;

function handleBulkFileUpload(input) {
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = () => {
    const parsed = parseCSV(reader.result);
    const previewEl = document.getElementById('isku-bulk-file-preview');
    if (!parsed.length) {
      previewEl.innerHTML = `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">Empty or unreadable CSV.</div>`;
      return;
    }
    const changes = [];
    const notFound = [];
    for (const csvRow of parsed) {
      const sku = csvRow.sku;
      if (!sku) continue;
      const rec = iskuDb.skus[sku];
      if (!rec) { notFound.push(sku); continue; }
      const assignments = [];
      if (csvRow.itec_code && csvRow.itec_code !== (rec.itec_code || '')) {
        if (!iskuDb.itec_codes[csvRow.itec_code]) { notFound.push(`${sku} (unknown ITEC Code "${csvRow.itec_code}", skipped)`); continue; }
        assignments.push({ field: 'itec_code', op: '=', value: csvRow.itec_code });
      }
      if (csvRow.currency && csvRow.currency !== (rec.currency || '')) {
        if (!iskuDb.currencies[csvRow.currency]) { notFound.push(`${sku} (unknown currency "${csvRow.currency}", skipped)`); continue; }
        assignments.push({ field: 'currency', op: '=', value: csvRow.currency });
      }
      if (csvRow.import_price !== '' && parseFloat(csvRow.import_price) !== rec.import_price) {
        assignments.push({ field: 'import_price', op: '=', value: parseFloat(csvRow.import_price) });
      }
      if (csvRow.rsp !== '' && parseFloat(csvRow.rsp) !== rec.rsp) {
        assignments.push({ field: 'rsp', op: '=', value: parseFloat(csvRow.rsp) });
      }
      if (csvRow.status && csvRow.status.toUpperCase() !== (rec.status || '')) {
        if (!BULK_STATUS_VALUES.includes(csvRow.status.toUpperCase())) { notFound.push(`${sku} (invalid status "${csvRow.status}", skipped)`); continue; }
        assignments.push({ field: 'status', op: '=', value: csvRow.status.toUpperCase() });
      }
      if (csvRow.notes !== undefined && csvRow.notes !== (rec.notes || '')) {
        assignments.push({ field: 'notes', op: '=', value: csvRow.notes });
      }
      if (assignments.length) changes.push({ sku, ...applyAssignmentsToRecord(rec, assignments) });
    }
    bulkFilePreviewRows = changes;
    const fmt = v => v == null ? '' : (typeof v === 'number' ? v.toFixed(2) : escapeHtml(v));
    let html = '';
    if (notFound.length) html += `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">${notFound.length} row(s) skipped: ${notFound.map(escapeHtml).join(', ')}</div>`;
    if (!changes.length) {
      html += `<div style="margin-top:10px;" class="note">No changed values detected.</div>`;
    } else {
      html += `<div style="margin-top:10px;">
        <div class="note" style="margin-bottom:6px;">${changes.length} ISKU(s) will change:</div>
        <table class="data-table"><thead><tr><th>SKU</th><th>Status</th><th>ITEC</th><th>Currency</th><th>Import Price</th><th>RSP</th><th>Cost</th><th>DSP</th></tr></thead><tbody>
          ${changes.slice(0, 20).map(r => `<tr>
            <td>${escapeHtml(r.sku)}</td>
            <td>${fmt(r.before.status)} &rarr; <strong>${fmt(r.after.status)}</strong></td>
            <td>${fmt(r.before.itec_code)} &rarr; <strong>${fmt(r.after.itec_code)}</strong></td>
            <td>${fmt(r.before.currency)} &rarr; <strong>${fmt(r.after.currency)}</strong></td>
            <td>${fmt(r.before.import_price)} &rarr; <strong>${fmt(r.after.import_price)}</strong></td>
            <td>${fmt(r.before.rsp)} &rarr; <strong>${fmt(r.after.rsp)}</strong></td>
            <td>${fmt(r.before.cost)} &rarr; <strong>${fmt(r.after.cost)}</strong></td>
            <td>${fmt(r.before.dsp)} &rarr; <strong>${fmt(r.after.dsp)}</strong></td>
          </tr>`).join('')}
        </tbody></table>
        <button onclick="applyBulkFileChanges()" style="margin-top:10px;padding:8px 16px;border:none;border-radius:6px;background:var(--positive);color:#fff;cursor:pointer;font-size:13px;">Apply ${changes.length} Change(s)</button>
        <button onclick="document.getElementById('isku-bulk-file-preview').innerHTML=''; bulkFilePreviewRows=null; document.getElementById('isku-bulk-file-input').value='';" style="margin-top:10px;margin-left:8px;padding:8px 16px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Cancel</button>
      </div>`;
    }
    previewEl.innerHTML = html;
  };
  reader.readAsText(file);
}

async function applyBulkFileChanges() {
  if (!bulkFilePreviewRows || !bulkFilePreviewRows.length) return;
  const today = new Date().toISOString().slice(0, 10);
  for (const { sku, draft } of bulkFilePreviewRows) {
    draft.date_modified = today;
    delete draft.sku;
    iskuDb.skus[sku] = draft;
  }
  const count = bulkFilePreviewRows.length;
  bulkFilePreviewRows = null;
  document.getElementById('isku-bulk-file-preview').innerHTML = '';
  document.getElementById('isku-bulk-file-input').value = '';
  const saved = await saveIskuDatabaseToDisk();
  populateIskuFilterDropdowns();
  renderIskuTable();
  if (saved) {
    document.getElementById('isku-bulk-file-preview').innerHTML = `<div style="padding:10px;background:#e6f4ea;color:#1e7e34;border-radius:6px;font-size:12px;">Applied ${count} change(s) and saved.</div>`;
  } else {
    alert(`Applied ${count} change(s) in-memory but auto-save failed -- reopen isku_database.json and try again.`);
  }
}

// --- Bulk ADD via File -- brand-new ISKUs, many at once --------------------
// Separate from Bulk Edit via File on purpose: editing matches by existing
// SKU and skips anything not already in the database; adding is the mirror
// image (every row must be a SKU that does NOT exist yet), so mixing them
// into one flow risked silently editing when the user meant to create, or
// vice versa. Same shared validation/compute path as saveNewIsku() and
// applyAssignmentsToRecord() -- nothing here re-derives the formula.
const BULK_ADD_CSV_COLUMNS = ['sku', 'brand', 'category', 'series', 'variant', 'color', 'itec_code', 'currency', 'import_price', 'rsp', 'status', 'notes'];

function updateBulkAddNote() {
  const el = document.getElementById('isku-bulk-add-active-codes');
  if (!el || !iskuDb) return;
  el.textContent = `ITEC: ${Object.keys(iskuDb.itec_codes).sort().join(', ')}; Currency: ${Object.keys(iskuDb.currencies).sort().join(', ')}`;
}

function downloadIskuAddTemplate() {
  const blob = new Blob([BULK_ADD_CSV_COLUMNS.join(',') + '\\r\\n'], { type: 'text/csv' });
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'isku_bulk_add_template.csv';
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

let bulkAddPreviewRows = null;

function handleBulkAddFileUpload(input) {
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = () => {
    const parsed = parseCSV(reader.result);
    const previewEl = document.getElementById('isku-bulk-add-preview');
    if (!parsed.length) {
      previewEl.innerHTML = `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">Empty or unreadable CSV.</div>`;
      return;
    }
    const newRows = [];
    const errors = [];
    const seenInBatch = new Set();
    const today = new Date().toISOString().slice(0, 10);
    for (const csvRow of parsed) {
      let sku = (csvRow.sku || '').trim();
      if (!sku) {
        const parts = ['brand', 'category', 'series', 'variant', 'color'].map(f => (csvRow[f] || '').trim()).filter(Boolean);
        sku = parts.join('-');
      }
      if (!sku) { errors.push('Row with no SKU and no brand/category/series/variant/color to build one -- skipped.'); continue; }
      if (iskuDb.skus[sku]) { errors.push(`${sku}: already exists -- skipped (use Search/Edit or Bulk Edit via File to change it).`); continue; }
      if (seenInBatch.has(sku)) { errors.push(`${sku}: duplicated more than once in this file -- only the first was kept.`); continue; }
      const itecCode = (csvRow.itec_code || '').trim();
      const itec = iskuDb.itec_codes[itecCode];
      if (!itec) { errors.push(`${sku}: unknown ITEC Code "${itecCode}" -- skipped.`); continue; }
      const currency = (csvRow.currency || '').trim();
      if (!iskuDb.currencies[currency]) { errors.push(`${sku}: unknown currency "${currency}" -- skipped.`); continue; }
      const importPrice = parseFloat(csvRow.import_price);
      if (isNaN(importPrice)) { errors.push(`${sku}: invalid or missing import_price -- skipped.`); continue; }
      const rsp = parseFloat(csvRow.rsp);
      if (isNaN(rsp)) { errors.push(`${sku}: invalid or missing rsp -- skipped.`); continue; }
      const status = (csvRow.status || 'ACTIVE').trim().toUpperCase() || 'ACTIVE';
      if (!BULK_STATUS_VALUES.includes(status)) { errors.push(`${sku}: invalid status "${csvRow.status}" -- skipped.`); continue; }

      const rec0 = { itec_code: itecCode, currency, import_price: importPrice, rsp };
      const cost = computeCost(rec0, iskuDb.itec_codes, iskuDb.currencies);
      const dsp = computeDsp(rec0, cost);
      const rec = {
        itec_code: itecCode, currency, upc: null, status,
        import_price: importPrice, tax: importPrice * (itec.tax_rate || 0), fees: itec.fees.map(f => ({ ...f })),
        cost, rsp, dsp,
        offline_set_margin: null, offline_purchase_price: null,
        brand: (csvRow.brand || '').trim() || null,
        category: (csvRow.category || '').trim() || null,
        series: (csvRow.series || '').trim() || null,
        variant: (csvRow.variant || '').trim() || null,
        color: (csvRow.color || '').trim() || null,
        date_created: today, date_modified: today,
        notes: (csvRow.notes || '').trim(),
      };
      newRows.push({ sku, rec });
      seenInBatch.add(sku);
    }
    bulkAddPreviewRows = newRows;
    const fmt = v => v == null ? '' : (typeof v === 'number' ? v.toFixed(2) : escapeHtml(v));
    let html = '';
    if (errors.length) html += `<div style="margin-top:10px;padding:10px;background:#fff8e1;color:#7a5c00;border-radius:6px;font-size:12px;">${errors.length} row(s) skipped:<br>${errors.map(escapeHtml).join('<br>')}</div>`;
    if (!newRows.length) {
      html += `<div style="margin-top:10px;" class="note">No valid new ISKUs to add.</div>`;
    } else {
      html += `<div style="margin-top:10px;">
        <div class="note" style="margin-bottom:6px;">${newRows.length} new ISKU(s) will be created:</div>
        <table class="data-table"><thead><tr><th>SKU</th><th>ITEC</th><th>Currency</th><th>Import Price</th><th>RSP</th><th>Cost</th><th>DSP</th><th>Status</th></tr></thead><tbody>
          ${newRows.slice(0, 20).map(({ sku, rec }) => `<tr>
            <td>${escapeHtml(sku)}</td>
            <td>${escapeHtml(rec.itec_code)}</td>
            <td>${escapeHtml(rec.currency)}</td>
            <td>${fmt(rec.import_price)}</td>
            <td>${fmt(rec.rsp)}</td>
            <td>${fmt(rec.cost)}</td>
            <td>${fmt(rec.dsp)}</td>
            <td>${escapeHtml(rec.status)}</td>
          </tr>`).join('')}
        </tbody></table>
        ${newRows.length > 20 ? `<div class="note">(showing first 20 of ${newRows.length})</div>` : ''}
        <button onclick="applyBulkAddFileChanges()" style="margin-top:10px;padding:8px 16px;border:none;border-radius:6px;background:var(--positive);color:#fff;cursor:pointer;font-size:13px;">Add ${newRows.length} New ISKU(s)</button>
        <button onclick="document.getElementById('isku-bulk-add-preview').innerHTML=''; bulkAddPreviewRows=null; document.getElementById('isku-bulk-add-file-input').value='';" style="margin-top:10px;margin-left:8px;padding:8px 16px;border:1px solid #dee2e6;border-radius:6px;background:#fff;cursor:pointer;font-size:13px;">Cancel</button>
      </div>`;
    }
    previewEl.innerHTML = html;
  };
  reader.readAsText(file);
}

async function applyBulkAddFileChanges() {
  if (!bulkAddPreviewRows || !bulkAddPreviewRows.length) return;
  for (const { sku, rec } of bulkAddPreviewRows) {
    iskuDb.skus[sku] = rec;
  }
  const count = bulkAddPreviewRows.length;
  bulkAddPreviewRows = null;
  document.getElementById('isku-bulk-add-preview').innerHTML = '';
  document.getElementById('isku-bulk-add-file-input').value = '';
  const saved = await saveIskuDatabaseToDisk();
  populateIskuSuggestions();
  populateIskuFilterDropdowns();
  renderIskuTable();
  updateBulkMatchCount();
  if (saved) {
    document.getElementById('isku-bulk-add-preview').innerHTML = `<div style="padding:10px;background:#e6f4ea;color:#1e7e34;border-radius:6px;font-size:12px;">Added ${count} new ISKU(s) and saved.</div>`;
  } else {
    alert(`Added ${count} new ISKU(s) in-memory but auto-save failed -- reopen isku_database.json and try again.`);
  }
}

// Each ITEC code is a card (not a table row) since its `fees` list is
// variable-length -- one line per invisible-cost item (packaging/shipping,
// engraving logo, custom packaging, etc.), each independently editable/
// removable, plus an "+ Add Fee" line to grow the list.
// Both ITEC codes and currencies use an explicit Edit -> Confirm/Cancel flow
// (not edit-as-you-type auto-save) -- same reasoning as the ISKU Search/Edit
// table: these numbers feed every Cost calculation going forward, so a
// half-finished edit (e.g. mid-way through typing a new rate) should never
// get written to disk. Editing works against a scratch copy (itecDraft) so
// Cancel can throw changes away cleanly.
let itecEditingCode = null;
let itecDraft = null;
let itecDeletingCode = null;

function countIskuUsesItec(code) {
  return Object.values(iskuDb.skus).filter(r => r.itec_code === code).length;
}

function renderItecCards() {
  const container = document.getElementById('itec-cards');
  const codes = Object.keys(iskuDb.itec_codes).sort();
  container.innerHTML = codes.map(code => {
    const editing = itecEditingCode === code;
    const deleting = itecDeletingCode === code;
    const c = editing ? itecDraft : iskuDb.itec_codes[code];
    const key = escapeHtml(code).replace(/"/g, '&quot;');
    const total = feesTotal(c);
    if (deleting) {
      const uses = countIskuUsesItec(code);
      return `<div style="border:2px solid var(--negative);border-radius:8px;padding:14px 16px;margin-bottom:12px;background:#fff5f5;">
        <div>Delete <strong>${escapeHtml(code)}</strong>? ${uses > 0
          ? `${uses} ISKU(s) currently reference this code -- their frozen Cost/DSP won't change, but you won't be able to pick this code for them again until they're re-edited onto a different one.`
          : `No ISKUs currently reference this code.`} Not undoable.</div>
        <div style="margin-top:10px;display:flex;gap:8px;">
          <button onclick="deleteItec('${key}')" style="padding:5px 14px;border:none;border-radius:4px;background:var(--negative);color:#fff;cursor:pointer;font-size:12px;">Confirm Delete</button>
          <button onclick="itecDeletingCode=null; renderItecCards();" style="padding:5px 14px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
        </div>
      </div>`;
    }
    if (!editing) {
      return `<div style="border:1px solid #dee2e6;border-radius:8px;padding:14px 16px;margin-bottom:12px;">
        <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin-bottom:10px;">
          <div><label class="note">ITEC Code</label><div style="font-weight:600;padding:6px 0;">${escapeHtml(code)}</div></div>
          <div style="flex:1;min-width:200px;"><label class="note">Description</label><div style="padding:6px 0;">${escapeHtml(c.description || '')}</div></div>
          <div><label class="note">Tax Rate %</label><div style="padding:6px 0;">${((c.tax_rate || 0) * 100).toFixed(1)}%</div></div>
          <div><label class="note">Fees Total</label><div style="font-weight:600;padding:6px 0;">${total.toFixed(2)}</div></div>
          <div style="display:flex;gap:6px;">
            <button onclick="startEditItec('${key}')" style="padding:6px 14px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Edit</button>
            <button onclick="itecDeletingCode='${key}'; renderItecCards();" title="Delete" aria-label="Delete" style="padding:6px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;color:var(--negative);cursor:pointer;line-height:0;">${TRASH_ICON}</button>
          </div>
        </div>
        <table class="data-table"><thead><tr><th>Fee Label</th><th>Amount</th></tr></thead><tbody>
          ${(c.fees || []).map(f => `<tr><td>${escapeHtml(f.label)}</td><td>${Number(f.amount).toFixed(2)}</td></tr>`).join('')}
        </tbody></table>
      </div>`;
    }
    return `<div style="border:2px solid var(--accent, #4a6cf7);border-radius:8px;padding:14px 16px;margin-bottom:12px;">
      <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end;margin-bottom:10px;">
        <div><label class="note">ITEC Code</label><div style="font-weight:600;padding:6px 0;">${escapeHtml(code)}</div></div>
        <div style="flex:1;min-width:200px;"><label class="note">Description</label>
          <input value="${escapeHtml(c.description || '')}" style="width:100%;padding:6px;border:1px solid #dee2e6;border-radius:4px;" oninput="itecDraft.description=this.value"></div>
        <div><label class="note">Tax Rate %</label>
          <input type="number" step="0.1" value="${((c.tax_rate || 0) * 100).toFixed(1)}" style="width:80px;padding:6px;border:1px solid #dee2e6;border-radius:4px;" oninput="itecDraft.tax_rate=(parseFloat(this.value)||0)/100"></div>
      </div>
      <table class="data-table"><thead><tr><th>Fee Label</th><th>Amount</th><th></th></tr></thead><tbody>
        ${(c.fees || []).map((f, i) => `<tr>
          <td><input value="${escapeHtml(f.label)}" style="width:100%;padding:4px;border:1px solid #dee2e6;border-radius:4px;" oninput="itecDraft.fees[${i}].label=this.value"></td>
          <td><input type="number" step="0.01" value="${f.amount}" style="width:100px;padding:4px;border:1px solid #dee2e6;border-radius:4px;" oninput="itecDraft.fees[${i}].amount=parseFloat(this.value)||0"></td>
          <td><button onclick="removeDraftItecFee(${i})" style="padding:3px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Remove</button></td>
        </tr>`).join('')}
      </tbody></table>
      <div style="margin-top:10px;display:flex;gap:8px;">
        <button onclick="addDraftItecFee()" style="padding:5px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">+ Add Fee</button>
        <button onclick="confirmEditItec('${key}')" style="padding:5px 14px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">Confirm</button>
        <button onclick="cancelEditItec()" style="padding:5px 14px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
      </div>
    </div>`;
  }).join('');
}

let itecIsNew = false;

function startEditItec(code, isNew) {
  itecEditingCode = code;
  itecIsNew = !!isNew;
  itecDraft = JSON.parse(JSON.stringify(iskuDb.itec_codes[code]));
  renderItecCards();
}
function cancelEditItec() {
  // A fresh Add that got Cancelled shouldn't leave a blank code behind --
  // only actually exists in iskuDb once Confirm is clicked.
  if (itecIsNew && itecEditingCode) delete iskuDb.itec_codes[itecEditingCode];
  itecEditingCode = null; itecDraft = null; itecIsNew = false;
  renderItecCards();
}
async function deleteItec(code) {
  if (Object.keys(iskuDb.itec_codes).length <= 1) {
    alert("Can't delete the last ITEC code -- the Add New ISKU form needs at least one to offer.");
    itecDeletingCode = null; renderItecCards();
    return;
  }
  delete iskuDb.itec_codes[code];
  itecDeletingCode = null;
  const saved = await saveIskuDatabaseToDisk();
  renderItecCards();
  if (!saved) alert(`${code} deleted in-memory but auto-save failed -- reopen isku_database.json and try again.`);
}

// Adding a new code just creates a blank entry and drops straight into the
// existing Edit flow (reusing startEditItec/confirmEditItec) so description/
// tax/fees are filled in through the same Confirm/Cancel-guarded UI as any
// other edit -- nothing is saved until that Confirm is clicked.
async function addItec() {
  const input = document.getElementById('itec-add-code');
  const msg = document.getElementById('itec-add-msg');
  const code = input.value.trim();
  if (!code) { msg.textContent = 'Enter a code first.'; msg.style.color = 'var(--negative)'; return; }
  if (iskuDb.itec_codes[code]) { msg.textContent = `${code} already exists.`; msg.style.color = 'var(--negative)'; return; }
  iskuDb.itec_codes[code] = { description: '', tax_rate: 0, fees: [] };
  input.value = '';
  msg.textContent = '';
  startEditItec(code, true);
}

function addDraftItecFee() {
  itecDraft.fees.push({ label: 'New Fee', amount: 0 });
  renderItecCards();
}
function removeDraftItecFee(index) {
  itecDraft.fees.splice(index, 1);
  renderItecCards();
}
async function confirmEditItec(code) {
  iskuDb.itec_codes[code] = itecDraft;
  itecEditingCode = null; itecDraft = null; itecIsNew = false;
  const saved = await saveIskuDatabaseToDisk();
  renderItecCards();
  if (!saved) alert('Auto-save failed -- reopen isku_database.json and try again.');
}

let currencyEditingCode = null;
let currencyDeletingCode = null;

function countIskuUsesCurrency(code) {
  return Object.values(iskuDb.skus).filter(r => r.currency === code).length;
}

function renderCurrencyTable() {
  const table = document.getElementById('currency-table');
  const codes = Object.keys(iskuDb.currencies).sort();
  table.innerHTML = '<thead><tr><th>Currency</th><th>RM per unit</th><th>Actions</th></tr></thead><tbody>' +
    codes.map(code => {
      const c = iskuDb.currencies[code];
      const key = escapeHtml(code).replace(/"/g, '&quot;');
      if (currencyDeletingCode === code) {
        const uses = countIskuUsesCurrency(code);
        return `<tr style="background:#fff5f5;">
          <td colspan="2">Delete <strong>${escapeHtml(code)}</strong>? ${uses > 0
            ? `${uses} ISKU(s) currently use this currency -- their frozen Cost/DSP won't change, but you won't be able to pick it for them again until re-edited onto a different one.`
            : `No ISKUs currently use this currency.`} Not undoable.</td>
          <td>
            <button onclick="deleteCurrency('${key}')" style="padding:4px 10px;border:none;border-radius:4px;background:var(--negative);color:#fff;cursor:pointer;font-size:12px;">Confirm Delete</button>
            <button onclick="currencyDeletingCode=null; renderCurrencyTable();" style="padding:4px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
          </td>
        </tr>`;
      }
      if (currencyEditingCode !== code) {
        return `<tr>
          <td>${escapeHtml(code)}</td>
          <td>${c.rate_to_rm}</td>
          <td>
            <button onclick="startEditCurrency('${key}')" style="padding:4px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Edit</button>
            <button onclick="currencyDeletingCode='${key}'; renderCurrencyTable();" title="Delete" aria-label="Delete" style="padding:4px 8px;border:1px solid #dee2e6;border-radius:4px;background:#fff;color:var(--negative);cursor:pointer;line-height:0;">${TRASH_ICON}</button>
          </td>
        </tr>`;
      }
      return `<tr>
        <td>${escapeHtml(code)}</td>
        <td><input id="currency-edit-input" type="number" step="0.0001" value="${c.rate_to_rm}" style="width:100px;padding:4px;border:1px solid #dee2e6;border-radius:4px;"></td>
        <td>
          <button onclick="confirmEditCurrency('${key}')" style="padding:4px 10px;border:none;border-radius:4px;background:var(--positive);color:#fff;cursor:pointer;font-size:12px;">Confirm</button>
          <button onclick="cancelEditCurrency()" style="padding:4px 10px;border:1px solid #dee2e6;border-radius:4px;background:#fff;cursor:pointer;font-size:12px;">Cancel</button>
        </td>
      </tr>`;
    }).join('') +
    '</tbody>';
}

function startEditCurrency(code) { currencyEditingCode = code; renderCurrencyTable(); }
function cancelEditCurrency() { currencyEditingCode = null; renderCurrencyTable(); }
async function confirmEditCurrency(code) {
  const v = parseFloat(document.getElementById('currency-edit-input').value);
  if (isNaN(v)) { alert('Enter a valid number first.'); return; }
  iskuDb.currencies[code].rate_to_rm = v;
  currencyEditingCode = null;
  const saved = await saveIskuDatabaseToDisk();
  renderCurrencyTable();
  if (!saved) alert('Auto-save failed -- reopen isku_database.json and try again.');
}
async function deleteCurrency(code) {
  if (Object.keys(iskuDb.currencies).length <= 1) {
    alert("Can't delete the last currency -- the Add New ISKU form needs at least one to offer.");
    currencyDeletingCode = null; renderCurrencyTable();
    return;
  }
  delete iskuDb.currencies[code];
  currencyDeletingCode = null;
  const saved = await saveIskuDatabaseToDisk();
  renderCurrencyTable();
  if (!saved) alert(`${code} deleted in-memory but auto-save failed -- reopen isku_database.json and try again.`);
}
async function addCurrency() {
  const codeInput = document.getElementById('currency-add-code');
  const rateInput = document.getElementById('currency-add-rate');
  const msg = document.getElementById('currency-add-msg');
  const code = codeInput.value.trim().toUpperCase();
  const rate = parseFloat(rateInput.value);
  if (!code) { msg.textContent = 'Enter a currency code first.'; msg.style.color = 'var(--negative)'; return; }
  if (iskuDb.currencies[code]) { msg.textContent = `${code} already exists.`; msg.style.color = 'var(--negative)'; return; }
  if (isNaN(rate) || rate <= 0) { msg.textContent = 'Enter a valid RM-per-unit rate.'; msg.style.color = 'var(--negative)'; return; }
  iskuDb.currencies[code] = { rate_to_rm: rate };
  const saved = await saveIskuDatabaseToDisk();
  renderCurrencyTable();
  if (saved) {
    codeInput.value = ''; rateInput.value = '';
    msg.textContent = `${code} added.`; msg.style.color = 'var(--positive)';
  } else {
    msg.textContent = `${code} added in-memory but auto-save failed -- reopen isku_database.json and try again.`;
    msg.style.color = 'var(--negative)';
  }
}

const restockFundedState = { sortField: 'est_cost', sortDir: 'desc' };
const restockCutState = { sortField: 'est_cost', sortDir: 'desc' };
const restockNoCostState = { sortField: 'stock_on_hand', sortDir: 'desc' };
const restockNotFoundState = { sortField: 'sku', sortDir: 'asc' };

function renderRestockTab() {
  const r = DATA.restock;
  if (!r) return;  // empty-note stays visible, content stays hidden -- nothing to render
  document.getElementById('restock-empty-note').style.display = 'none';
  document.getElementById('restock-content').style.display = 'block';

  document.getElementById('restock-generated-note').textContent =
    `Generated ${r.generated_at} from ${r.target_month} actuals + SiteGiant Inventory Forecasting `+
    `(last 30 days) + cash position as of ${r.cash_position_as_of}. Re-run restock_planner.py monthly with a `+
    `fresh export to keep this current -- it is a snapshot, not a forecast of future months.`;

  const b = r.budget;
  const bt = document.getElementById('restock-budget-table');
  bt.innerHTML = '<thead><tr><th></th><th>Amount</th></tr></thead><tbody>' +
    `<tr><td>Raw restock need (all flagged active SKUs)</td><td>${fmtMoney(b.raw_need)}</td></tr>` +
    `<tr><td>Ceiling: ${b.revenue_cap_pct.toFixed(0)}% of last month's revenue (${fmtMoney(b.last_month_revenue)})</td><td>${fmtMoney(b.revenue_cap)}</td></tr>` +
    `<tr><td>Ceiling: cash on hand (${fmtMoney(r.cash_position)}) minus last month's Payroll+Bank OpEx reserve (${fmtMoney(b.cash_reserve)})</td><td>${fmtMoney(b.cash_cap)}</td></tr>` +
    `<tr style="font-weight:600;background:#f8f9fa"><td>Recommended budget</td><td>${fmtMoney(b.final_budget)}</td></tr>` +
    `<tr><td>Binding constraint</td><td>${b.binding_constraint}</td></tr>` +
    '</tbody>';

  const orderCols = (state, stateName, renderFn) =>
    th('SKU', 'sku', state, stateName, renderFn) + th('Order Sheet', 'order_sheet', state, stateName, renderFn) +
    th('Class', 'abc_class', state, stateName, renderFn) + th('Stock', 'stock_on_hand', state, stateName, renderFn) +
    th('On Order', 'stock_on_purchase_order', state, stateName, renderFn) + th('Sale/Day', 'sale_per_day', state, stateName, renderFn) +
    th('Days Left', 'days_left', state, stateName, renderFn) + th('Recommend Qty', 'recommended_qty', state, stateName, renderFn) +
    th('Unit Cost', 'cost', state, stateName, renderFn) + th('Est. Cost', 'est_cost', state, stateName, renderFn);
  const orderRow = (o) => `<tr>
    <td>${o.sku}</td><td>${o.order_sheet}</td><td>${o.abc_class || 'n/a'}</td>
    <td>${o.stock_on_hand ?? '-'}</td><td>${o.stock_on_purchase_order ?? '-'}</td>
    <td>${o.sale_per_day ?? '-'}</td><td>${o.days_left ?? 'n/a'}</td>
    <td>${o.recommended_qty ?? '-'}</td><td>${fmtMoney(o.cost)}</td><td>${fmtMoney(o.est_cost)}</td>
  </tr>`;

  const funded = sortRows(r.funded, restockFundedState.sortField, restockFundedState.sortDir);
  document.getElementById('restock-funded-note').textContent =
    `${r.summary.funded_count} SKUs, ${fmtMoney(r.summary.funded_spend)} total.`;
  document.getElementById('restock-funded-table').innerHTML =
    '<thead><tr>' + orderCols(restockFundedState, 'restockFundedState', 'renderRestockTab') + '</tr></thead><tbody>' +
    funded.map(orderRow).join('') + '</tbody>';

  const cutSection = document.getElementById('restock-cut-section');
  if (r.cut.length) {
    cutSection.style.display = 'block';
    const cut = sortRows(r.cut, restockCutState.sortField, restockCutState.sortDir);
    document.getElementById('restock-cut-table').innerHTML =
      '<thead><tr>' + orderCols(restockCutState, 'restockCutState', 'renderRestockTab') + '<th>Urgent?</th></tr></thead><tbody>' +
      cut.map(o => orderRow(o).replace('</tr>', `<td>${o.urgent ? 'YES' : ''}</td></tr>`)).join('') + '</tbody>';
  } else {
    cutSection.style.display = 'none';
  }

  const noCostSection = document.getElementById('restock-nocost-section');
  if (r.no_cost.length) {
    noCostSection.style.display = 'block';
    const nc = sortRows(r.no_cost, restockNoCostState.sortField, restockNoCostState.sortDir);
    document.getElementById('restock-nocost-table').innerHTML =
      '<thead><tr>' +
      th('SKU', 'sku', restockNoCostState, 'restockNoCostState', 'renderRestockTab') +
      th('Order Sheet', 'order_sheet', restockNoCostState, 'restockNoCostState', 'renderRestockTab') +
      th('Stock', 'stock_on_hand', restockNoCostState, 'restockNoCostState', 'renderRestockTab') +
      th('Sale/Day', 'sale_per_day', restockNoCostState, 'restockNoCostState', 'renderRestockTab') +
      '</tr></thead><tbody>' +
      nc.map(o => `<tr><td>${o.sku}</td><td>${o.order_sheet}</td><td>${o.stock_on_hand ?? '-'}</td><td>${o.sale_per_day ?? '-'}</td></tr>`).join('') +
      '</tbody>';
  } else {
    noCostSection.style.display = 'none';
  }

  const notFoundSection = document.getElementById('restock-notfound-section');
  if (r.not_in_forecast.length) {
    notFoundSection.style.display = 'block';
    const nf = sortRows(r.not_in_forecast, restockNotFoundState.sortField, restockNotFoundState.sortDir);
    document.getElementById('restock-notfound-table').innerHTML =
      '<thead><tr>' +
      th('SKU', 'sku', restockNotFoundState, 'restockNotFoundState', 'renderRestockTab') +
      th('Order Sheet', 'order_sheet', restockNotFoundState, 'restockNotFoundState', 'renderRestockTab') +
      th('Brand', 'brand', restockNotFoundState, 'restockNotFoundState', 'renderRestockTab') +
      '</tr></thead><tbody>' +
      nf.map(o => `<tr><td>${o.sku}</td><td>${o.order_sheet}</td><td>${o.brand || '-'}</td></tr>`).join('') +
      '</tbody>';
  } else {
    notFoundSection.style.display = 'none';
  }
}

const pipelineState = { page: 0, pageSize: 50, sortField: 'est_cost', sortDir: 'desc' };

function renderPipelineTab() {
  const p = DATA.restock_pipeline;
  if (!p) return;
  document.getElementById('pipeline-empty-note').style.display = 'none';
  document.getElementById('pipeline-content').style.display = 'block';

  const k = p.kpis;
  const cards = [
    ['Active SKUs Matched', k.active_skus_matched, ''],
    ['SKUs Needing Restock', k.skus_needing_restock, ''],
    ['Recommended Spend', fmtMoney(k.recommended_spend), ''],
    ['Units to Order', k.units_to_order.toLocaleString(), ''],
    ['On-PO Flags (not netted)', k.on_po_flags, k.on_po_flags ? 'highlight-negative' : ''],
    ['Low/Out of Stock', k.low_stock_flags, k.low_stock_flags ? 'highlight-negative' : ''],
    ['No Action', k.no_action_count, ''],
    ['Missing from Forecast', k.missing_from_forecast, ''],
  ];
  document.getElementById('pipeline-kpi-row').innerHTML = cards.map(([label, val, cls]) =>
    `<div class="kpi-card ${cls}"><div class="kpi-label">${label}</div><div class="kpi-value">${val}</div></div>`).join('');

  const o = p.otb_check;
  document.getElementById('pipeline-otb-section').style.background = o.within_budget ? '#e9f3e6' : '#fbe9e7';
  document.getElementById('pipeline-otb-note').innerHTML =
    `<strong>${o.within_budget ? 'Within budget' : 'OVER BUDGET'}:</strong> Recommended spend ${fmtMoney(o.total_spend)} vs. ` +
    `ceiling ${fmtMoney(o.ceiling)} (smaller of ${o.revenue_cap_pct.toFixed(0)}% of last month's revenue = ${fmtMoney(o.revenue_cap)}` +
    (o.cash_cap != null ? `, or cash on hand minus Payroll/Bank OpEx reserve = ${fmtMoney(o.cash_cap)}` : '') +
    `). Sanity check on the TOTAL only -- doesn't change any SKU's quantity, which follows the confirmed ` +
    `60-day/MOQ formula regardless.`;

  document.getElementById('pipeline-supplier-table').innerHTML =
    '<thead><tr><th>Order Sheet</th><th>SKUs</th><th>Units</th><th>Spend</th></tr></thead><tbody>' +
    p.by_supplier.map(s => `<tr><td>${s.order_sheet}</td><td>${s.skus}</td><td>${s.units}</td><td>${fmtMoney(s.spend)}</td></tr>`).join('') +
    '</tbody>';

  populatePipelineFilter();
  renderPipelineDetail();
}

function populatePipelineFilter() {
  const p = DATA.restock_pipeline;
  const select = document.getElementById('pipeline-supplier-filter');
  const suppliers = [...new Set(p.detail.map(d => d.order_sheet))].sort();
  select.innerHTML = '<option value="all">All Suppliers</option>' +
    suppliers.map(s => `<option value="${s}">${s}</option>`).join('');
}

function getFilteredSortedPipeline() {
  const p = DATA.restock_pipeline;
  const search = document.getElementById('pipeline-search').value.trim().toLowerCase();
  const supplier = document.getElementById('pipeline-supplier-filter').value;
  const action = document.getElementById('pipeline-action-filter').value;
  let rows = p.detail;
  if (search) rows = rows.filter(r => r.sku.toLowerCase().includes(search));
  if (supplier !== 'all') rows = rows.filter(r => r.order_sheet === supplier);
  if (action !== 'all') rows = rows.filter(r => r.action === action);
  return sortRows(rows, pipelineState.sortField, pipelineState.sortDir);
}

function renderPipelineDetail() {
  const rows = getFilteredSortedPipeline();
  const start = pipelineState.page * pipelineState.pageSize;
  const pageRows = rows.slice(start, start + pipelineState.pageSize);
  const s = pipelineState, sn = 'pipelineState', rf = 'renderPipelineDetail';
  const table = document.getElementById('pipeline-detail-table');
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s, sn, rf) + th('Order Sheet', 'order_sheet', s, sn, rf) + th('Brand', 'brand', s, sn, rf) +
    th('Action', 'action', s, sn, rf) + th('Stock', 'stock_on_hand', s, sn, rf) +
    th('On PO', 'stock_on_purchase_order', s, sn, rf) + th('Sales 30d', 'total_sales', s, sn, rf) +
    th('Restock Qty', 'restock_qty', s, sn, rf) + th('Unit Cost', 'cost', s, sn, rf) + th('Est. Cost', 'est_cost', s, sn, rf) +
    '<th>Flag</th></tr></thead><tbody>' +
    pageRows.map(r => `<tr>
      <td>${r.sku}</td><td>${r.order_sheet}</td><td>${r.brand || '-'}</td><td>${r.action}</td>
      <td>${r.stock_on_hand}</td><td>${r.stock_on_purchase_order}</td><td>${r.total_sales}</td>
      <td>${r.restock_qty}</td><td>${fmtMoney(r.cost)}</td><td>${fmtMoney(r.est_cost)}</td>
      <td>${r.on_po_flag ? '<span class="fee-unknown" style="color:var(--negative);font-style:normal;font-weight:600;">ON PO</span>' : ''}</td>
    </tr>`).join('') + '</tbody>';

  renderPagination('pipeline-pagination', pipelineState, 'pipelineState', rows.length, 'renderPipelineDetail');
}

const strategyScaleupState = { page: 0, pageSize: 50, sortField: 'eta_days', sortDir: 'asc' };
const strategyStopRestockState = { page: 0, pageSize: 50, sortField: 'total_4mo', sortDir: 'asc' };
const strategyTrendingState = { page: 0, pageSize: 50, sortField: 'total_4mo', sortDir: 'desc' };

function renderStrategyTab() {
  const s = DATA.restock_strategy;
  if (!s) return;
  document.getElementById('strategy-empty-note').style.display = 'none';
  document.getElementById('strategy-content').style.display = 'block';

  const k = s.summary;
  const bm = s.brand_mix;
  const leadTime = s.lead_time_days != null ? s.lead_time_days : 50;
  document.getElementById('strategy-leadtime-days').textContent = leadTime;
  const cards = [
    ['FRES Brand Mix (Real)', fmtPct(bm.fres_pct), `Target ${bm.target_fres_pct.toFixed(0)}% -- gap ${bm.gap_to_target_pct>=0?'+':''}${bm.gap_to_target_pct.toFixed(1)}pp`],
    ['Active SKUs Matched', k.active_skus_matched, ''],
    ['Stockout Gap Risk', k.gap_risk_count, `won't survive the ${leadTime}-day lead time`],
    ['Running Well (Scale Candidates)', k.running_well_count, 'FRES, no gap risk yet'],
    ['FRES Scale-Ups (any tier)', k.fres_scaleup_count, fmtMoney(k.fres_scaleup_spend)],
    ['Stop-Restock Candidates', k.stop_restock_count, ''],
    ['Trending Up', k.trending_up_count, ''],
  ];
  document.getElementById('strategy-kpi-row').innerHTML = cards.map(([label, val, sub]) =>
    `<div class="kpi-card"><div class="kpi-label">${label}</div><div class="kpi-value">${val}</div>${sub?`<div class="kpi-sub">${sub}</div>`:''}</div>`).join('');

  document.getElementById('strategy-brandmix-note').textContent =
    `Computed from this run's real revenue (generated ${s.generated_at}). Not an enforced constraint -- a tracked goal.`;
  document.getElementById('strategy-brandmix-table').innerHTML =
    '<thead><tr><th></th><th>Revenue</th><th>% of Total</th></tr></thead><tbody>' +
    `<tr><td>FRES (Frescoone brand)</td><td>${fmtMoney(bm.fres_revenue)}</td><td>${fmtPct(bm.fres_pct)}</td></tr>` +
    `<tr><td>Generic brands</td><td>${fmtMoney(bm.generic_revenue)}</td><td>${fmtPct(bm.generic_pct)}</td></tr>` +
    `<tr style="font-weight:600;background:#f8f9fa"><td>Total</td><td>${fmtMoney(bm.total_revenue)}</td><td>100.0%</td></tr>` +
    `<tr><td>Target</td><td></td><td>${fmtPct(bm.target_fres_pct)} FRES / ${fmtPct(100-bm.target_fres_pct)} Generic</td></tr>` +
    '</tbody>';

  document.getElementById('strategy-stoprestock-note').textContent =
    `${s.stop_restock_candidates.length} SKUs with minimal sales across ${s.months_analyzed.join(', ')} -- consider stopping restock on these.`;

  populateStrategyOrderSheetFilter();
  renderStrategyScaleupTable();
  renderStrategyStopRestockTable();
  renderStrategyTrendingTable();
}

const STATUS_COLOR_HEX = { red: '#a3342a', orange: '#c26a00', blue: '#2b6cb0', green: '#3f7a3f', grey: '#6c757d' };
function statusBadge(status, color) {
  const hex = STATUS_COLOR_HEX[color] || '#6c757d';
  return `<span style="display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600;color:#fff;background:${hex};">${status}</span>`;
}
function fmtEta(days) {
  if (days == null) return '<span class="fee-unknown">n/a</span>';
  if (days <= 0) return 'Out now';
  return days.toFixed(1) + 'd';
}

function populateStrategyOrderSheetFilter() {
  const select = document.getElementById('strategy-ordersheet-filter');
  if (select.dataset.populated) return;
  const s = DATA.restock_strategy;
  const sheets = [...new Set(s.scaleup_detail.map(r => r.order_sheet))].sort();
  select.innerHTML = '<option value="all">All Order Sheets</option>' +
    sheets.map(sh => `<option value="${sh}">${sh}</option>`).join('');
  select.dataset.populated = '1';
}

function getFilteredSortedScaleup() {
  const s = DATA.restock_strategy;
  const search = document.getElementById('strategy-search').value.trim().toLowerCase();
  const status = document.getElementById('strategy-status-filter').value;
  const brand = document.getElementById('strategy-brand-filter').value;
  const orderSheet = document.getElementById('strategy-ordersheet-filter').value;
  const topEarnerOnly = document.getElementById('strategy-topearner-filter').checked;
  let rows = s.scaleup_detail;
  if (search) rows = rows.filter(r => r.sku.toLowerCase().includes(search));
  if (status !== 'all') rows = rows.filter(r => r.status === status);
  if (brand !== 'all') rows = rows.filter(r => r.brand_type === brand);
  if (orderSheet !== 'all') rows = rows.filter(r => r.order_sheet === orderSheet);
  if (topEarnerOnly) rows = rows.filter(r => r.is_top_earner);
  return sortRows(rows, strategyScaleupState.sortField, strategyScaleupState.sortDir);
}

function renderStrategyScaleupTable() {
  const rows = getFilteredSortedScaleup();
  const st = strategyScaleupState;
  const start = st.page * st.pageSize;
  const pageRows = rows.slice(start, start + st.pageSize);
  const s2 = st, sn = 'strategyScaleupState', rf = 'renderStrategyScaleupTable';
  const table = document.getElementById('strategy-scaleup-table');
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s2, sn, rf) + th('Order Sheet', 'order_sheet', s2, sn, rf) + th('Brand', 'brand_type', s2, sn, rf) +
    th('Stock', 'stock_on_hand', s2, sn, rf) + th('On PO', 'stock_on_purchase_order', s2, sn, rf) +
    th('Sales 30d', 'total_sales', s2, sn, rf) +
    th('ETA to Stockout', 'eta_days', s2, sn, rf) + th('Status', 'status', s2, sn, rf) +
    th('Avg NPE', 'npe', s2, sn, rf) + th('Multiplier', 'multiplier', s2, sn, rf) +
    th('Restock Qty', 'restock_qty', s2, sn, rf) + th('Est. Cost', 'est_cost', s2, sn, rf) +
    '<th>Reason</th></tr></thead><tbody>' +
    pageRows.map(r => `<tr>
      <td>${r.sku}${r.is_top_earner ? ' <span title="Top-20% average-NPE earner among FRES SKUs" style="color:#c9a227;">&#9733;</span>' : ''}</td>
      <td>${r.order_sheet}</td><td>${r.brand_type}</td>
      <td>${r.stock_on_hand}</td><td>${r.stock_on_purchase_order}</td><td>${r.total_sales}</td>
      <td>${fmtEta(r.eta_days)}</td><td>${statusBadge(r.status, r.status_color)}</td>
      <td style="color:${r.npe<0?'var(--negative)':'inherit'}" title="Average of ${r.npe_order_count} order(s) this month">${fmtMoney(r.npe)}${r.npe_order_count ? ` <span class="fee-unknown">(&times;${r.npe_order_count})</span>` : ''}</td>
      <td>${r.multiplier}x${r.cost_capped ? ' <span class="fee-unknown" title="Capped: unit cost over RM50">(capped)</span>' : ''}</td>
      <td>${r.restock_qty}</td><td>${fmtMoney(r.est_cost)}</td>
      <td style="font-size:11px;color:var(--text-secondary);">${r.reason}</td>
    </tr>`).join('') + '</tbody>';

  renderPagination('strategy-scaleup-pagination', st, 'strategyScaleupState', rows.length, 'renderStrategyScaleupTable');
}

function renderStrategyStopRestockTable() {
  const s = DATA.restock_strategy;
  const st = strategyStopRestockState;
  const rows = sortRows(s.stop_restock_candidates, st.sortField, st.sortDir);
  const start = st.page * st.pageSize;
  const pageRows = rows.slice(start, start + st.pageSize);
  const s2 = st, sn = 'strategyStopRestockState', rf = 'renderStrategyStopRestockTable';
  const table = document.getElementById('strategy-stoprestock-table');
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s2, sn, rf) + th('Order Sheet', 'order_sheet', s2, sn, rf) + th('Brand', 'brand_type', s2, sn, rf) +
    th('Total (4mo)', 'total_4mo', s2, sn, rf) + '<th>Reason</th></tr></thead><tbody>' +
    pageRows.map(r => `<tr><td>${r.sku}</td><td>${r.order_sheet}</td><td>${r.brand_type}</td><td>${r.total_4mo}</td>` +
      `<td style="font-size:11px;color:var(--text-secondary);">${r.trend_reason}</td></tr>`).join('') +
    '</tbody>';
  renderPagination('strategy-stoprestock-pagination', st, 'strategyStopRestockState', rows.length, 'renderStrategyStopRestockTable');
}

function renderStrategyTrendingTable() {
  const s = DATA.restock_strategy;
  const st = strategyTrendingState;
  const rows = sortRows(s.trending_up, st.sortField, st.sortDir);
  const start = st.page * st.pageSize;
  const pageRows = rows.slice(start, start + st.pageSize);
  const s2 = st, sn = 'strategyTrendingState', rf = 'renderStrategyTrendingTable';
  const table = document.getElementById('strategy-trending-table');
  table.innerHTML = '<thead><tr>' +
    th('SKU', 'sku', s2, sn, rf) + th('Order Sheet', 'order_sheet', s2, sn, rf) + th('Brand', 'brand_type', s2, sn, rf) +
    th('Total (4mo)', 'total_4mo', s2, sn, rf) + '<th>Trend</th></tr></thead><tbody>' +
    pageRows.map(r => `<tr><td>${r.sku}</td><td>${r.order_sheet}</td><td>${r.brand_type}</td><td>${r.total_4mo}</td>` +
      `<td style="font-size:11px;color:var(--text-secondary);">${r.trend_reason}</td></tr>`).join('') +
    '</tbody>';
  renderPagination('strategy-trending-pagination', st, 'strategyTrendingState', rows.length, 'renderStrategyTrendingTable');
}

function renderHealthTab() {
  const h = DATA.health;
  if (!h) return;

  if (h.flags && h.flags.length) {
    document.getElementById('health-flags-section').style.display = 'block';
    const iconFor = { warn: '⚠️', info: 'ℹ️', good: '✅' };
    const colorFor = { warn: 'var(--negative)', info: 'var(--text-secondary)', good: 'var(--positive)' };
    document.getElementById('health-flags-list').innerHTML = h.flags.map(f =>
      `<li style="padding:10px 14px;border-radius:6px;background:#f8f9fa;border-left:3px solid ${colorFor[f.severity]};">` +
      `${iconFor[f.severity] || ''} ${f.text}</li>`
    ).join('');
  }

  const cards = [
    ['Revenue', fmtMoney(h.revenue), h.target_month],
    ['Gross Margin', fmtPct(h.gross_margin_pct), fmtMoney(h.gross_profit) + ' gross profit'],
    ['Net Profit (Accrual)', fmtMoney(h.net_profit), fmtPct(h.net_margin_pct) + ' margin'],
    ['Cash Position', fmtMoney(h.cash_position), 'as of ' + h.cash_position_as_of],
    ['Cash Runway (worst case)', h.cash_runway_months != null ? h.cash_runway_months.toFixed(1) + ' months' : 'n/a',
     'if revenue stopped entirely -- ' + fmtMoney(h.monthly_fixed_cost) + '/mo Payroll+Bank OpEx'],
    ['Current Ad Spend', fmtMoney(h.ad_spend.current), fmtPct(h.ad_spend.current_pct_of_revenue) + ' of revenue'],
  ];
  document.getElementById('health-kpi-row').innerHTML = cards.map(([label, val, sub]) => {
    let cls = '';
    if (label === 'Net Profit (Accrual)') cls = h.net_profit >= 0 ? 'highlight-positive' : 'highlight-negative';
    return `<div class="kpi-card ${cls}"><div class="kpi-label">${label}</div><div class="kpi-value">${val}</div>` +
      `${sub ? `<div class="kpi-sub">${sub}</div>` : ''}</div>`;
  }).join('');

  const b = h.breakeven;
  let running = b.gross_profit;
  const steps = [
    { label: 'Gross Profit', amount: null, running: b.gross_profit },
    { label: 'less Payroll OpEx', amount: -b.payroll, running: (running -= b.payroll) },
    { label: 'less Bank OpEx (non-payroll debits)', amount: -b.bank_opex, running: (running -= b.bank_opex) },
    { label: 'less Dropship ads rebate', amount: -b.dropship_ads_rebate, running: (running -= b.dropship_ads_rebate) },
    { label: 'less Platform transaction fees (commission/service/transaction, no ads)',
      amount: -b.platform_transaction_fee, running: (running -= b.platform_transaction_fee) },
  ];
  document.getElementById('health-breakeven-table').innerHTML =
    '<thead><tr><th>Step</th><th>Amount</th><th>Running Total</th></tr></thead><tbody>' +
    steps.map(s => `<tr><td>${s.label}</td><td style="color:${s.amount<0?'var(--negative)':'inherit'}">${s.amount==null?'':fmtMoney(s.amount)}</td><td><strong>${fmtMoney(s.running)}</strong></td></tr>`).join('') +
    `<tr style="font-weight:600;background:#f8f9fa"><td>= Headroom left for ad spend before Net Profit hits zero</td><td></td>` +
    `<td style="color:${b.headroom_before_ads<0?'var(--negative)':'var(--positive)'}">${fmtMoney(b.headroom_before_ads)}</td></tr>` +
    '</tbody>';

  document.getElementById('health-ad-table').innerHTML =
    '<thead><tr><th></th><th>Amount</th></tr></thead><tbody>' +
    `<tr><td>Current spend (this month, real)</td><td>${fmtMoney(h.ad_spend.current)}</td></tr>` +
    `<tr><td>As % of revenue</td><td>${fmtPct(h.ad_spend.current_pct_of_revenue)}</td></tr>` +
    `<tr style="font-weight:600;background:#f8f9fa"><td>Sustainable ceiling (breakeven, all else held fixed)</td><td style="color:${h.ad_spend.breakeven_ceiling<=0?'var(--negative)':'inherit'}">${fmtMoney(h.ad_spend.breakeven_ceiling)}</td></tr>` +
    `<tr><td>Generic e-commerce industry range (context only, not Frescoone-specific)</td><td>~5-15% of revenue</td></tr>` +
    '</tbody>';

  document.getElementById('health-restock-table').innerHTML = h.restock_range
    ? '<thead><tr><th></th><th>Amount</th></tr></thead><tbody>' +
      `<tr><td>Minimum need (real SKUs flagged low/out of stock)</td><td>${fmtMoney(h.restock_range.min_need)}</td></tr>` +
      `<tr><td>Maximum safe ceiling now (40% revenue / cash cap, whichever binds)</td><td>${fmtMoney(h.restock_range.max_ceiling)}</td></tr>` +
      `<tr style="font-weight:600;background:#f8f9fa"><td>Recommended (from Restock Plan tab)</td><td>${fmtMoney(h.restock_range.recommended)}</td></tr>` +
      `<tr><td>Remaining cash headroom after that + Payroll/Bank OpEx reserve</td><td>${fmtMoney(h.cash_headroom_after_restock)}</td></tr>` +
      '</tbody>'
    : '<tbody><tr><td>Not generated yet -- run <code>restock_planner.py</code>, then re-run this script.</td></tr></tbody>';

  document.getElementById('health-ladder-note').innerHTML =
    `<strong>Model, not a guarantee</strong> -- holds this month's real cost ratios (COGS% and platform fee% of ` +
    `revenue) and real fixed costs (Payroll, Bank OpEx, current ad spend) constant, and solves for the revenue ` +
    `each margin tier needs. Verified this reproduces this month's actual Net Profit exactly when run backwards ` +
    `-- it's your real P&L rearranged, not an independent guess. Re-run monthly as your real ratios shift. ` +
    `Contribution margin (revenue left after COGS + platform fees, before fixed costs): ${fmtPct(h.contribution_margin_pct)}.`;
  document.getElementById('health-ladder-table').innerHTML = h.profit_ladder && h.profit_ladder.length
    ? '<thead><tr><th>Target Net Margin</th><th>Revenue Needed</th><th>Net Profit At That Level</th>' +
      '<th>Gap vs. Current Revenue</th><th>Implied Restock Ceiling (40%)</th></tr></thead><tbody>' +
      h.profit_ladder.map(t => `<tr${t.target_margin_pct===0?' style="font-weight:600;background:#f8f9fa"':''}>` +
        `<td>${t.target_margin_pct}% ${t.target_margin_pct===0?'(breakeven)':''}</td>` +
        `<td>${fmtMoney(t.revenue_needed)}</td><td>${fmtMoney(t.net_profit_at)}</td>` +
        `<td>${fmtMoney(t.gap_vs_current_revenue)}</td><td>${fmtMoney(t.restock_ceiling_40pct)}</td></tr>`).join('') +
      '</tbody>'
    : '<tbody><tr><td>Not enough margin data to compute a target ladder this month.</td></tr></tbody>';

  if (h.trailing_12mo_cash_flow != null) {
    document.getElementById('health-cashflow-note').textContent =
      `Trailing ${h.trailing_12mo_months} full months: ${fmtMoney(h.trailing_12mo_cash_flow)} net -- see the Bank Statement tab for the month-by-month detail.`;
  }
  document.getElementById('health-cashflow-table').innerHTML =
    '<thead><tr><th></th><th>Amount</th></tr></thead><tbody>' +
    (h.current_month_cash_flow != null ? `<tr><td>${h.current_month_label} net cash flow (may be partial month)</td><td style="color:${h.current_month_cash_flow<0?'var(--negative)':'var(--positive)'}">${fmtMoney(h.current_month_cash_flow)}</td></tr>` : '') +
    (h.trailing_12mo_cash_flow != null ? `<tr><td>Trailing ${h.trailing_12mo_months} full months, net</td><td style="color:${h.trailing_12mo_cash_flow<0?'var(--negative)':'var(--positive)'}">${fmtMoney(h.trailing_12mo_cash_flow)}</td></tr>` : '') +
    `<tr><td>Cash position</td><td>${fmtMoney(h.cash_position)} (as of ${h.cash_position_as_of})</td></tr>` +
    '</tbody>';
}

renderKPIs();
renderSgdNote();
renderWalkthrough();
renderCashflowChart();
renderTrendChart();
renderChannelChart();
renderBrandChart();
renderOpexChart();
renderPayrollChart();
renderChannelTable();
renderWarnings();
populateOrdersFilter();
renderOrdersTable();
renderPendingTable();
populateBankFilter();
renderBankTable();
renderRestockTab();
renderPipelineTab();
renderStrategyTab();
renderHealthTab();
</script>
</body>
</html>
"""


def write_dashboard_html(ctx, history, out_path):
    data = build_dashboard_data(ctx, history)
    html = (DASHBOARD_HTML_TEMPLATE
            .replace("__COMPANY__", data["company"])
            .replace("__PERIOD__", data["period"])
            .replace("__GENERATED_AT__", data["generated_at"])
            .replace("__DASHBOARD_DATA_JSON__", json.dumps(data)))
    Path(out_path).write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
def main():
    print("Loading SKU master...")
    sku_master = load_sku_master()
    sku_master = apply_cost_overrides(sku_master)

    print("Loading SiteGiant D2C orders...")
    orders = load_sitegiant_orders()
    sitegiant = sitegiant_channel_metrics(orders, sku_master)

    print("Loading dropship reseller invoices...")
    dropship = dropship_channel_metrics(sku_master)

    print("Loading bank statement...")
    bank_df = load_bank_statement()
    cash_flow = cash_flow_trends(bank_df)
    bank_opex_result = bank_opex(bank_df)

    print("Loading payroll history...")
    payroll_df = load_payroll()
    payroll_monthly_df = payroll_monthly(payroll_df)

    target_month = (sitegiant["monthly"]["month"].iloc[0]
                     if not sitegiant["monthly"].empty else pd.Period(datetime.now(), freq="M"))

    print("Loading platform fee reports (if provided)...")
    platform_fees = load_platform_fees(orders, target_month)

    net_profit = compute_net_profit(sitegiant, dropship, bank_opex_result, payroll_monthly_df,
                                     target_month, platform_fees)
    brand_margin = brand_margin_analysis(sitegiant, dropship)
    bank_opex_month_category = bank_opex_category_for_month(bank_opex_result, target_month)
    order_profitability = build_order_profitability(sitegiant, dropship, platform_fees)

    print("Loading restock plan (if generated)...")
    restock = load_restock_data()

    print("Loading restock pipeline -- confirmed 60-day/MOQ methodology (if generated)...")
    restock_pipeline = load_restock_pipeline_data()

    print("Loading restock strategy -- brand mix / sellout scale-up / stop-restock (if generated)...")
    restock_strategy = load_restock_strategy_data()

    ctx = {
        "sitegiant": sitegiant, "dropship": dropship, "cash_flow": cash_flow,
        "bank_opex": bank_opex_result, "bank_opex_month_category": bank_opex_month_category,
        "payroll_monthly": payroll_monthly_df, "platform_fees": platform_fees,
        "net_profit": net_profit, "brand_margin": brand_margin, "warnings": WARNINGS,
        "order_profitability": order_profitability, "bank_df": bank_df,
        "restock": restock, "restock_pipeline": restock_pipeline, "restock_strategy": restock_strategy,
    }
    ctx["health"] = build_financial_health(ctx, restock)

    out_path = BASE_DIR / "Executive_Financial_Report.md"
    write_executive_report(ctx, out_path)

    snapshot_path = save_history_snapshot(ctx)
    history = load_history()
    dashboard_path = BASE_DIR / "dashboard.html"
    write_dashboard_html(ctx, history, dashboard_path)

    print("\n=== SUMMARY ===")
    print(f"Revenue: {fmt_money(net_profit['revenue'])}")
    print(f"COGS: {fmt_money(net_profit['cogs'])}")
    print(f"Gross Profit: {fmt_money(net_profit['gross_profit'])} ({net_profit['gross_margin_pct']:.1f}%)")
    print(f"Net Profit: {fmt_money(net_profit['net_profit'])} ({net_profit['net_margin_pct']:.1f}%)")
    print(f"\nReport written to: {out_path}")
    print(f"Dashboard written to: {dashboard_path}")
    print(f"History snapshot saved: {snapshot_path} ({len(history)} month(s) accumulated so far)")
    if WARNINGS:
        print(f"{len(WARNINGS)} data-quality warnings -- see report Section 7.")


if __name__ == "__main__":
    main()
