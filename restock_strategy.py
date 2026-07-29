"""
Restock Strategy -- a decision layer on top of restock_planner.py (ABC/OTB)
and restock_pipeline.py (60-day/MOQ), per Bun's direction 2026-07-21.
Neither of those two systems is touched or replaced -- this reads the same
real underlying data (SiteGiant forecasts, the active ISKU catalog, the
finance pipeline's own per-order profitability) and answers four questions
Bun described as slow/manual today:

  1. Brand Mix -- what % of real revenue is FRES- (Frescoone's own brand)
     vs. generic right now, against the stated 85%/15% goal? This is a
     tracked KPI, not an enforced constraint -- Bun was explicit that the
     goal isn't achievable immediately given current traffic/sales, so the
     job here is to report the real number every time, not pretend it's 85%.

  2. Restock Status (per SKU) -- one priority-ordered status that replaces
     what used to be two disconnected columns (a fixed-bucket "urgency" and
     a separate FRES "scale tier"). Fixed 2026-07-21 after Bun pointed out
     two real problems with the first version:

     (a) The original urgency buckets (Critical/Urgent/Moderate/Healthy at
         3/7/14 days) were arbitrary and disconnected from reality -- Bun's
         actual restock lead time is 45-60 days for EVERY SKU, so a SKU
         showing "Healthy" at 25 days left was actually in real trouble
         (it'll run dry ~25 days before a freshly-placed order would even
         arrive). Status is now anchored to LEAD_TIME_DAYS (standardized at
         50, the middle of Bun's 45-60 day range) instead of fixed buckets:
         if a SKU's stock won't last until a new order would land, that's a
         real gap risk regardless of whether it's 5 days away or 45.

     (b) Bun wanted one combined answer per SKU -- "which ISKU are running
         well and need restock more" alongside "which will actually run
         out" -- rather than two separate signals to mentally combine.
         compute_restock_status() now returns exactly one of: Out of Stock,
         Stockout Risk -- Order Now, Stockout Risk -- Reorder Placed, Scale
         Up -- Running Well (FRES top earner or trending-up, no gap risk
         yet -- a proactive "grow this before it's urgent" signal, not a
         reactive one), Stop-Restock Candidate, or Covered.

     Multiplier logic (how much extra to order, FRES only -- generic always
     stays at baseline/1x regardless of status): 2x on any real gap risk,
     3x if that SKU is also a top-20%-NPE earner among FRES SKUs, 1.5x for
     "Scale Up -- Running Well" (smaller than a reactive scale-up since
     there's no emergency, just an opportunity), capped at 1.5x for any SKU
     costing over RM50 regardless of tier, per Bun's cost-caution rule.

     IMPORTANT CAVEAT Bun should know about: this still multiplies the
     baseline DEMAND-DRIVEN need (same formula restock_pipeline.py uses),
     not a "previous restock batch quantity" -- actual placed-order history
     isn't tracked anywhere yet (confirmed 2026-07-21). Once that exists,
     it should be added as a second input alongside this, not a replacement.

  3. Stop-Restock Candidates -- SKUs with negligible sales across the last
     4 calendar months (Mar-Jun 2026, the monthly forecast exports already
     in inventory-forecast-SiteGiant/), with a plain-language reason, not
     just a number.

  4. Scale-Potential Categorization -- every active SKU color-coded by its
     4-month sales trend (Trending Up / Stable / Declining / Stop-Restock
     Candidate). "Trending Up" is a real, verifiable signal (sales rose
     every single month) -- not a statistical forecast. Nothing here claims
     to predict a specific future sales number; it flags real momentum so
     Bun doesn't have to eyeball 4 months of data per SKU by hand.

Output: restock_strategy_data.json (optional/additive input to
finance_analyzer.py's dashboard, same pattern as restock_data.json and
restock_pipeline_data.json).
"""
import json
import math
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import finance_analyzer as fa
import restock_planner as rp
import restock_pipeline as pipeline  # reuses confirmed Excel styling + the SiteGiant bulk-upload PO writer

BASE_DIR = fa.BASE_DIR

TARGET_FRES_PCT = 85.0

# Restock lead time -- Bun confirmed 2026-07-21 this applies to EVERY SKU,
# not just fast movers: placing an order today means it lands in 45-60 days.
# Standardized on 50 as the middle of that range. This single number now
# drives two things that used to be inconsistent with each other: (a) the
# baseline restock quantity target (how many days of cover to buy up to),
# and (b) the restock STATUS below (will this SKU survive until a
# freshly-placed order would arrive?). Before this fix, urgency used
# arbitrary 3/7/14-day buckets that had nothing to do with the real 45-60
# day replenishment cycle -- a SKU with 25 days of stock left showed
# "Healthy" even though it will actually run dry ~25 days before a new
# order lands. Status is now anchored to the real cycle instead.
LEAD_TIME_DAYS = 50
SALES_WINDOW_DAYS = 30
ROUND_TO = 5
MOQ_LOW_COST_QTY = 5
MOQ_HIGH_COST_QTY = 3
COST_DAMPENER_THRESHOLD = 50.0
COST_DAMPENER_CAP_MULTIPLIER = 1.5
STOP_RESTOCK_MAX_4MO_SALES = 3   # total units across 4 months at/below this = stop-restock candidate
RUNNING_WELL_SCALE_MULTIPLIER = 1.5  # proactive bump for FRES top earners/trending-up SKUs, even without gap risk

STATUS_COLOR = {
    "Out of Stock": "red", "Stockout Risk -- Order Now": "red",
    "Stockout Risk -- Reorder Placed": "orange", "Scale Up -- Running Well": "blue",
    "Stop-Restock Candidate": "grey", "Covered": "green", "No Sales Signal": "grey",
}


def compute_eta_days(stock_on_hand, total_sales):
    """Days until stock runs out at the current 30-day sale rate. None when
    there's no sales velocity to estimate one from -- never a fabricated
    infinite/zero number standing in for "unknown"."""
    if stock_on_hand <= 0:
        return 0.0
    if total_sales <= 0:
        return None
    sale_per_day = total_sales / SALES_WINDOW_DAYS
    return stock_on_hand / sale_per_day


def compute_restock_status(stock_on_hand, eta_days, stock_on_purchase_order, brand_type,
                            is_top_earner, trend_category):
    """One priority-ordered status per SKU -- replaces the old separate
    Urgency + Tier columns Bun had to mentally combine. Checked in order:
    actual/imminent stockout gap always wins (most actionable), then FRES
    growth opportunity, then the 4-month stop-restock signal, else healthy.
    """
    if stock_on_hand <= 0:
        return "Out of Stock"
    if eta_days is not None and eta_days < LEAD_TIME_DAYS:
        if stock_on_purchase_order <= 0:
            return "Stockout Risk -- Order Now"
        return "Stockout Risk -- Reorder Placed"
    if brand_type == "FRES" and (is_top_earner or trend_category == "Trending Up"):
        return "Scale Up -- Running Well"
    if trend_category == "Stop-Restock Candidate":
        return "Stop-Restock Candidate"
    if eta_days is None:
        return "No Sales Signal"
    return "Covered"


def classify_brand(sku):
    """Same rule as finance_analyzer.brand_margin_analysis() -- kept
    identical on purpose so "FRES %" here always matches the dashboard's
    own Brand Performance numbers."""
    return "FRES" if str(sku).upper().startswith("FRES") else "Generic"


# ---------------------------------------------------------------------------
# 1. Brand Mix -- real revenue split vs. the 85/15 goal
# ---------------------------------------------------------------------------
def compute_brand_mix(sitegiant, dropship):
    bm = fa.brand_margin_analysis(sitegiant, dropship)
    total_revenue = float(bm["revenue"].sum())
    fres_row = bm[bm["brand"] == "Frescoone Proprietary (FRES)"]
    fres_revenue = float(fres_row["revenue"].sum()) if not fres_row.empty else 0.0
    fres_pct = fres_revenue / total_revenue * 100 if total_revenue else 0.0
    return {
        "fres_revenue": fres_revenue, "generic_revenue": total_revenue - fres_revenue,
        "total_revenue": total_revenue, "fres_pct": fres_pct, "generic_pct": 100 - fres_pct,
        "target_fres_pct": TARGET_FRES_PCT, "gap_to_target_pct": TARGET_FRES_PCT - fres_pct,
    }


# ---------------------------------------------------------------------------
# Per-SKU NPE -- reuses build_order_profitability()'s already-verified
# per-line net_profit rather than recomputing revenue/cost/fee logic again.
# AVERAGE net profit per order, not a total. A summed total scales with how
# many orders happened to land in the month -- a SKU with 10 small orders can
# out-rank a SKU with 2 large, more-profitable ones, and the total is easy to
# misread as tied to whatever window a SKU's "Sales 30D" figure covers (it
# isn't -- this draws from every SiteGiant D2C + dropship reseller order with
# a known cost, same as before, just averaged instead of summed). Confirmed
# with Bun 2026-07-22.
# ---------------------------------------------------------------------------
def compute_sku_npe(order_profitability):
    if order_profitability is None or order_profitability.empty:
        return pd.DataFrame(columns=["sku", "npe", "npe_revenue", "npe_order_count"])
    costed = order_profitability[order_profitability["cost_known"]]
    return (costed.groupby("sku").agg(npe=("net_profit", "mean"), npe_revenue=("revenue", "sum"),
                                        npe_order_count=("net_profit", "count"))
            .reset_index())


# ---------------------------------------------------------------------------
# 2. Unified restock status + quantity -- one status per SKU (see
# compute_restock_status docstring), combined with the multiplier logic
# that decides how much to actually order. Generic-brand SKUs get the same
# real status (they can absolutely show "Stockout Risk", nobody wants ANY
# SKU to gap) but never scale beyond baseline -- confirmed 2026-07-21 that
# scale-up is FRES-only.
# ---------------------------------------------------------------------------
def compute_restock_recommendations(active_df, last30, sku_master, sku_npe, trend_lookup):
    df = active_df.merge(last30, left_on="sku", right_on="isku", how="left", indicator=True)
    not_in_forecast = df[df["_merge"] == "left_only"][["sku", "order_sheet", "brand"]].copy()
    df = df[df["_merge"] == "both"].drop(columns=["_merge", "isku"]).copy()
    for c in ("total_sales", "stock_on_hand", "stock_on_purchase_order"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df = df.merge(sku_master[["cost"]], left_on="sku", right_index=True, how="left")
    no_cost = df[df["cost"].isna()][["sku", "order_sheet", "stock_on_hand", "total_sales"]].copy()
    df = df[df["cost"].notna()].copy()

    df["brand_type"] = df["sku"].apply(classify_brand)
    df = df.merge(sku_npe, on="sku", how="left")
    df["npe"] = df["npe"].fillna(0.0)
    df["npe_order_count"] = df["npe_order_count"].fillna(0).astype(int)
    df["trend_category"] = df["sku"].map(trend_lookup).fillna("Stable")

    # "Top earner" = top 20% of FRES SKUs by NPE, among those actually
    # profitable this month -- an item that's losing money isn't a "top
    # earner" just because it also sold out.
    fres_positive_npe = df[(df["brand_type"] == "FRES") & (df["npe"] > 0)]["npe"]
    top_earner_cutoff = float(fres_positive_npe.quantile(0.80)) if len(fres_positive_npe) >= 5 else None

    def compute_row(row):
        has_sales = row["total_sales"] > 0
        baseline_need = 0.0
        if has_sales:
            baseline_need = math.ceil((row["total_sales"] / SALES_WINDOW_DAYS) * LEAD_TIME_DAYS) - row["stock_on_hand"]
        baseline_need = max(0.0, baseline_need)

        eta_days = compute_eta_days(row["stock_on_hand"], row["total_sales"])
        is_top_earner = bool(row["brand_type"] == "FRES" and top_earner_cutoff is not None
                              and row["npe"] >= top_earner_cutoff)
        status = compute_restock_status(row["stock_on_hand"], eta_days, row["stock_on_purchase_order"],
                                          row["brand_type"], is_top_earner, row["trend_category"])
        status_color = STATUS_COLOR.get(status, "grey")

        gap_risk = status in ("Out of Stock", "Stockout Risk -- Order Now")
        if row["brand_type"] != "FRES":
            multiplier, reason = 1.0, "Generic brand -- safety restock only, no scale-up"
        elif status == "Stockout Risk -- Reorder Placed":
            multiplier, reason = 1.0, "Gap risk, but stock is already on order -- verify quantity/timing before adding more"
        elif gap_risk and is_top_earner:
            multiplier, reason = 3.0, f"FRES top earner, {status.lower()} -- scale up aggressively"
        elif gap_risk:
            multiplier, reason = 2.0, f"FRES item, {status.lower()} -- scale up"
        elif status == "Scale Up -- Running Well":
            multiplier, reason = RUNNING_WELL_SCALE_MULTIPLIER, "FRES SKU running well (top earner or trending up), no gap risk yet -- proactively order more before it becomes urgent"
        else:
            multiplier, reason = 1.0, f"FRES, status={status.lower()} -- normal restock"

        cost_capped = False
        if row["cost"] > COST_DAMPENER_THRESHOLD and multiplier > COST_DAMPENER_CAP_MULTIPLIER:
            multiplier = COST_DAMPENER_CAP_MULTIPLIER
            cost_capped = True
            reason += f" (capped at {COST_DAMPENER_CAP_MULTIPLIER:.1f}x -- unit cost over RM{COST_DAMPENER_THRESHOLD:.0f})"

        moq = MOQ_LOW_COST_QTY if row["cost"] < COST_DAMPENER_THRESHOLD else MOQ_HIGH_COST_QTY
        base_fields = {"status": status, "status_color": status_color, "eta_days": eta_days,
                        "is_top_earner": is_top_earner}

        # Stop-restock overrides everything else on quantity, even if the
        # baseline formula would technically suggest a small top-up -- the
        # whole point of that status is "there's no case for restocking this."
        if status == "Stop-Restock Candidate":
            return pd.Series({**base_fields, "baseline_need": 0.0, "multiplier": 1.0,
                               "reason": "4-month sales trend says stop -- see Stop-Restock Candidates",
                               "cost_capped": False, "restock_qty": 0})

        raw_qty = baseline_need * multiplier
        if raw_qty <= 0:
            if row["stock_on_hand"] < 3 and not has_sales:
                return pd.Series({**base_fields, "baseline_need": 0.0, "multiplier": 1.0,
                                   "reason": "No recent sales, low stock -- safety restock", "cost_capped": False,
                                   "restock_qty": moq})
            return pd.Series({**base_fields, "baseline_need": baseline_need, "multiplier": multiplier,
                               "reason": reason, "cost_capped": cost_capped, "restock_qty": 0})

        rounded = math.ceil(raw_qty / ROUND_TO) * ROUND_TO
        qty = max(rounded, moq)
        return pd.Series({**base_fields, "baseline_need": baseline_need, "multiplier": multiplier,
                           "reason": reason, "cost_capped": cost_capped, "restock_qty": qty})

    computed = df.apply(compute_row, axis=1)
    df = pd.concat([df, computed], axis=1)
    df["est_cost"] = df["restock_qty"] * df["cost"]

    return df, not_in_forecast, no_cost


# ---------------------------------------------------------------------------
# 3 & 4. Monthly trend -- Stop-Restock Candidates + Scale-Potential tiers.
# Built from the 4 calendar-month SiteGiant exports already saved in
# inventory-forecast-SiteGiant/ (2026-03 through 2026-06) -- no new input
# needed. These are month-SCOPED exports (that month's real sales), not
# rolling windows -- confirmed when they were first pulled.
# ---------------------------------------------------------------------------
def build_monthly_trend(active_df, monthly_forecasts):
    months = sorted(monthly_forecasts.keys())
    trend = active_df[["sku", "order_sheet", "brand"]].copy()
    trend["brand_type"] = trend["sku"].apply(classify_brand)
    for m in months:
        mf = monthly_forecasts[m][["isku", "total_sales"]].rename(columns={"isku": "sku", "total_sales": m})
        mf[m] = pd.to_numeric(mf[m], errors="coerce").fillna(0)
        mf = mf.drop_duplicates("sku", keep="last")
        trend = trend.merge(mf, on="sku", how="left")
    for m in months:
        trend[m] = trend[m].fillna(0)
    trend["total_4mo"] = trend[months].sum(axis=1) if months else 0
    return trend, months


def classify_trend(trend_df, months):
    if not months:
        trend_df["category"], trend_df["color"], trend_df["trend_reason"] = "Insufficient Data", "grey", "No monthly forecast files found"
        return trend_df

    def categorize(row):
        vals = [row[m] for m in months]
        total = sum(vals)
        path = " -> ".join(str(int(v)) for v in vals)
        if total <= STOP_RESTOCK_MAX_4MO_SALES:
            return pd.Series({"category": "Stop-Restock Candidate", "color": "red",
                               "trend_reason": f"Only {int(total)} unit(s) sold across {months[0]} to {months[-1]} "
                                               f"({path}) -- minimal ongoing demand, no case for continued restock"})
        rising = all(vals[i] >= vals[i - 1] for i in range(1, len(vals))) and vals[-1] > vals[0]
        declining = all(vals[i] <= vals[i - 1] for i in range(1, len(vals))) and vals[-1] < vals[0]
        if rising:
            return pd.Series({"category": "Trending Up", "color": "green",
                               "trend_reason": f"Sales rose every month ({path}) -- real momentum, worth scaling up further"})
        if declining:
            return pd.Series({"category": "Declining", "color": "amber",
                               "trend_reason": f"Sales fell every month ({path}) -- watch closely, may become a stop-restock candidate"})
        return pd.Series({"category": "Stable", "color": "grey", "trend_reason": f"Sales stable/mixed ({path})"})

    result = trend_df.apply(categorize, axis=1)
    return pd.concat([trend_df, result], axis=1)


# ---------------------------------------------------------------------------
# JSON output -- same additive/optional pattern as restock_data.json and
# restock_pipeline_data.json.
# ---------------------------------------------------------------------------
def _clean(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (bool,)):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return str(v)


def build_strategy_json(brand_mix, scaleup_df, trend_df, months, not_in_forecast, no_cost):
    scaleup_cols = ["sku", "order_sheet", "brand_type", "stock_on_hand", "stock_on_purchase_order",
                     "total_sales", "npe", "npe_order_count", "status", "status_color", "eta_days",
                     "is_top_earner", "multiplier", "cost_capped", "cost", "restock_qty",
                     "est_cost", "reason"]
    scaleup_detail = [{c: _clean(r[c]) for c in scaleup_cols} for _, r in scaleup_df.iterrows()]

    trend_cols = ["sku", "order_sheet", "brand_type", "category", "color", "trend_reason", "total_4mo"] + months
    trend_detail = [{c: _clean(r[c]) for c in trend_cols} for _, r in trend_df.iterrows()]

    stop_restock = [r for r in trend_detail if r["category"] == "Stop-Restock Candidate"]
    trending_up = sorted([r for r in trend_detail if r["category"] == "Trending Up"],
                          key=lambda r: r["total_4mo"], reverse=True)

    fres_scaled = [r for r in scaleup_detail if r["brand_type"] == "FRES" and r["multiplier"] > 1.0]
    gap_risk = [r for r in scaleup_detail if r["status"] in ("Out of Stock", "Stockout Risk -- Order Now",
                                                                "Stockout Risk -- Reorder Placed")]
    running_well = [r for r in scaleup_detail if r["status"] == "Scale Up -- Running Well"]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "months_analyzed": months,
        "lead_time_days": LEAD_TIME_DAYS,
        "brand_mix": {k: _clean(v) for k, v in brand_mix.items()},
        "summary": {
            "active_skus_matched": int(len(scaleup_df)),
            "gap_risk_count": len(gap_risk),
            "running_well_count": len(running_well),
            "fres_scaleup_count": len(fres_scaled),
            "fres_scaleup_spend": round(sum(r["est_cost"] or 0 for r in fres_scaled), 2),
            "stop_restock_count": len(stop_restock),
            "trending_up_count": len(trending_up),
            "missing_from_forecast": int(len(not_in_forecast)),
            "no_cost_count": int(len(no_cost)),
        },
        "scaleup_detail": scaleup_detail,
        "trend_detail": trend_detail,
        "stop_restock_candidates": stop_restock,
        "trending_up": trending_up,
        "not_in_forecast": [{"sku": r["sku"], "order_sheet": r["order_sheet"], "brand": r.get("brand")}
                             for _, r in not_in_forecast.iterrows()],
        "no_cost": [{"sku": r["sku"], "order_sheet": r["order_sheet"], "stock_on_hand": _clean(r["stock_on_hand"]),
                      "total_sales": _clean(r["total_sales"])} for _, r in no_cost.iterrows()],
    }


# ---------------------------------------------------------------------------
# 4. Purchase-order artifacts for the top-earner SKUs specifically -- run on
# request (Bun 2026-07-22: "perform restock action for top earner"), scoped
# narrowly to is_top_earner == True (the top-20%-avg-NPE-among-FRES signal,
# the star in the dashboard), not the broader "Scale Up -- Running Well"
# status (which also includes trending-up SKUs that aren't top earners).
# Reuses restock_pipeline.py's confirmed SiteGiant bulk-upload PO format
# as-is (write_supplier_po_files doesn't care which methodology produced the
# restock_qty), but writes to a SEPARATE directory/filename so this doesn't
# collide with or silently overwrite restock_pipeline's own confirmed
# 60-day/MOQ supplier PO files for the same month. The purchase-list workbook
# here is purpose-built (not reused) because it keeps the Strategy-specific
# "why" -- status, avg NPE, multiplier, reason -- which is the entire point
# of this list and isn't part of restock_pipeline's generic schema.
# ---------------------------------------------------------------------------
TOP_EARNER_PO_DIR_NAME = "restock-po-files-top-earner"


def build_top_earner_po_list(scaleup_df):
    df = scaleup_df[scaleup_df["is_top_earner"] & (scaleup_df["restock_qty"] > 0)].copy()
    df["action"] = "RESTOCK"
    df["brand"] = df["brand_type"]
    return df.sort_values(["order_sheet", "sku"])


def write_top_earner_workbook(po_df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TopEarner_PO_List"

    r = 1
    ws.cell(row=r, column=1, value="Top-Earner Restock -- Spend Summary by Supplier").font = pipeline.SECTION_FONT
    r += 2
    r = pipeline._write_header_row(ws, r, ["Order Sheet", "SKUs", "Total Units", "Est. Spend (RM)"])
    summary = (po_df.groupby("order_sheet").agg(skus=("sku", "count"), units=("restock_qty", "sum"),
                                                   spend=("est_cost", "sum"))
               .reset_index().sort_values("spend", ascending=False))
    for _, row in summary.iterrows():
        ws.cell(row=r, column=1, value=row["order_sheet"])
        ws.cell(row=r, column=2, value=int(row["skus"]))
        ws.cell(row=r, column=3, value=int(row["units"]))
        ws.cell(row=r, column=4, value=round(float(row["spend"]), 2))
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=int(summary["skus"].sum()))
    ws.cell(row=r, column=3, value=int(summary["units"].sum()))
    ws.cell(row=r, column=4, value=round(float(summary["spend"].sum()), 2))
    r += 3

    headers = ["SKU", "Status", "Stock On Hand", "On Purchase Order (not netted)", "Sales (30d)",
               "Avg NPE (RM)", "Multiplier", "Restock Qty", "Unit Cost (RM)", "Est. Cost (RM)", "Reason"]
    for order_sheet in sorted(po_df["order_sheet"].unique()):
        section = po_df[po_df["order_sheet"] == order_sheet]
        ws.cell(row=r, column=1, value=order_sheet).font = pipeline.SECTION_FONT
        r += 1
        r = pipeline._write_header_row(ws, r, headers)
        for _, row in section.iterrows():
            vals = [row["sku"], row["status"], row["stock_on_hand"], row["stock_on_purchase_order"],
                    row["total_sales"], round(float(row["npe"]), 2), f"{row['multiplier']}x",
                    int(row["restock_qty"]), round(float(row["cost"]), 2), round(float(row["est_cost"]), 2),
                    row["reason"]]
            for col_idx, v in enumerate(vals, start=1):
                ws.cell(row=r, column=col_idx, value=v)
            r += 1
        r += 2
    pipeline._autosize(ws, len(headers))

    wb.save(out_path)
    return out_path


def perform_top_earner_restock_action(scaleup_df, last30):
    po_df = build_top_earner_po_list(scaleup_df)
    if po_df.empty:
        fa.warn("restock_strategy: no top-earner SKU currently has a positive restock_qty -- "
                 "nothing to write for the top-earner restock action.")
        return None, []

    workbook_path = BASE_DIR / f"TopEarner_Restock_{pipeline.MONTH_LABEL}.xlsx"
    write_top_earner_workbook(po_df, workbook_path)

    po_dir = BASE_DIR / TOP_EARNER_PO_DIR_NAME
    po_files = pipeline.write_supplier_po_files(po_df, po_dir)

    return workbook_path, po_files


def main():
    print("Loading active ISKU catalog...")
    active_df = rp.load_active_isku_catalog()

    print("Loading SiteGiant Inventory Forecasting exports (last-30-days + 4 monthly)...")
    last30, alltime, monthly = rp.load_all_forecast_files()
    if len(monthly) < 2:
        fa.warn(f"restock_strategy: only {len(monthly)} monthly forecast file(s) found -- trend "
                 f"classification will be unreliable with fewer than a few months of data.")

    print("Loading finance context (cost master, revenue, per-order profitability)...")
    sku_master = fa.apply_cost_overrides(fa.load_sku_master())
    orders = fa.load_sitegiant_orders()
    sitegiant = fa.sitegiant_channel_metrics(orders, sku_master)
    dropship = fa.dropship_channel_metrics(sku_master)
    target_month = (sitegiant["monthly"]["month"].iloc[0] if not sitegiant["monthly"].empty
                     else pd.Period(fa.datetime.now(), freq="M"))
    platform_fees = fa.load_platform_fees(orders, target_month)
    order_profitability = fa.build_order_profitability(sitegiant, dropship, platform_fees)

    print("Computing brand mix (FRES vs. Generic, real revenue)...")
    brand_mix = compute_brand_mix(sitegiant, dropship)

    print("Computing per-SKU NPE...")
    sku_npe = compute_sku_npe(order_profitability)

    print("Building 4-month sales trend (stop-restock + scale-potential)...")
    trend_df, months = build_monthly_trend(active_df, monthly)
    trend_df = classify_trend(trend_df, months)
    trend_lookup = dict(zip(trend_df["sku"], trend_df["category"]))

    print(f"Computing unified restock status (lead time = {LEAD_TIME_DAYS} days)...")
    scaleup_df, not_in_forecast, no_cost = compute_restock_recommendations(
        active_df, last30, sku_master, sku_npe, trend_lookup)

    print("Writing restock_strategy_data.json...")
    data = build_strategy_json(brand_mix, scaleup_df, trend_df, months, not_in_forecast, no_cost)
    json_path = BASE_DIR / "restock_strategy_data.json"
    json_path.write_text(json.dumps(data, indent=2), encoding="utf-8")

    print("Performing restock action for top-earner SKUs (PO files + workbook)...")
    top_earner_workbook_path, top_earner_po_files = perform_top_earner_restock_action(scaleup_df, last30)

    s = data["summary"]
    bm = data["brand_mix"]
    print(f"\n=== SUMMARY ===")
    print(f"Brand mix: {bm['fres_pct']:.1f}% FRES / {bm['generic_pct']:.1f}% Generic "
          f"(target {bm['target_fres_pct']:.0f}% FRES, gap {bm['gap_to_target_pct']:+.1f}pp)")
    print(f"SKUs at stockout gap risk (won't survive {LEAD_TIME_DAYS}-day lead time): {s['gap_risk_count']}")
    print(f"FRES SKUs running well, worth scaling proactively: {s['running_well_count']}")
    print(f"FRES SKUs flagged for scale-up (any tier): {s['fres_scaleup_count']} / {fa.fmt_money(s['fres_scaleup_spend'])}")
    print(f"Stop-restock candidates: {s['stop_restock_count']}")
    print(f"Trending up: {s['trending_up_count']}")
    print(f"\nDashboard data written to: {json_path} -- re-run finance_analyzer.py to fold it into dashboard.html")
    if top_earner_workbook_path:
        print(f"\nTop-earner restock action: {top_earner_workbook_path.name} "
              f"+ {len(top_earner_po_files)} per-supplier PO file(s) in {TOP_EARNER_PO_DIR_NAME}/")
    if fa.WARNINGS:
        print(f"\n{len(fa.WARNINGS)} data-quality warnings logged during load.")


if __name__ == "__main__":
    main()
