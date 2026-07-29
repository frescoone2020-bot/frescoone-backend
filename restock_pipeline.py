"""
Monthly Restock Pipeline -- confirmed 60-day lead-time / MOQ methodology
==========================================================================
This is Frescoone's SECOND, separate restock system, kept deliberately
independent from restock_planner.py (the ABC-tiered / Open-To-Buy version)
per Bun's direction 2026-07-21: "keep both, side by side." This one
replicates the exact confirmed methodology already in use in the
"Frescoone Sales & Stock Analyst" project (documented in
Restock_Workflow.html, uploaded by Bun), so the two systems don't produce
silently conflicting numbers under the same name.

Reuses restock_planner.py's loaders (active ISKU catalog parser, SiteGiant
forecast file parser) and finance_analyzer.py's SKU cost master -- never a
second, drifting copy of the same source-of-truth parsing logic.

Confirmed rules (from Restock_Workflow.html):
  1. Match every isku in the last-30-day forecast against the active ISKU
     catalog (14 supplier order sheets). Unmatched active SKUs logged
     separately, not silently dropped.
  2. SKUs with sales:  need = ROUNDUP((total_sales / 30) * 60) - stock_on_hand
  3. Zero-sales SKUs:  restock (at MOQ) only if stock_on_hand < 3 units
  4. Stock on Purchase Order is NEVER netted off the need -- flagged for
     manual review instead (Bun reviews these by eye; automating a netting
     assumption here would silently override that check).
  5. Final quantity rounded UP to the nearest multiple of 5.
  6. MOQ floor: unit cost < 50 -> min 5 units; unit cost >= 50 -> min 3
     units. Source doc says "RMB50" (Chinese Yuan, most of these suppliers
     are China-based) -- the only cost figure available in this project is
     RM (Malaysian Ringgit, from isku_database.json), so RM cost is used
     here instead. Confirmed 2026-07-21 with Bun that this floor rarely
     changes anything in practice (round-to-5 already guarantees >=5 for
     any SKU with a positive need), so the RMB-vs-RM ambiguity has limited
     practical impact -- flagged in the report regardless.

Outputs:
  Restock_Purchase_List_<MONTH>.xlsx  -- 3-tab workbook (PO list, forecast
                                          audit trail, no-action list)
  PO_<Supplier>_<MONTH>.xlsx           -- one per supplier needing restock,
                                          SiteGiant bulk-upload format
  Restock_Dashboard_<MONTH>.html       -- standalone interactive dashboard
"""
import json
import math
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import finance_analyzer as fa
import restock_planner as rp

BASE_DIR = fa.BASE_DIR
MONTH_LABEL = datetime.now().strftime("%Y-%m")

LEAD_TIME_DAYS = 60
SALES_WINDOW_DAYS = 30
ZERO_SALES_STOCK_THRESHOLD = 3
ROUND_TO = 5
MOQ_LOW_COST_THRESHOLD = 50.0   # see module docstring -- RM, not confirmed RMB
MOQ_LOW_COST_QTY = 5
MOQ_HIGH_COST_QTY = 3


# ---------------------------------------------------------------------------
# 1-2. Match + compute restock need
# ---------------------------------------------------------------------------
def compute_restock_list(active_df, last30, sku_master):
    df = active_df.merge(last30, left_on="sku", right_on="isku", how="left", indicator=True)
    not_in_forecast = df[df["_merge"] == "left_only"][["sku", "order_sheet", "brand"]].copy()
    df = df[df["_merge"] == "both"].drop(columns=["_merge", "isku"]).copy()
    for c in ("total_sales", "stock_on_hand", "stock_on_purchase_order"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df = df.merge(sku_master[["cost"]], left_on="sku", right_index=True, how="left")
    no_cost = df[df["cost"].isna()][["sku", "order_sheet", "stock_on_hand", "total_sales"]].copy()
    df = df[df["cost"].notna()].copy()

    def moq_for(cost):
        return MOQ_LOW_COST_QTY if cost < MOQ_LOW_COST_THRESHOLD else MOQ_HIGH_COST_QTY

    def compute_row(row):
        has_sales = row["total_sales"] > 0
        moq = moq_for(row["cost"])
        if has_sales:
            raw_need = math.ceil((row["total_sales"] / SALES_WINDOW_DAYS) * LEAD_TIME_DAYS) - row["stock_on_hand"]
            if raw_need <= 0:
                return pd.Series({"raw_need": raw_need, "restock_qty": 0, "moq": moq,
                                   "reason": f"{row['total_sales']:.0f} units sold/30d covers {LEAD_TIME_DAYS}d "
                                             f"lead time already ({row['stock_on_hand']:.0f} on hand)"})
            rounded = math.ceil(raw_need / ROUND_TO) * ROUND_TO
            qty = max(rounded, moq)
            return pd.Series({"raw_need": raw_need, "restock_qty": qty, "moq": moq, "reason": None})
        else:
            if row["stock_on_hand"] < ZERO_SALES_STOCK_THRESHOLD:
                return pd.Series({"raw_need": moq, "restock_qty": moq, "moq": moq, "reason": None})
            return pd.Series({"raw_need": 0, "restock_qty": 0, "moq": moq,
                               "reason": f"No sales in last {SALES_WINDOW_DAYS}d, "
                                         f"{row['stock_on_hand']:.0f} units on hand (>= {ZERO_SALES_STOCK_THRESHOLD})"})

    computed = df.apply(compute_row, axis=1)
    df = pd.concat([df, computed], axis=1)
    df["est_cost"] = df["restock_qty"] * df["cost"]
    df["on_po_flag"] = df["stock_on_purchase_order"] > 0
    df["action"] = df["restock_qty"].apply(lambda q: "RESTOCK" if q > 0 else "NO_ACTION")

    return df, not_in_forecast, no_cost


# ---------------------------------------------------------------------------
# 3. Restock_Purchase_List_<MONTH>.xlsx -- 3 tabs
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12, color="305496")
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_header_row(ws, row_idx, headers):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    return row_idx + 1


def _autosize(ws, n_cols, min_width=10, max_width=45):
    for i in range(1, n_cols + 1):
        col = get_column_letter(i)
        length = max((len(str(c.value)) for c in ws[col] if c.value is not None), default=min_width)
        ws.column_dimensions[col].width = min(max(length + 2, min_width), max_width)


def write_purchase_list_workbook(restock_df, not_in_forecast, no_cost, last30, out_path):
    wb = Workbook()

    # --- Tab 1: Restock_PO_List ---
    ws = wb.active
    ws.title = "Restock_PO_List"
    to_order = restock_df[restock_df["action"] == "RESTOCK"].sort_values(["order_sheet", "sku"])

    r = 1
    ws.cell(row=r, column=1, value="Spend Summary by Supplier").font = SECTION_FONT
    r += 2
    r = _write_header_row(ws, r, ["Order Sheet", "SKUs to Restock", "Total Units", "Est. Spend (RM)"])
    summary = to_order.groupby("order_sheet").agg(skus=("sku", "count"), units=("restock_qty", "sum"),
                                                     spend=("est_cost", "sum")).reset_index().sort_values("spend", ascending=False)
    for _, row in summary.iterrows():
        ws.cell(row=r, column=1, value=row["order_sheet"])
        ws.cell(row=r, column=2, value=int(row["skus"]))
        ws.cell(row=r, column=3, value=int(row["units"]))
        ws.cell(row=r, column=4, value=round(float(row["spend"]), 2))
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=int(summary["skus"].sum()))
    ws.cell(row=r, column=3, value=int(summary["units"].sum()))
    ws.cell(row=r, column=4, value=round(float(summary["spend"].sum()), 2))
    r += 3

    for order_sheet in sorted(to_order["order_sheet"].unique()):
        section = to_order[to_order["order_sheet"] == order_sheet]
        ws.cell(row=r, column=1, value=order_sheet).font = SECTION_FONT
        r += 1
        r = _write_header_row(ws, r, ["SKU", "Stock On Hand", "On Purchase Order (not netted)",
                                        "Sales (30d)", "Restock Qty", "MOQ", "Unit Cost (RM)", "Est. Cost (RM)", "Flag"])
        for _, row in section.iterrows():
            flag = "ON PO -- REVIEW" if row["on_po_flag"] else ""
            vals = [row["sku"], row["stock_on_hand"], row["stock_on_purchase_order"], row["total_sales"],
                    int(row["restock_qty"]), int(row["moq"]), round(float(row["cost"]), 2),
                    round(float(row["est_cost"]), 2), flag]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col_idx, value=v)
                if col_idx == 9 and flag:
                    cell.fill = FLAG_FILL
            r += 1
        r += 2
    _autosize(ws, 9)

    # --- Tab 2: Forecast Upload Proof (raw audit trail) ---
    ws2 = wb.create_sheet("Forecast Upload Proof")
    cols = ["product_name", "isku", "stock_on_hand", "stock_on_purchase_order", "safety_stock",
            "total_sales", "sale_per_day", "lead_time", "recommended_quantity"]
    r = _write_header_row(ws2, 1, cols)
    for _, row in last30.iterrows():
        for col_idx, c in enumerate(cols, start=1):
            ws2.cell(row=r, column=col_idx, value=row.get(c))
        r += 1
    _autosize(ws2, len(cols))

    # --- Tab 3: No_Action_ISKU_List ---
    ws3 = wb.create_sheet("No_Action_ISKU_List")
    no_action = restock_df[restock_df["action"] == "NO_ACTION"].sort_values(["order_sheet", "sku"])
    r = 1
    ws3.cell(row=r, column=1, value=f"Sufficient Stock -- No Restock Needed ({len(no_action)} SKUs)").font = SECTION_FONT
    r += 2
    r = _write_header_row(ws3, r, ["SKU", "Order Sheet", "Stock On Hand", "Sales (30d)", "Reason"])
    for _, row in no_action.iterrows():
        for col_idx, v in enumerate([row["sku"], row["order_sheet"], row["stock_on_hand"],
                                       row["total_sales"], row["reason"]], start=1):
            ws3.cell(row=r, column=col_idx, value=v)
        r += 1
    r += 2
    if len(no_cost):
        ws3.cell(row=r, column=1, value=f"No Cost on File -- Cannot Size ({len(no_cost)} SKUs)").font = SECTION_FONT
        r += 2
        r = _write_header_row(ws3, r, ["SKU", "Order Sheet", "Stock On Hand", "Sales (30d)"])
        for _, row in no_cost.iterrows():
            for col_idx, v in enumerate([row["sku"], row["order_sheet"], row["stock_on_hand"], row["total_sales"]], start=1):
                ws3.cell(row=r, column=col_idx, value=v)
            r += 1
        r += 2
    if len(not_in_forecast):
        ws3.cell(row=r, column=1, value=f"Missing From Forecast Export ({len(not_in_forecast)} active SKUs)").font = SECTION_FONT
        r += 2
        r = _write_header_row(ws3, r, ["SKU", "Order Sheet", "Brand"])
        for _, row in not_in_forecast.iterrows():
            for col_idx, v in enumerate([row["sku"], row["order_sheet"], row["brand"]], start=1):
                ws3.cell(row=r, column=col_idx, value=v)
            r += 1
    _autosize(ws3, 5)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 4. Per-supplier PO upload files -- SiteGiant bulk-upload format
# ---------------------------------------------------------------------------
# The active-isku-master.xlsx source file has two "SR" sections that differ
# ONLY by capitalization ("SR under FRESCOONE" = Ring products, "SR Under
# FRESCOONE" = AG Case products -- confirmed 2026-07-21 by checking each
# section's actual SKUs: FRES-RING-* vs FRES-PC-AG-*, matching the "SR
# (Ring)" / "SR (AG Case)" labels in Restock_Workflow.html). Windows
# filenames are case-INSENSITIVE, so naive slugging made both suppliers'
# PO files collide on disk -- one silently overwrote the other. Fixed by
# using the confirmed disambiguated names for this known case, plus a
# generic case-insensitive collision guard for anything else.
_KNOWN_SLUG_OVERRIDES = {
    "Order Sheet-SR under FRESCOONE": "SR (Ring)",
    "Order Sheet-SR Under FRESCOONE": "SR (AG Case)",
}


def write_supplier_po_files(restock_df, out_dir):
    out_dir.mkdir(exist_ok=True)
    to_order = restock_df[restock_df["action"] == "RESTOCK"]
    written = []
    used_slugs_lower = {}
    for order_sheet in sorted(to_order["order_sheet"].unique()):
        section = to_order[to_order["order_sheet"] == order_sheet].sort_values("sku")
        if order_sheet in _KNOWN_SLUG_OVERRIDES:
            supplier_slug = _KNOWN_SLUG_OVERRIDES[order_sheet]
        else:
            supplier_slug = order_sheet.replace("Order Sheet-", "").replace("Order Sheet- ", "").strip().replace("/", "-")
        key = supplier_slug.lower()
        if key in used_slugs_lower:
            used_slugs_lower[key] += 1
            fa.warn(f"restock_pipeline: order sheet {order_sheet!r} collides on a case-insensitive filename "
                     f"with {used_slugs_lower[key]-1} other order sheet(s) already written -- appending a "
                     f"disambiguator so neither file gets silently overwritten. Check restock-po-files/ for "
                     f"a '{supplier_slug} ({used_slugs_lower[key]})' file and confirm it's the right supplier.")
            supplier_slug = f"{supplier_slug} ({used_slugs_lower[key]})"
        else:
            used_slugs_lower[key] = 1
        wb = Workbook()
        ws = wb.active
        ws.title = "PO"
        _write_header_row(ws, 1, ["product_name", "supplier_sku", "isku", "tax", "cost", "quantity"])
        r = 2
        for _, row in section.iterrows():
            ws.cell(row=r, column=1, value=None)   # product_name -- manual entry
            ws.cell(row=r, column=2, value=None)   # supplier_sku -- manual entry
            ws.cell(row=r, column=3, value=row["sku"])
            ws.cell(row=r, column=4, value=None)   # tax -- manual entry
            ws.cell(row=r, column=5, value=0)      # cost fixed to 0 per confirmed format
            ws.cell(row=r, column=6, value=int(row["restock_qty"]))
            r += 1
        _autosize(ws, 6)
        path = out_dir / f"PO_{supplier_slug}_{MONTH_LABEL}.xlsx"
        wb.save(path)
        written.append(path)
    return written


# ---------------------------------------------------------------------------
# OTB sanity check -- Bun's direction 2026-07-21: keep the cash/revenue-based
# ceiling from restock_planner.py as a check ON TOP of this pipeline's total
# spend, not as something that changes any individual SKU's quantity. This
# pipeline's per-SKU math (60-day cover, MOQ, round-to-5) is fully confirmed
# and untouched -- this only tells you whether the TOTAL fits your real cash
# position and revenue, same real numbers as everywhere else in this project.
# ---------------------------------------------------------------------------
def otb_sanity_check(total_spend, finance_ctx):
    cash_position = finance_ctx["cash_position"]
    revenue = finance_ctx["revenue"]
    monthly_fixed_cost = finance_ctx["payroll_cost"] + finance_ctx["bank_opex"]
    revenue_cap = revenue * rp.REVENUE_CAP_PCT
    cash_cap = max(0.0, (cash_position or 0) - monthly_fixed_cost) if cash_position is not None else None
    ceiling = min(revenue_cap, cash_cap) if cash_cap is not None else revenue_cap
    return {
        "total_spend": total_spend, "revenue_cap": revenue_cap, "cash_cap": cash_cap,
        "ceiling": ceiling, "within_budget": total_spend <= ceiling,
        "revenue_cap_pct": rp.REVENUE_CAP_PCT * 100, "cash_position": cash_position,
        "monthly_fixed_cost": monthly_fixed_cost,
    }


# ---------------------------------------------------------------------------
# 5. Restock_Dashboard_<MONTH>.html
# ---------------------------------------------------------------------------
def _clean(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return str(v)


def build_dashboard_json(restock_df, not_in_forecast, no_cost, otb_check):
    to_order = restock_df[restock_df["action"] == "RESTOCK"]
    no_action = restock_df[restock_df["action"] == "NO_ACTION"]

    by_supplier = (to_order.groupby("order_sheet")
                   .agg(skus=("sku", "count"), units=("restock_qty", "sum"), spend=("est_cost", "sum"))
                   .reset_index().sort_values("spend", ascending=False))
    by_brand = (to_order.groupby("brand").agg(spend=("est_cost", "sum"), skus=("sku", "count"))
                .reset_index().sort_values("spend", ascending=False).head(10))

    detail_cols = ["sku", "order_sheet", "brand", "action", "stock_on_hand", "stock_on_purchase_order",
                    "total_sales", "restock_qty", "cost", "est_cost", "on_po_flag", "reason"]
    detail = [{c: _clean(r[c]) for c in detail_cols} for _, r in restock_df.iterrows()]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"), "month_label": MONTH_LABEL,
        "kpis": {
            "active_skus_matched": int(len(restock_df)),
            "skus_needing_restock": int(len(to_order)),
            "recommended_spend": float(to_order["est_cost"].sum()),
            "units_to_order": int(to_order["restock_qty"].sum()),
            "on_po_flags": int(to_order["on_po_flag"].sum()),
            "low_stock_flags": int((restock_df["stock_on_hand"] <= 0).sum()),
            "no_action_count": int(len(no_action)),
            "missing_from_forecast": int(len(not_in_forecast)),
            "no_cost_count": int(len(no_cost)),
        },
        "by_supplier": [{"order_sheet": r["order_sheet"], "skus": int(r["skus"]), "units": int(r["units"]),
                          "spend": round(float(r["spend"]), 2)} for _, r in by_supplier.iterrows()],
        "by_brand": [{"brand": r["brand"] or "Unknown", "spend": round(float(r["spend"]), 2), "skus": int(r["skus"])}
                     for _, r in by_brand.iterrows()],
        "breakdown": {"restock": int(len(to_order)), "no_action": int(len(no_action)),
                      "missing": int(len(not_in_forecast)), "no_cost": int(len(no_cost))},
        "otb_check": {k: _clean(v) if not isinstance(v, bool) else v for k, v in otb_check.items()},
        "detail": detail,
    }


DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Restock Dashboard -- __MONTH__</title>
<style>
:root{--ink:#1a1d29;--sub:#6b7280;--page:#f4f5f7;--card:#fff;--line:#ececf0;--blue:#305496;--blue-soft:#e8edf7;
--green:#3f7a3f;--green-soft:#e9f3e6;--amber:#8a6d00;--amber-soft:#fff2cc;--red:#a3342a;--red-soft:#fbe9e7;--radius:12px;}
*{box-sizing:border-box;}
body{margin:0;background:var(--page);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;}
.wrap{max-width:1200px;margin:0 auto;padding:32px 24px 60px;}
header .eyebrow{font-size:12px;letter-spacing:1.5px;text-transform:uppercase;color:var(--blue);font-weight:700;}
header h1{font-size:26px;margin:6px 0 4px;font-weight:750;}
header p{margin:0;color:var(--sub);font-size:13px;}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px;margin:24px 0;}
.kpi{background:var(--card);border-radius:var(--radius);padding:16px 18px;box-shadow:0 1px 3px rgba(0,0,0,.06);border:1px solid var(--line);}
.kpi .label{font-size:11.5px;color:var(--sub);text-transform:uppercase;letter-spacing:.4px;font-weight:600;}
.kpi .value{font-size:22px;font-weight:750;margin-top:4px;}
.kpi.warn .value{color:var(--amber);}
.kpi.flag .value{color:var(--red);}
.otb-banner{border-radius:var(--radius);padding:16px 20px;margin-bottom:24px;font-size:13.5px;line-height:1.6;}
.otb-banner.ok{background:var(--green-soft);color:var(--green);}
.otb-banner.over{background:var(--red-soft);color:var(--red);}
.card{background:var(--card);border-radius:var(--radius);padding:18px 22px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.06);border:1px solid var(--line);}
.card h2{font-size:15px;margin:0 0 12px;font-weight:700;}
.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px;}
.chart-row canvas{max-height:260px;}
.controls{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;}
.controls input,.controls select{padding:8px 10px;border:1px solid var(--line);border-radius:6px;font-size:13px;}
.controls input{flex:1;min-width:160px;}
table{width:100%;border-collapse:collapse;font-size:12.5px;}
th{text-align:left;padding:8px 10px;border-bottom:2px solid var(--line);color:var(--sub);font-weight:700;font-size:11px;text-transform:uppercase;cursor:pointer;white-space:nowrap;}
th:hover{color:var(--blue);}
td{padding:7px 10px;border-bottom:1px solid #f2f2f5;}
tr:hover td{background:#fafbfc;}
.badge{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:20px;}
.badge.restock{background:var(--blue-soft);color:var(--blue);}
.badge.no-action{background:#f0f0f2;color:var(--sub);}
.flag-po{color:var(--amber);font-weight:600;}
footer{margin-top:20px;font-size:11.5px;color:var(--sub);text-align:center;}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <div class="eyebrow">Frescoone · Inventory Ops</div>
    <h1>Restock Dashboard -- __MONTH__</h1>
    <p id="generated-note"></p>
  </header>

  <div id="otb-banner"></div>
  <section class="kpi-row" id="kpi-row"></section>

  <div class="chart-row">
    <div class="card"><h2>Spend by Supplier</h2><canvas id="supplierChart"></canvas></div>
    <div class="card"><h2>Top Brands by Spend</h2><canvas id="brandChart"></canvas></div>
  </div>
  <div class="chart-row">
    <div class="card"><h2>Restock / No-Action / Missing Breakdown</h2><canvas id="breakdownChart"></canvas></div>
    <div class="card">
      <h2>Summary</h2>
      <table id="summary-table"></table>
    </div>
  </div>

  <div class="card">
    <h2>Full Detail</h2>
    <div class="controls">
      <input type="text" id="search" placeholder="Search SKU...">
      <select id="supplier-filter"></select>
      <select id="action-filter">
        <option value="all">All Actions</option>
        <option value="RESTOCK">Restock Only</option>
        <option value="NO_ACTION">No Action Only</option>
      </select>
    </div>
    <div style="overflow-x:auto;">
      <table id="detail-table"></table>
    </div>
  </div>

  <footer>Frescoone Trading Sdn Bhd -- generated locally by restock_pipeline.py -- not published anywhere</footer>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.1"></script>
<script>
const DATA = __DATA_JSON__;
const COLORS = ['#305496','#3f7a3f','#8a6d00','#a3342a','#5b4b8a','#1f8a8a','#c25b00','#6b7280'];
function fmtMoney(v){ if(Math.abs(v)>=1e3) return 'RM '+(v/1e3).toFixed(1)+'K'; return 'RM '+v.toFixed(2); }

document.getElementById('generated-note').textContent = 'Generated ' + DATA.generated_at;

function renderKpis(){
  const k = DATA.kpis;
  const cards = [
    ['Active SKUs Matched', k.active_skus_matched, ''],
    ['SKUs Needing Restock', k.skus_needing_restock, ''],
    ['Recommended Spend', fmtMoney(k.recommended_spend), ''],
    ['Units to Order', k.units_to_order.toLocaleString(), ''],
    ['On-PO Flags (not netted)', k.on_po_flags, 'warn'],
    ['Low/Out of Stock', k.low_stock_flags, 'flag'],
    ['No Action', k.no_action_count, ''],
    ['Missing from Forecast', k.missing_from_forecast, 'warn'],
  ];
  document.getElementById('kpi-row').innerHTML = cards.map(([label,val,cls]) =>
    `<div class="kpi ${cls}"><div class="label">${label}</div><div class="value">${val}</div></div>`).join('');
}

function renderOtbBanner(){
  const o = DATA.otb_check;
  const el = document.getElementById('otb-banner');
  el.className = 'otb-banner ' + (o.within_budget ? 'ok' : 'over');
  el.innerHTML = `<strong>${o.within_budget ? 'Within budget' : 'Over budget'}:</strong> ` +
    `Recommended spend ${fmtMoney(o.total_spend)} vs. ceiling ${fmtMoney(o.ceiling)} ` +
    `(smaller of ${o.revenue_cap_pct.toFixed(0)}% of last month's revenue = ${fmtMoney(o.revenue_cap)}` +
    (o.cash_cap != null ? `, or cash on hand minus Payroll/Bank OpEx reserve = ${fmtMoney(o.cash_cap)}` : '') +
    `). This is a sanity check on the TOTAL only -- it does not change any SKU's quantity above, which follows the confirmed 60-day/MOQ formula regardless.`;
}

function renderCharts(){
  new Chart(document.getElementById('supplierChart'), {
    type: 'bar',
    data: { labels: DATA.by_supplier.map(s=>s.order_sheet), datasets: [{ label:'Spend', data: DATA.by_supplier.map(s=>s.spend), backgroundColor: COLORS[0]+'CC' }] },
    options: { indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{ticks:{callback:v=>fmtMoney(v)}}} }
  });
  new Chart(document.getElementById('brandChart'), {
    type: 'bar',
    data: { labels: DATA.by_brand.map(b=>b.brand), datasets: [{ label:'Spend', data: DATA.by_brand.map(b=>b.spend), backgroundColor: COLORS[1]+'CC' }] },
    options: { indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{ticks:{callback:v=>fmtMoney(v)}}} }
  });
  const b = DATA.breakdown;
  new Chart(document.getElementById('breakdownChart'), {
    type: 'doughnut',
    data: { labels:['Restock','No Action','Missing','No Cost'], datasets:[{ data:[b.restock,b.no_action,b.missing,b.no_cost], backgroundColor: COLORS.map(c=>c+'CC') }] },
    options: { responsive:true, maintainAspectRatio:false, plugins:{legend:{position:'right'}} }
  });
}

function renderSummaryTable(){
  const k = DATA.kpis;
  document.getElementById('summary-table').innerHTML =
    '<tbody>' +
    `<tr><td>Total active SKUs matched to forecast</td><td>${k.active_skus_matched}</td></tr>` +
    `<tr><td>Needing restock</td><td>${k.skus_needing_restock}</td></tr>` +
    `<tr><td>No action (sufficient stock)</td><td>${k.no_action_count}</td></tr>` +
    `<tr><td>Missing from forecast export</td><td>${k.missing_from_forecast}</td></tr>` +
    `<tr><td>No cost on file</td><td>${k.no_cost_count}</td></tr>` +
    `<tr style="font-weight:700"><td>Total recommended spend</td><td>${fmtMoney(k.recommended_spend)}</td></tr>` +
    '</tbody>';
}

let sortField = 'est_cost', sortDir = 'desc';
function toggleSort(field){
  if (sortField === field) sortDir = sortDir === 'asc' ? 'desc' : 'asc';
  else { sortField = field; sortDir = 'desc'; }
  renderDetailTable();
}

function populateFilters(){
  const suppliers = [...new Set(DATA.detail.map(d=>d.order_sheet))].sort();
  document.getElementById('supplier-filter').innerHTML = '<option value="all">All Suppliers</option>' +
    suppliers.map(s=>`<option value="${s}">${s}</option>`).join('');
}

function getFilteredSorted(){
  const search = document.getElementById('search').value.trim().toLowerCase();
  const supplier = document.getElementById('supplier-filter').value;
  const action = document.getElementById('action-filter').value;
  let rows = DATA.detail;
  if (search) rows = rows.filter(r => r.sku.toLowerCase().includes(search));
  if (supplier !== 'all') rows = rows.filter(r => r.order_sheet === supplier);
  if (action !== 'all') rows = rows.filter(r => r.action === action);
  rows = [...rows].sort((a,b) => {
    let av = a[sortField], bv = b[sortField];
    if (av == null) return 1; if (bv == null) return -1;
    if (typeof av === 'string') { av = av.toLowerCase(); bv = String(bv).toLowerCase(); }
    const cmp = av < bv ? -1 : av > bv ? 1 : 0;
    return sortDir === 'asc' ? cmp : -cmp;
  });
  return rows;
}

function renderDetailTable(){
  const rows = getFilteredSorted();
  const cols = [['SKU','sku'],['Order Sheet','order_sheet'],['Brand','brand'],['Action','action'],
    ['Stock','stock_on_hand'],['On PO','stock_on_purchase_order'],['Sales 30d','total_sales'],
    ['Restock Qty','restock_qty'],['Unit Cost','cost'],['Est. Cost','est_cost']];
  const table = document.getElementById('detail-table');
  table.innerHTML = '<thead><tr>' + cols.map(([label,field]) =>
    `<th onclick="toggleSort('${field}')">${label}${sortField===field ? (sortDir==='asc'?' ▲':' ▼') : ''}</th>`).join('') +
    '<th>Flag</th></tr></thead><tbody>' +
    rows.map(r => `<tr>
      <td>${r.sku}</td><td>${r.order_sheet}</td><td>${r.brand||'-'}</td>
      <td><span class="badge ${r.action==='RESTOCK'?'restock':'no-action'}">${r.action}</span></td>
      <td>${r.stock_on_hand}</td><td>${r.stock_on_purchase_order}</td><td>${r.total_sales}</td>
      <td>${r.restock_qty}</td><td>${fmtMoney(r.cost)}</td><td>${fmtMoney(r.est_cost)}</td>
      <td>${r.on_po_flag ? '<span class="flag-po">ON PO</span>' : ''}</td>
    </tr>`).join('') + '</tbody>';
}

renderKpis();
renderOtbBanner();
renderCharts();
renderSummaryTable();
populateFilters();
renderDetailTable();
document.getElementById('search').addEventListener('input', () => renderDetailTable());
document.getElementById('supplier-filter').addEventListener('change', () => renderDetailTable());
document.getElementById('action-filter').addEventListener('change', () => renderDetailTable());
</script>
</body>
</html>
"""


def write_dashboard_html(dashboard_data, out_path):
    html = DASHBOARD_TEMPLATE.replace("__MONTH__", MONTH_LABEL).replace(
        "__DATA_JSON__", json.dumps(dashboard_data))
    out_path.write_text(html, encoding="utf-8")
    return out_path


def write_pipeline_json(dashboard_data, out_path):
    """Same data as the standalone dashboard, also dropped as JSON so
    finance_analyzer.py's main dashboard.html can pick it up as an
    additional tab (additive/optional, same pattern as restock_data.json --
    dashboard.html builds fine without this file)."""
    out_path.write_text(json.dumps(dashboard_data, indent=2), encoding="utf-8")
    return out_path


def main():
    print("Loading active ISKU catalog...")
    active_df = rp.load_active_isku_catalog()

    print("Loading SiteGiant Inventory Forecasting exports...")
    last30, alltime, monthly = rp.load_all_forecast_files()

    print("Loading finance context (cost master, cash position, revenue)...")
    finance_ctx = rp.load_finance_context()

    print("Computing restock needs (60-day lead time, MOQ, round-to-5)...")
    restock_df, not_in_forecast, no_cost = compute_restock_list(active_df, last30, finance_ctx["sku_master"])

    print("Checking against OTB cash/revenue ceiling (sanity check only)...")
    total_spend = float(restock_df[restock_df["action"] == "RESTOCK"]["est_cost"].sum())
    otb_check = otb_sanity_check(total_spend, finance_ctx)

    print("Writing Restock_Purchase_List workbook...")
    workbook_path = BASE_DIR / f"Restock_Purchase_List_{MONTH_LABEL}.xlsx"
    write_purchase_list_workbook(restock_df, not_in_forecast, no_cost, last30, workbook_path)

    print("Writing per-supplier PO upload files...")
    po_dir = BASE_DIR / "restock-po-files"
    po_files = write_supplier_po_files(restock_df, po_dir)

    print("Building Restock Dashboard...")
    dashboard_data = build_dashboard_json(restock_df, not_in_forecast, no_cost, otb_check)
    dashboard_path = BASE_DIR / f"Restock_Dashboard_{MONTH_LABEL}.html"
    write_dashboard_html(dashboard_data, dashboard_path)
    json_path = BASE_DIR / "restock_pipeline_data.json"
    write_pipeline_json(dashboard_data, json_path)

    to_order = restock_df[restock_df["action"] == "RESTOCK"]
    print(f"\n=== SUMMARY ===")
    print(f"Active SKUs matched: {len(restock_df)}")
    print(f"SKUs needing restock: {len(to_order)} / {int(to_order['restock_qty'].sum())} units")
    print(f"Recommended spend: {fa.fmt_money(total_spend)}")
    print(f"OTB ceiling: {fa.fmt_money(otb_check['ceiling'])} -- {'within budget' if otb_check['within_budget'] else 'OVER BUDGET'}")
    print(f"\nWorkbook: {workbook_path}")
    print(f"Per-supplier PO files: {len(po_files)} written to {po_dir}")
    print(f"Dashboard: {dashboard_path}")
    print(f"Dashboard data written to: {json_path} -- re-run finance_analyzer.py to fold it into dashboard.html")
    if fa.WARNINGS:
        print(f"\n{len(fa.WARNINGS)} data-quality warnings logged during load.")


if __name__ == "__main__":
    main()
